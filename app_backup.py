import os
import pandas as pd
import requests
import json
import time
from datetime import datetime
import sys
from io import BytesIO
import warnings
import concurrent.futures
from threading import Lock
from difflib import SequenceMatcher
import model_GRAND_match.model_grand_match
from concurrent.futures import ThreadPoolExecutor, as_completed
warnings.filterwarnings('ignore')
import re
import jieba
import streamlit as st
import io
import zipfile
from pathlib import Path
import glob
import webbrowser
import shutil
import difflib
import threading
import openpyxl
from queue import Queue
import concurrent.futures
import zipfile
from io import BytesIO
from typing import List, Dict, Tuple, Optional, Union
import streamlit as st
import yt_dlp
import os
import json
import pandas as pd
import xml.etree.ElementTree as ET
import subprocess
import platform
import time
import glob
import tempfile
import io
from datetime import datetime
import jieba
try:
    from wordcloud import WordCloud
    HAS_WORDCLOUD = True
except ImportError:
    HAS_WORDCLOUD = False
import matplotlib.pyplot as plt



# --- 核心工具函数类 (放在函数外以便复用) ---
class Utils:
    @staticmethod
    def load_config(config_file):
        if os.path.exists(config_file):
            with open(config_file, 'r') as f: return json.load(f)
        return {
            "save_path": os.path.join(os.path.expanduser("~"), "Downloads", "Yt-DLP-Data"),
            "proxy": "",
            "naming_tmpl": "%(title)s"
        }

    @staticmethod
    def save_config(config_file, config):
        with open(config_file, 'w') as f: json.dump(config, f)

    @staticmethod
    def open_folder(path):
        if platform.system() == "Windows": os.startfile(path)
        elif platform.system() == "Darwin": subprocess.Popen(["open", path])
        else: subprocess.Popen(["xdg-open", path])

    @staticmethod
    def create_netscape_cookie_file(raw_cookie_str):
        if not raw_cookie_str or "=" not in raw_cookie_str: return None
        try:
            fd, path = tempfile.mkstemp(suffix='.txt', text=True)
            with os.fdopen(fd, 'w') as f:
                f.write("# Netscape HTTP Cookie File\n\n")
                for item in raw_cookie_str.split(';'):
                    if '=' in item:
                        key, value = item.strip().split('=', 1)
                        f.write(f".bilibili.com\tTRUE\t/\tFALSE\t253402300799\t{key}\t{value}\n")
            return path
        except: return None

    @staticmethod
    def get_chinese_font():
        system = platform.system()
        if system == "Windows":
            fonts = ["simhei.ttf", "msyh.ttc", "simsun.ttc"]
            for f in fonts:
                path = os.path.join("C:\\Windows\\Fonts", f)
                if os.path.exists(path): return path
        elif system == "Darwin": return "/System/Library/Fonts/PingFang.ttc"
        return None

    @staticmethod
    def generate_wordcloud_img(text_list):
        if not HAS_WORDCLOUD: return None
        if not text_list: return None
        
        full_text = " ".join([str(t) for t in text_list if str(t)])
        cut_text = " ".join(jieba.cut(full_text))
        
        font_path = Utils.get_chinese_font()
        # 如果没有中文字体，为了不报错，不传font_path参数(虽然会乱码)
        params = {
            'background_color': 'white', 'width': 800, 'height': 400,
            'max_words': 200, 'colormap': 'viridis',
            'stopwords': {'的', '了', '是', '在', '也', '就', '不', '都', '吗', '啊', '吧', '我', '这'}
        }
        if font_path: params['font_path'] = font_path

        wc = WordCloud(**params).generate(cut_text)
        return wc

    @staticmethod
    def process_xml_to_excel(xml_path, excel_path):
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            data = []
            for d in root.findall('d'):
                content = d.text
                p_attr = d.get('p')
                if p_attr and content:
                    attrs = p_attr.split(',')
                    if len(attrs) >= 7:
                        data.append({
                            "时间": f"{int(float(attrs[0])//60):02d}:{int(float(attrs[0])%60):02d}",
                            "秒数": round(float(attrs[0]), 2),
                            "内容": content,
                            "用户Hash": attrs[6],
                            "日期": datetime.fromtimestamp(int(attrs[4])).strftime('%Y-%m-%d')
                        })
            if data:
                pd.DataFrame(data).sort_values(by="秒数").to_excel(excel_path, index=False)
                return True, len(data)
            return False, 0
        except: return False, 0

    @staticmethod
    def process_json_to_excel(json_path, excel_path):
        try:
            with open(json_path, 'r', encoding='utf-8') as f: info = json.load(f)
            comments = info.get('comments', [])
            data = []
            for c in comments:
                data.append({
                    "用户": c.get('author'),
                    "内容": c.get('text'),
                    "点赞": c.get('like_count'),
                    "时间": datetime.fromtimestamp(c.get('timestamp', 0)).strftime('%Y-%m-%d %H:%M') if c.get('timestamp') else "-"
                })
            if data:
                pd.DataFrame(data).to_excel(excel_path, index=False)
                return True, len(data)
            return False, 0
        except: return False, 0


# --- 主函数组件 ---
def ytdlp_downloader_app():
    """
    Streamlit 主函数组件。
    调用此函数即可在任何页面渲染下载器。
    """
    
    # 1. 局部状态初始化 (State Init)
    if 'ytdlp_queue' not in st.session_state: st.session_state.ytdlp_queue = []
    if 'ytdlp_history' not in st.session_state: st.session_state.ytdlp_history = []
    if 'current_meta' not in st.session_state: st.session_state.current_meta = None
    if 'available_formats' not in st.session_state: st.session_state.available_formats = []
    
    CONFIG_FILE = "ytdlp_config.json"
    config = Utils.load_config(CONFIG_FILE)

    st.title("📺 YT-DLP 全能媒体终端")
    if not HAS_WORDCLOUD:
        st.warning("⚠️ 检测到未安装 wordcloud 库，词云功能将不可用，但下载功能正常。")

    # --- 侧边栏 (Sidebar) ---
    with st.sidebar:
        st.header("⚙️ 设置")
        new_path = st.text_input("📂 保存路径", value=config['save_path'])
        if new_path != config['save_path']:
            config['save_path'] = new_path
            Utils.save_config(CONFIG_FILE, config)

        st.divider()
        st.subheader("🍪 Cookie (VIP/评论)")
        raw_cookie = st.text_area("粘贴 Cookie (SESSDATA=...)", height=100, help="F12抓取B站请求头中的Cookie")
        
        temp_cookie_path = None
        if raw_cookie and "SESSDATA" in raw_cookie:
            temp_cookie_path = Utils.create_netscape_cookie_file(raw_cookie)
            if temp_cookie_path: st.success("✅ Cookie 已激活")
        
        st.divider()
        config['proxy'] = st.text_input("代理 (Proxy)", value=config['proxy'])
        if st.button("📂 打开文件夹"): 
            if os.path.exists(config['save_path']): Utils.open_folder(config['save_path'])

    # --- 页面主体 Tabs ---
    tab_dl, tab_review = st.tabs(["⬇️ 下载与解析", "👁️ 资产管理与词云"])

    # === Tab 1: 下载中心 ===
    with tab_dl:
        col1, col2 = st.columns([4,1])
        with col1: url = st.text_input("视频链接", key="url_input")
        with col2: btn_analyze = st.button("🔍 解析", use_container_width=True, type="primary")

        # 基础参数构造器
        def get_opts():
            opts = {'quiet': True, 'proxy': config['proxy'] or None, 'no_warnings': True, 'extractor_args': {'bilibili': {'comment_sort': 'time'}}}
            if temp_cookie_path: opts['cookiefile'] = temp_cookie_path
            return opts

        # 解析逻辑
        if btn_analyze and url:
            with st.spinner("正在解析流..."):
                try:
                    with yt_dlp.YoutubeDL(get_opts()) as ydl:
                        meta = ydl.extract_info(url, download=False)
                        st.session_state.current_meta = meta
                        formats = meta.get('formats', [])
                        heights = sorted(list(set([f.get('height') for f in formats if f.get('height')])), reverse=True)
                        st.session_state.available_formats = [f"{h}p" for h in heights]
                except Exception as e: st.error(f"解析错误: {e}")

        # 任务配置卡片
        if st.session_state.current_meta:
            meta = st.session_state.current_meta
            st.divider()
            c1, c2 = st.columns([1, 2])
            with c1: 
                if meta.get('thumbnail'): st.image(meta['thumbnail'], use_container_width=True)
            with c2:
                st.subheader(meta.get('title'))
                quality = st.selectbox("画质选择", ["✨ 最佳 (MP4)"] + st.session_state.available_formats + ["🎵 纯音频"])
                
                cd1, cd2 = st.columns(2)
                with cd1: get_danmaku = st.checkbox("导出弹幕 Excel", value=True)
                with cd2: get_comments = st.checkbox("导出评论 Excel", value=True)
                
                limit_cmt = 100
                if get_comments: limit_cmt = st.slider("评论抓取量", 10, 5000, 500, step=50)

                if st.button("➕ 加入队列", type="primary"):
                    st.session_state.ytdlp_queue.append({
                        "url": meta['webpage_url'], "title": meta['title'], "quality": quality,
                        "danmaku": get_danmaku, "comments": get_comments, "limit_cmt": limit_cmt
                    })
                    st.success("已加入下载队列")

        # 队列执行
        if st.session_state.ytdlp_queue:
            st.divider()
            if st.button(f"🚀 开始下载 ({len(st.session_state.ytdlp_queue)} 个任务)", type="primary", use_container_width=True):
                prog = st.progress(0)
                for idx, task in enumerate(st.session_state.ytdlp_queue):
                    opts = get_opts()
                    opts.update({'outtmpl': os.path.join(config['save_path'], f"{task['title']}.%(ext)s"), 'ignoreerrors': True, 'merge_output_format': 'mp4', 'writeinfojson': True})
                    
                    # 画质参数
                    if "纯音频" in task['quality']: opts['format'] = 'bestaudio/best'
                    elif "最佳" in task['quality']: opts['format'] = 'bestvideo+bestaudio/best'
                    else: opts['format'] = f"bestvideo[height={task['quality'].replace('p','')}]" + "+bestaudio/best"

                    if task['danmaku']: opts.update({'writesubtitles': True, 'allsubtitles': True})
                    if task['comments']: opts.update({'getcomments': True, 'max_comments': task['limit_cmt']})

                    try:
                        with yt_dlp.YoutubeDL(opts) as ydl:
                            ydl.download([task['url']])
                            base = os.path.join(config['save_path'], task['title'])
                            
                            # Excel 转换
                            if task['danmaku']:
                                xmls = glob.glob(f"{base}*.xml")
                                if xmls: 
                                    Utils.process_xml_to_excel(xmls[0], f"{base}_弹幕.xlsx")
                                    try: os.remove(xmls[0])
                                    except: pass
                            
                            if task['comments']:
                                json_f = f"{base}.info.json"
                                if os.path.exists(json_f):
                                    Utils.process_json_to_excel(json_f, f"{base}_评论.xlsx")
                                    try: os.remove(json_f)
                                    except: pass

                            st.session_state.ytdlp_history.append({"title": task['title'], "video_path": f"{base}.mp4", "base_name": base})
                    except Exception as e: st.error(f"任务失败: {e}")
                    prog.progress((idx+1)/len(st.session_state.ytdlp_queue))
                
                st.session_state.ytdlp_queue = []
                st.success("全部任务完成！")

    # === Tab 2: 资产与词云 ===
    with tab_review:
        if not st.session_state.ytdlp_history: st.info("暂无历史记录")
        
        for item in reversed(st.session_state.ytdlp_history):
            with st.expander(f"🎥 {item['title']}", expanded=True):
                c_vid, c_data = st.columns([1, 1.5])
                with c_vid:
                    if os.path.exists(item['video_path']): st.video(item['video_path'])
                    else: st.warning("文件未找到")
                
                with c_data:
                    dm_path = f"{item['base_name']}_弹幕.xlsx"
                    cm_path = f"{item['base_name']}_评论.xlsx"
                    
                    t1, t2 = st.tabs(["📊 数据", "☁️ 词云"])
                    with t1:
                        if os.path.exists(dm_path): st.dataframe(pd.read_excel(dm_path), height=150)
                        if os.path.exists(cm_path): st.dataframe(pd.read_excel(cm_path), height=150)
                    
                    with t2:
                        if not HAS_WORDCLOUD:
                            st.error("词云库缺失，请安装 wordcloud")
                        else:
                            wc1, wc2 = st.columns(2)
                            with wc1:
                                if os.path.exists(dm_path) and st.button("弹幕词云", key=f"d_{item['title']}"):
                                    wc = Utils.generate_wordcloud_img(pd.read_excel(dm_path)['内容'].tolist())
                                    if wc: 
                                        st.image(wc.to_array(), use_container_width=True)
                                        buf = io.BytesIO()
                                        wc.to_image().save(buf, format='PNG')
                                        st.download_button("下载", buf.getvalue(), "dm_wc.png", "image/png", key=f"dd_{item['title']}")
                            with wc2:
                                if os.path.exists(cm_path) and st.button("评论词云", key=f"c_{item['title']}"):
                                    wc = Utils.generate_wordcloud_img(pd.read_excel(cm_path)['内容'].tolist())
                                    if wc: 
                                        st.image(wc.to_array(), use_container_width=True)
                                        buf = io.BytesIO()
                                        wc.to_image().save(buf, format='PNG')
                                        st.download_button("下载", buf.getvalue(), "cm_wc.png", "image/png", key=f"dc_{item['title']}")
class MultiAPIExcelTranslator:
    def __init__(self, api_key, api_provider, api_url, model, context_size=10, max_retries=10):
        self.api_key = api_key
        self.api_provider = api_provider
        self.api_url = api_url
        self.model = model
        self.headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        self.context_history = {}  # 按语言分别存储上下文
        self.term_dict = {}
        self.role_column = None
        self.context_size = context_size
        self.max_retries = max_retries
        
        # 改为按语言存储术语库
        self.term_base_dict = {}  # {语言: [{source: xxx, target: xxx}]}
        
        self.role_personality_dict = {}
        self.current_text_terms = {}
        self.current_role_personality = None
        self.target_languages = ["英文"]
        self.language_column_names = {"英文": "英文翻译结果"}

        # 新增：角色映射表
        self.role_mapping = {}
        self.enable_fuzzy_match = False
        self.fuzzy_threshold = 0.6

        self.init_chinese_tokenizer()

    def init_chinese_tokenizer(self):
        try:
            self.chinese_tokenizer = jieba
            st.success("✅ 中文分词器初始化成功")
        except Exception as e:
            st.warning(f"⚠️ 中文分词器初始化失败: {e}")
            self.chinese_tokenizer = None

    def tokenize_chinese_text(self, text):
        if not text or pd.isna(text):
            return []

        text = str(text).strip()
        if not text:
            return []

        try:
            if self.chinese_tokenizer:
                words = self.chinese_tokenizer.cut(text)
                return [word for word in words if word.strip() and re.search(r'[\w\u4e00-\u9fa5]', word)]
            else:
                return [char for char in text if char.strip() and re.search(r'[\w\u4e00-\u9fa5]', char)]
        except Exception as e:
            st.warning(f"⚠️ 中文分词失败: {e}")
            return [char for char in text if char.strip()]

    def clean_role_name(self, role_name):
        """清理角色名称：去除额外标记、空格等"""
        if not role_name or pd.isna(role_name):
            return ""

        role_name = str(role_name).strip()
        role_name = re.sub(r'\|.*$', '', role_name)
        role_name = re.sub(r'[\s\u3000]+', '', role_name)
        role_name = role_name.strip()

        return role_name

    def fuzzy_match_role(self, role_name, threshold=None):
        """模糊匹配角色名称"""
        if not role_name or not self.role_personality_dict:
            return None, 0

        if threshold is None:
            threshold = self.fuzzy_threshold

        cleaned_role = self.clean_role_name(role_name)
        if not cleaned_role:
            return None, 0

        if role_name in self.role_mapping:
            return self.role_mapping[role_name], 1.0

        best_match = None
        best_score = 0

        for official_role in self.role_personality_dict.keys():
            cleaned_official = self.clean_role_name(official_role)

            if cleaned_role == cleaned_official:
                return official_role, 1.0

            score = difflib.SequenceMatcher(None, cleaned_role, cleaned_official).ratio()

            if cleaned_role in cleaned_official or cleaned_official in cleaned_role:
                score = max(score, 0.8)

            if score > best_score:
                best_score = score
                best_match = official_role

        if best_score >= threshold:
            return best_match, best_score

        return None, best_score

    def analyze_role_matches(self, df, role_col):
        """分析数据中的所有角色名称"""
        if not role_col or role_col not in df.columns:
            return {}

        unique_roles = df[role_col].dropna().unique()
        fuzzy_matches = {}

        for role in unique_roles:
            role_str = str(role).strip()
            if not role_str:
                continue

            if role_str in self.role_mapping:
                continue

            if role_str in self.role_personality_dict:
                continue

            matched_role, score = self.fuzzy_match_role(role_str)

            if matched_role:
                if role_str not in fuzzy_matches:
                    fuzzy_matches[role_str] = []
                fuzzy_matches[role_str].append((matched_role, score))

                if score < 1.0:
                    for official_role in self.role_personality_dict.keys():
                        if official_role == matched_role:
                            continue
                        alt_score = difflib.SequenceMatcher(
                            None,
                            self.clean_role_name(role_str),
                            self.clean_role_name(official_role)
                        ).ratio()

                        if alt_score >= self.fuzzy_threshold * 0.8:
                            fuzzy_matches[role_str].append((official_role, alt_score))

                fuzzy_matches[role_str].sort(key=lambda x: x[1], reverse=True)

        return fuzzy_matches

    def find_role_personality(self, role_name):
        """查找角色性格描述"""
        if not role_name or not self.role_personality_dict:
            return None

        role_name = str(role_name).strip()
        if not role_name:
            return None

        if role_name in self.role_mapping:
            mapped_role = self.role_mapping[role_name]
            return self.role_personality_dict.get(mapped_role)

        if role_name in self.role_personality_dict:
            return self.role_personality_dict[role_name]

        if self.enable_fuzzy_match:
            matched_role, score = self.fuzzy_match_role(role_name)
            if matched_role:
                return self.role_personality_dict[matched_role]

        return None

    def add_to_context(self, original, translation, role=None, language="英文"):
        """为指定语言添加上下文（确保语言独立）"""
        if language not in self.context_history:
            self.context_history[language] = []
        
        self.context_history[language].append((original, translation, role))
        if len(self.context_history[language]) > self.context_size:
            self.context_history[language].pop(0)

    def build_context_prompt(self, language="英文"):
        """为指定语言构建上下文提示（确保语言独立）"""
        if language not in self.context_history or not self.context_history[language]:
            return ""

        context_str = f"\n\n### 重要上下文参考（{language}翻译）：\n"
        for i, (orig, trans, role) in enumerate(self.context_history[language], 1):
            role_info = f" [{role}]" if role else ""
            context_str += f"前文{i}{role_info}:\n原文: {orig}\n{language}译文: {trans}\n\n"
        return context_str

    def find_matched_terms(self, text, language):
        """为指定语言查找匹配的术语"""
        if language not in self.term_base_dict or not self.term_base_dict[language]:
            return {}

        words = self.tokenize_chinese_text(text)
        matched_terms = {}

        # 词级别匹配
        for word in words:
            for term_entry in self.term_base_dict[language]:
                if term_entry['source'] == word:
                    if word not in matched_terms:
                        matched_terms[word] = []
                    if term_entry['target'] not in matched_terms[word]:
                        matched_terms[word].append(term_entry['target'])

        # 短语级别匹配
        for term_entry in self.term_base_dict[language]:
            term = term_entry['source']
            if term in text:
                if term not in matched_terms:
                    matched_terms[term] = []
                if term_entry['target'] not in matched_terms[term]:
                    matched_terms[term].append(term_entry['target'])

        return matched_terms

    def build_term_base_prompt(self, text, language):
        """为指定语言构建术语库提示"""
        matched_terms = self.find_matched_terms(text, language)

        if not matched_terms:
            return ""

        term_base_str = f"\n\n### 术语库匹配：\n"
        
        for orig, trans_list in matched_terms.items():
            if len(trans_list) == 1:
                term_base_str += f"- 「{orig}」 → {language}译名：「{trans_list[0]}」\n"
            else:
                term_base_str += f"- 「{orig}」 → {language}译名候选：{' / '.join([f'「{t}」' for t in trans_list])} （根据上下文选择最合适的）\n"

        return term_base_str

    def build_role_personality_prompt(self, role_name):
        if not role_name:
            return ""

        personality = self.find_role_personality(role_name)
        self.current_role_personality = personality

        if not personality:
            return ""

        mapped_role = self.role_mapping.get(str(role_name).strip())
        if mapped_role and mapped_role != str(role_name).strip():
            role_personality_str = f"\n\n### 角色性格描述：\n角色「{role_name}」(映射为「{mapped_role}」)的性格特点：{personality}\n"
        else:
            role_personality_str = f"\n\n### 角色性格描述：\n角色「{role_name}」的性格特点：{personality}\n"

        return role_personality_str

    def set_target_languages(self, languages, column_names):
        """设置目标语言列表和对应的列名"""
        self.target_languages = languages
        self.language_column_names = column_names
        # 为每种语言初始化独立的上下文历史
        for lang in languages:
            if lang not in self.context_history:
                self.context_history[lang] = []
    def set_target_language(self, language):
        self.target_language = language
    def get_language_specific_requirements(self, language):
        language_requirements = {
            "英文": """
英文翻译要求：
- 保持自然流畅，符合英语母语者的表达习惯
- 游戏UI文本要简洁明了，避免冗长
- 角色对话要符合人物性格，使用恰当的语气
- 专有名词和术语要保持一致性
- 文化特定表达要进行适当的本地化处理
""",
            "日文": """
日文翻译要求：
- 注意敬体和常体的使用，根据角色关系和场景选择合适的语体
- 游戏UI文本要简洁明了，符合日语表达习惯
- 角色对话要符合人物性格，使用恰当的语气和语尾
- 专有名词和术语要保持一致性
- 文化特定表达要进行适当的本地化处理
""",
            "韩文": """
韩文翻译要求：
- 注意尊敬语和非尊敬语的使用，根据角色关系和场景选择合适的语体
- 游戏UI文本要简洁明了，符合韩语表达习惯
- 角色对话要符合人物性格，使用恰当的语气
- 专有名词和术语要保持一致性
- 文化特定表达要进行适当的本地化处理
"""
        }

        if language in language_requirements:
            return language_requirements[language]
        else:
            return f"""
{language}翻译要求：
- 注意正式和非正式语体的使用，根据角色关系和场景选择合适的语体
- 游戏UI文本要简洁明了，符合{language}表达习惯
- 角色对话要符合人物性格，使用恰当的语气
- 专有名词和术语要保持一致性
- 文化特定表达要进行适当的本地化处理
"""

    def is_translation_error(self, response_text, original_text):
        if not response_text or response_text.strip() == "":
            return True
        if len(response_text) < len(original_text) * 0.1:
            return True
        return False

    def translate_text_with_retry(self, text, target_language, custom_requirements="", role=None):
        if not text or pd.isna(text) or str(text).strip() == "":
            return text

        last_exception = None

        for attempt in range(self.max_retries):
            try:
                translated_text = self._translate_single_attempt(text, target_language, custom_requirements, role)

                if not self.is_translation_error(translated_text, text):
                    return translated_text
                else:
                    st.warning(f"⚠️ [{target_language}] 第 {attempt + 1} 次翻译结果异常，准备重试...")

            except requests.exceptions.RequestException as e:
                last_exception = e
                st.warning(f"⚠️ [{target_language}] 网络错误 (第 {attempt + 1} 次尝试): {e}")

            except requests.exceptions.Timeout as e:
                last_exception = e
                st.warning(f"⚠️ [{target_language}] 请求超时 (第 {attempt + 1} 次尝试): {e}")

            except requests.exceptions.ConnectionError as e:
                last_exception = e
                st.warning(f"⚠️ [{target_language}] 连接错误 (第 {attempt + 1} 次尝试): {e}")

            except Exception as e:
                last_exception = e
                st.warning(f"⚠️ [{target_language}] API错误 (第 {attempt + 1} 次尝试): {e}")

            if attempt < self.max_retries - 1:
                wait_time = min(2 ** attempt, 60)
                time.sleep(wait_time)

        st.error(f"❌ [{target_language}] 翻译失败，已达到最大重试次数 {self.max_retries}")
        if last_exception:
            st.error(f"最后错误: {last_exception}")

        return text

    def _translate_single_attempt(self, text, target_language, custom_requirements="", role=None):
        # 为当前语言构建独立的上下文和术语提示
        context_prompt = self.build_context_prompt(target_language)
        term_base_prompt = self.build_term_base_prompt(text, target_language)
        role_personality_prompt = self.build_role_personality_prompt(role) if role else ""
        language_requirements = self.get_language_specific_requirements(target_language)

        role_prompt = ""
        if role and not pd.isna(role) and str(role).strip() != "":
            role_prompt = f"\n当前文本的说话人: {role}\n"

        prompt = f"""
请将以下文本翻译成{target_language}。

## 角色信息：
{role_prompt}{role_personality_prompt}

## {target_language}翻译规范（优先级最低）：
{language_requirements}

## 用户自定义要求（优先级第一高）：
{custom_requirements}

{context_prompt}

{term_base_prompt}

## 待翻译文本：
{text}

## 重要说明（优先级第二高）：
1. 请只返回{target_language}翻译结果，不要添加任何解释或备注
2. 术语库中的特定词汇翻译，如果是人名或者固定特殊名称需要严格采用相同的翻译,但注意如果是一些普通的词汇则看句子翻译不必一定按照术语库来
3. 如果一个术语有多个候选译名，请根据上下文选择最合适的
4. 请根据角色性格描述调整翻译风格和语气
5. 参考上下文中的{target_language}译文风格，保持翻译一致性，但是有的时候上下文角色可能不是一个人或者只是UI翻译，你还是需要参考原文判断是否参考上下文

{target_language}翻译结果：
"""

        data = {
            "model": self.model,
            "messages": [
                {
                    "role": "system",
                    "content": f"你是一名专业的{target_language}翻译专家，擅长游戏本地化、UI界面翻译和角色文案翻译。你正在进行中文到{target_language}的翻译工作。请确保术语一致性和风格统一，并根据角色特点调整翻译风格。"
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "temperature": 0.2,
            "max_tokens": 4000
        }
        print(str(prompt))
        response = requests.post(self.api_url, headers=self.headers, json=data, timeout=60)
        response.raise_for_status()
        result = response.json()
        translated_text = result["choices"][0]["message"]["content"].strip()
        
        translated_text = self.clean_translation(translated_text)
        # 将翻译结果添加到该语言的独立上下文中
        self.add_to_context(text, translated_text, role, target_language)

        return translated_text

    def translate_text(self, text, target_language, custom_requirements="", role=None):
        return self.translate_text_with_retry(text, target_language, custom_requirements, role)

    def clean_translation(self, text):
        if text.startswith('"') and text.endswith('"'):
            text = text[1:-1]
        elif text.startswith("'") and text.endswith("'"):
            text = text[1:-1]

        prefixes = [
            "monior：", "monior:", "角色：", "角色:",
            "翻译：", "翻译:", "译文：", "译文:",
            "结果：", "结果:", "翻译结果：", "翻译结果:",
            "英文翻译结果：", "日文翻译结果：", "韩文翻译结果：",
            "英文：", "日文：", "韩文：", "法文：", "德文：",
        ]
        for prefix in prefixes:
            if text.startswith(prefix):
                text = text[len(prefix):].strip()

        return text

    def reset_context(self):
        self.context_history = {}
        self.term_dict = {}
        self.term_base_dict = {}
        self.role_column = None
        self.current_text_terms = {}
        self.current_role_personality = None
        self.role_mapping = {}
    def load_term_base(self, df, source_col, target_col):
        """加载术语库 - 支持重复术语"""
        try:
            # 构建术语库列表，支持重复术语
            self.term_base_list = []
            missing_count = 0
            
            for _, row in df.iterrows():
                source = row[source_col]
                target = row[target_col]
                
                if pd.isna(source) or pd.isna(target):
                    missing_count += 1
                    continue
                    
                source = str(source).strip()
                target = str(target).strip()
                
                if source and target:
                    # 不再检查重复，直接添加到列表
                    self.term_base_list.append({
                        'source': source,
                        'target': target
                    })
            
            st.success(f"✅ 成功加载术语: {len(self.term_base_list)} 条")
            if missing_count > 0:
                st.warning(f"⚠️ 跳过 {missing_count} 条不完整的记录")
            
            return True
        except Exception as e:
            st.error(f"❌❌ 加载术语库失败: {e}")
            return False
    def load_term_base_multilang(self, df, source_col, target_cols_dict):
        """
        加载多语言术语库
        df: 术语库DataFrame
        source_col: 原文列名
        target_cols_dict: {语言: 列名} 例如 {"英文": "English", "日文": "Japanese"}
        """
        try:
            self.term_base_dict = {}
            
            for language, target_col in target_cols_dict.items():
                if target_col not in df.columns:
                    st.warning(f"⚠️ 术语库中未找到 {language} 对应的列: {target_col}")
                    continue
                
                self.term_base_dict[language] = []
                missing_count = 0

                for _, row in df.iterrows():
                    source = row[source_col]
                    target = row[target_col]

                    if pd.isna(source) or pd.isna(target):
                        missing_count += 1
                        continue

                    source = str(source).strip()
                    target = str(target).strip()

                    if source and target:
                        # 将原文添加到分词词典
                        try:
                            self.chinese_tokenizer.add_word(source)
                        except:
                            pass

                        self.term_base_dict[language].append({
                            'source': source,
                            'target': target
                        })

                st.success(f"✅ {language} 术语加载成功: {len(self.term_base_dict[language])} 条")
                if missing_count > 0:
                    st.info(f"   跳过 {missing_count} 条不完整的 {language} 术语")

            # 显示术语库统计
            total_terms = sum(len(terms) for terms in self.term_base_dict.values())
            st.success(f"📊 总计加载术语: {total_terms} 条，覆盖 {len(self.term_base_dict)} 种语言")
            
            # 显示术语示例
            with st.expander("📋 查看术语库示例"):
                for language, terms in self.term_base_dict.items():
                    if terms:
                        st.write(f"**{language}术语示例：**")
                        for i, term in enumerate(terms[:5]):
                            st.write(f"  {i+1}. {term['source']} → {term['target']}")
                        if len(terms) > 5:
                            st.write(f"  ... 还有 {len(terms)-5} 条")

            return True
            
        except Exception as e:
            st.error(f"❌ 加载多语言术语库失败: {e}")
            import traceback
            st.error(traceback.format_exc())
            return False

    def load_role_personality(self, df, role_col, personality_col):
        try:
            self.role_personality_dict = {}
            missing_count = 0

            for _, row in df.iterrows():
                role = row[role_col]
                personality = row[personality_col]

                if pd.isna(role) or pd.isna(personality):
                    missing_count += 1
                    continue

                role = str(role).strip()
                personality = str(personality).strip()

                if role and personality:
                    self.role_personality_dict[role] = personality

            st.success(f"✅ 成功加载角色性格: {len(self.role_personality_dict)} 条")
            if missing_count > 0:
                st.warning(f"⚠️ 跳过 {missing_count} 条不完整的记录")

            self.analyze_role_personality()

            return True
        except Exception as e:
            st.error(f"❌ 加载角色性格库失败: {e}")
            return False

    def analyze_role_personality(self):
        if not self.role_personality_dict:
            return

        st.write(f"📊 角色性格库统计: {len(self.role_personality_dict)} 个角色")

        st.write("📋 部分角色性格预览:")
        count = 0
        for role, personality in list(self.role_personality_dict.items())[:5]:
            st.write(f"  - {role}: {personality[:50]}..." if len(personality) > 50 else f"  - {role}: {personality}")
            count += 1
            if count >= 5:
                break


def get_api_providers():
    providers = {
        "DeepSeek": {
            "url": "https://api.deepseek.com/v1/chat/completions",
            "models": ["deepseek-chat", "deepseek-coder"]
        },
        "OpenAI": {
            "url": "https://api.openai.com/v1/chat/completions",
            "models": ["gpt-3.5-turbo", "gpt-4", "gpt-4-turbo"]
        },
        "自定义API": {
            "url": "https://tb.api.mkeai.com/v1/chat/completions",
            "models": ["custom-model"]
        }
    }
    return providers


def render_role_matching_interface(translator, df, role_col):
    """渲染角色匹配确认界面"""
    st.header("🎭 角色模糊匹配确认")

    fuzzy_matches = translator.analyze_role_matches(df, role_col)

    if not fuzzy_matches:
        st.success("✅ 所有角色都已精确匹配，无需确认")
        return True

    st.warning(f"⚠️ 发现 {len(fuzzy_matches)} 个需要确认的角色匹配")
    st.info("💡 提示：系统会自动为相同的角色名批量应用您的选择")

    if 'role_confirmations' not in st.session_state:
        st.session_state.role_confirmations = {}

    with st.form("role_matching_form"):
        for idx, (original_role, candidates) in enumerate(fuzzy_matches.items()):
            st.markdown(f"---")
            st.markdown(f"### 角色 {idx + 1}: `{original_role}`")

            cleaned = translator.clean_role_name(original_role)
            st.caption(f"清理后: `{cleaned}`")

            role_count = len(df[df[role_col] == original_role])
            st.caption(f"📊 在文档中出现 **{role_count}** 次")

            options = ["❌ 不匹配任何角色"] + [
                f"✅ {candidate} (相似度: {score:.2%})"
                for candidate, score in candidates
            ]

            default_idx = 1 if candidates else 0

            selected = st.radio(
                f"请选择匹配的官方角色:",
                options=options,
                index=default_idx,
                key=f"role_match_{idx}"
            )

            if selected.startswith("✅"):
                matched_role = selected.split("(")[0].replace("✅", "").strip()
                st.session_state.role_confirmations[original_role] = matched_role
            else:
                st.session_state.role_confirmations[original_role] = None

            if st.session_state.role_confirmations.get(original_role):
                matched = st.session_state.role_confirmations[original_role]
                personality = translator.role_personality_dict.get(matched)
                if personality:
                    with st.expander("👤 查看角色性格描述"):
                        st.write(personality)

        submitted = st.form_submit_button("✅ 确认所有匹配", use_container_width=True)

        if submitted:
            for original_role, matched_role in st.session_state.role_confirmations.items():
                if matched_role:
                    translator.role_mapping[original_role] = matched_role

            st.success(f"✅ 已确认 {len([v for v in st.session_state.role_confirmations.values() if v])} 个角色映射")

            with st.expander("📋 查看映射摘要"):
                for orig, matched in st.session_state.role_confirmations.items():
                    if matched:
                        st.write(f"• `{orig}` → `{matched}`")
                    else:
                        st.write(f"• `{orig}` → ❌ 未匹配")

            return True

    return False


def batch_translation_page():
    st.title("批量翻译工具 - 多语言优化版")
    text_col = None
    role_col = None
    source_col = None
    role_name_col = None
    personality_col = None
    
    st.markdown("### 支持同时翻译多种语言，独立的术语库和上下文管理")

    # 初始化会话状态
    if 'translator' not in st.session_state:
        st.session_state.translator = None
    if 'current_file' not in st.session_state:
        st.session_state.current_file = None
    if 'term_base_df' not in st.session_state:
        st.session_state.term_base_df = None
    if 'role_personality_df' not in st.session_state:
        st.session_state.role_personality_df = None
    if 'role_matching_confirmed' not in st.session_state:
        st.session_state.role_matching_confirmed = False
    if 'translation_progress' not in st.session_state:
        st.session_state.translation_progress = None
    if 'language_configs' not in st.session_state:
        st.session_state.language_configs = {}
    if 'term_language_mapping' not in st.session_state:
        st.session_state.term_language_mapping = {}

    # 侧边栏配置
    with st.sidebar:
        st.header("⚙️ API配置")

        api_providers = get_api_providers()
        api_provider = st.selectbox(
            "🌍 API提供商",
            options=list(api_providers.keys()),
            index=0,
            key="batch_api_provider"
        )

        api_key = st.text_input(
            "🔑 API密钥",
            type="password",
            key="batch_api_key"
        )

        if api_provider == "自定义API":
            api_url = st.text_input(
                "🔗 API URL",
                value="https://tb.api.mkeai.com/v1/chat/completions",
                key="batch_api_url"
            )
        else:
            api_url = api_providers[api_provider]["url"]

        model = st.text_input(
            "🤖 模型名称",
            value="deepseek-chat",
            key="batch_model"
        )

        st.markdown("---")
        st.header("🌐 多语言配置")
        
        # 可选语言列表
        available_languages = ["英文", "日文", "韩文", "法文", "德文", "西班牙文", "俄文", "阿拉伯文", "葡萄牙文", "意大利文"]
        
        # 选择要翻译的语言
        selected_languages = st.multiselect(
            "🎯 选择目标语言（可多选）",
            options=available_languages,
            default=["英文"],
            help="可以同时选择多种语言进行翻译",
            key="selected_languages"
        )
        
        if not selected_languages:
            st.warning("⚠️ 请至少选择一种目标语言")
        
        # 为每种语言配置列名
        st.subheader("📝 自定义结果列名")
        language_column_names = {}
        
        for lang in selected_languages:
            default_name = f"{lang}翻译结果"
            col_name = st.text_input(
                f"{lang} 结果列名",
                value=default_name,
                key=f"col_name_{lang}",
                help=f"设置{lang}翻译结果在Excel中的列名"
            )
            language_column_names[lang] = col_name
        
        st.session_state.language_configs = {
            'languages': selected_languages,
            'column_names': language_column_names
        }

        st.markdown("---")
        st.header("💾 自动保存设置")
        
        auto_save_interval = st.number_input(
            "自动保存间隔（每N行）",
            min_value=10,
            max_value=500,
            value=50,
            step=10,
            help="每翻译N行后自动保存一次"
        )
        
        save_directory = st.text_input(
            "保存目录",
            value="./translation_saves",
            help="自动保存文件的目录路径"
        )

        st.markdown("---")
        st.header("🎭 角色匹配设置")

        enable_fuzzy = st.checkbox(
            "启用模糊角色匹配",
            value=True,
            help="自动识别文档中的角色名变体（如空格、错别字等）",
            key="enable_fuzzy_match"
        )

        if enable_fuzzy:
            fuzzy_threshold = st.slider(
                "匹配相似度阈值",
                min_value=0.5,
                max_value=1.0,
                value=0.6,
                step=0.05,
                help="相似度越高越严格，0.6为推荐值",
                key="fuzzy_threshold"
            )
        else:
            fuzzy_threshold = 1.0

        st.markdown("---")
        context_size = st.slider(
            "📚 上下文记录数量",
            min_value=1,
            max_value=20,
            value=10,
            help="每种语言独立维护的上下文数量",
            key="batch_context_size"
        )

        max_retries = st.number_input(
            "🔄 最大重试次数",
            min_value=1,
            max_value=10000,
            value=10,
            key="batch_max_retries"
        )

    # 主界面
    col1, col2 = st.columns([1, 1])

    with col1:
        st.header("📁 文件上传")

        # 检查是否有保存的进度文件
        saved_files = []
        if os.path.exists(save_directory):
            saved_files = [f for f in os.listdir(save_directory) if f.endswith('_progress.xlsx')]
        
        resume_mode = st.checkbox(
            "🔄 从上次进度继续翻译",
            value=False,
            help="从之前保存的进度文件继续翻译"
        )
        
        if resume_mode and saved_files:
            st.info("📋 找到以下进度文件：")
            selected_progress_file = st.selectbox(
                "选择要继续的进度文件",
                options=saved_files,
                format_func=lambda x: f"{x} ({datetime.fromtimestamp(os.path.getmtime(os.path.join(save_directory, x))).strftime('%Y-%m-%d %H:%M:%S')})"
            )
            
            if st.button("📂 加载进度文件"):
                try:
                    progress_path = os.path.join(save_directory, selected_progress_file)
                    df = pd.read_excel(progress_path)
                    df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                    st.session_state.current_file = df
                    
                    # 统计每种语言的翻译进度
                    progress_info = []
                    for lang, col_name in st.session_state.language_configs['column_names'].items():
                        if col_name in df.columns:
                            translated_count = df[col_name].notna().sum()
                            total_count = len(df)
                            progress_info.append(f"{lang}: {translated_count}/{total_count}")
                    
                    st.success(f"✅ 成功加载进度文件！")
                    if progress_info:
                        st.info(f"📊 翻译进度: {', '.join(progress_info)}")
                    
                    with st.expander("📊 文件预览"):
                        st.dataframe(df.head(10))
                except Exception as e:
                    st.error(f"❌ 加载进度文件失败: {e}")
        else:
            uploaded_file = st.file_uploader(
                "📄 上传翻译文件 (Excel)",
                type=['xlsx', 'xls', 'csv'],
                key="batch_file_uploader"
            )

            if uploaded_file is not None:
                try:
                    if uploaded_file.name.endswith('.csv'):
                        df = pd.read_csv(uploaded_file)
                    else:
                        df = pd.read_excel(uploaded_file)
                    df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                    st.session_state.current_file = df
                    st.success(f"✅ 成功读取文件，共 {len(df)} 行数据")

                    with st.expander("📊 文件预览"):
                        st.dataframe(df.head(10))

                except Exception as e:
                    st.error(f"❌ 文件读取失败: {e}")
        
        if st.session_state.current_file is not None:
            df = st.session_state.current_file
            cols = df.columns.tolist()
            text_col = st.selectbox(
                "📝 选择文本列",
                options=cols,
                index=0,
                key="batch_text_col"
            )

            role_col = st.selectbox(
                "👥 选择角色列 (可选)",
                options=["无"] + cols,
                index=0,
                key="batch_role_col"
            )
            if role_col == "无":
                role_col = None

    with col2:
        st.header("📚 术语库和性格库")

        uploaded_term_base = st.file_uploader(
            "📚 上传多语言术语库 (Excel)",
            type=['xlsx', 'xls'],
            key="batch_term_base_uploader",
            help="术语库应包含原文列和多个目标语言列"
        )

        if uploaded_term_base is not None:
            try:
                term_df = pd.read_excel(uploaded_term_base)
                term_df.columns = term_df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.term_base_df = term_df
                st.success(f"✅ 成功读取术语库，共 {len(term_df)} 条术语")

                with st.expander("📋 配置术语库列映射", expanded=True):
                    term_cols = term_df.columns.tolist()
                    
                    # 选择原文列
                    source_col = st.selectbox(
                        "📤 选择原文列（中文）",
                        options=term_cols,
                        index=0,
                        key="batch_source_col"
                    )
                    
                    st.markdown("---")
                    st.subheader("🌐 为每种语言选择对应的术语列")
                    st.info("💡 提示：为每种目标语言选择术语库中对应的翻译列")
                    
                    # 为每种选定的语言配置术语列
                    term_language_mapping = {}
                    for lang in selected_languages:
                        st.markdown(f"**{lang} 术语列：**")
                        target_col = st.selectbox(
                            f"选择 {lang} 对应的术语列",
                            options=["不使用术语库"] + term_cols,
                            index=0,
                            key=f"term_col_{lang}",
                            help=f"选择术语库中 {lang} 翻译对应的列"
                        )
                        
                        if target_col != "不使用术语库":
                            term_language_mapping[lang] = target_col
                            
                            # 显示该语言的术语示例
                            sample_terms = term_df[[source_col, target_col]].head(3).dropna()
                            if not sample_terms.empty:
                                st.caption(f"示例：")
                                for _, row in sample_terms.iterrows():
                                    st.caption(f"  • {row[source_col]} → {row[target_col]}")
                    
                    st.session_state.term_language_mapping = term_language_mapping
                    
                    if term_language_mapping:
                        st.success(f"✅ 已配置 {len(term_language_mapping)} 种语言的术语映射")
                    else:
                        st.warning("⚠️ 未配置任何语言的术语映射")

            except Exception as e:
                st.error(f"❌ 术语库读取失败: {e}")

        uploaded_role_personality = st.file_uploader(
            "📋 上传角色性格库文件 (Excel)",
            type=['xlsx', 'xls'],
            key="batch_role_personality_uploader"
        )

        if uploaded_role_personality is not None:
            try:
                role_personality_df = pd.read_excel(uploaded_role_personality)
                role_personality_df.columns = role_personality_df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.role_personality_df = role_personality_df
                st.success(f"✅ 成功读取角色性格库，共 {len(role_personality_df)} 条记录")

                role_personality_cols = role_personality_df.columns.tolist()
                role_name_col = st.selectbox(
                    "👥 选择角色名称列",
                    options=role_personality_cols,
                    index=0,
                    key="batch_role_name_col"
                )

                personality_col = st.selectbox(
                    "💬 选择性格描述列",
                    options=role_personality_cols,
                    index=min(1, len(role_personality_cols) - 1) if len(role_personality_cols) > 1 else 0,
                    key="batch_personality_col"
                )

            except Exception as e:
                st.error(f"❌ 角色性格库读取失败: {e}")

    # 翻译要求设置
    st.header("🎯 翻译要求设置")

    custom_requirements = st.text_area(
        "💬 自定义翻译要求（适用于所有语言）",
        value="角色对话自然流畅；专业术语统一；保持原文风格；本地化适配；保持上下文一致性；根据角色调整语气;请注意使用语体，且所有角色除了微型机和班长，其他都为女生用语，不要用男性用语，现在角色们都十分熟悉彼此了，不需要使用太正式尊重的语体了例如日语的话不需要ですます型了。",
        height=100,
        key="batch_custom_requirements"
    )

    # 初始化翻译器和加载资源
    if st.button("🔧 初始化翻译器", type="secondary", use_container_width=True):
        if not api_key:
            st.error("❌ 请先输入API密钥")
        elif not selected_languages:
            st.error("❌ 请至少选择一种目标语言")
        else:
            try:
                translator = MultiAPIExcelTranslator(
                    api_key, api_provider, api_url, model,
                    context_size, max_retries
                )
                translator.enable_fuzzy_match = enable_fuzzy
                translator.fuzzy_threshold = fuzzy_threshold
                translator.set_target_languages(selected_languages, language_column_names)

                st.session_state.translator = translator
                st.session_state.role_matching_confirmed = False

                st.info(f"🌍 使用 {api_provider} API")
                st.info(f"🤖 使用模型: {model}")
                st.info(f"🎯 目标语言: {', '.join(selected_languages)}")
                
                # 显示列名配置
                with st.expander("📋 查看列名配置"):
                    for lang, col_name in language_column_names.items():
                        st.write(f"• {lang} → `{col_name}`")

                # 加载多语言术语库
                if st.session_state.term_base_df is not None and st.session_state.term_language_mapping:
                    if translator.load_term_base_multilang(
                        st.session_state.term_base_df, 
                        source_col, 
                        st.session_state.term_language_mapping
                    ):
                        st.success("✅ 多语言术语库加载成功")
                elif st.session_state.term_base_df is not None:
                    st.warning("⚠️ 术语库已上传但未配置语言映射")

                # 加载角色性格库
                if st.session_state.role_personality_df is not None:
                    if translator.load_role_personality(
                            st.session_state.role_personality_df,
                            role_name_col,
                            personality_col
                    ):
                        st.success("✅ 角色性格库加载成功")

                st.success("✅ 翻译器初始化完成！")

            except Exception as e:
                st.error(f"❌ 初始化失败: {e}")
                import traceback
                st.error(traceback.format_exc())

    # 角色匹配确认界面
    if (st.session_state.translator is not None and
            st.session_state.current_file is not None and
            role_col is not None and
            enable_fuzzy and
            not st.session_state.role_matching_confirmed):

        st.markdown("---")
        confirmed = render_role_matching_interface(
            st.session_state.translator,
            st.session_state.current_file,
            role_col
        )
        if confirmed:
            st.session_state.role_matching_confirmed = True
            st.rerun()

    # 开始翻译按钮
    st.markdown("---")
    translation_ready = (
            st.session_state.translator is not None and
            st.session_state.current_file is not None and
            selected_languages and
            (not enable_fuzzy or not role_col or st.session_state.role_matching_confirmed)
    )

    if not translation_ready and enable_fuzzy and role_col:
        st.info("💡 请先完成角色匹配确认后再开始翻译")

    if st.button(
            "🎯 开始多语言翻译",
            type="primary",
            use_container_width=True,
            disabled=not translation_ready,
            key="batch_start_translation"
    ):
        try:
            translator = st.session_state.translator
            df = st.session_state.current_file.copy()
            languages = st.session_state.language_configs['languages']
            column_names = st.session_state.language_configs['column_names']

            # 确保保存目录存在
            os.makedirs(save_directory, exist_ok=True)

            # 显示配置信息
            with st.expander("📋 查看翻译配置", expanded=True):
                st.write("**语言配置：**")
                for lang, col_name in column_names.items():
                    term_status = "✅ 已配置术语库" if lang in translator.term_base_dict and translator.term_base_dict[lang] else "⚠️ 未配置术语库"
                    st.write(f"• {lang} → `{col_name}` ({term_status})")
                
                if translator.role_mapping:
                    st.write("**角色映射：**")
                    for orig, mapped in translator.role_mapping.items():
                        st.write(f"• `{orig}` → `{mapped}`")

            # 为每种语言添加结果列
            for lang in languages:
                col_name = column_names[lang]
                if col_name not in df.columns:
                    df[col_name] = ''
            
            # 计算起始位置
            start_index = 0
            if resume_mode:
                min_translated_index = len(df)
                for lang in languages:
                    col_name = column_names[lang]
                    if col_name in df.columns:
                        last_index = -1
                        for idx in range(len(df)):
                            if not pd.isna(df.at[idx, col_name]) and str(df.at[idx, col_name]).strip() != '':
                                if not str(df.at[idx, col_name]).startswith('[翻译失败'):
                                    last_index = idx
                        min_translated_index = min(min_translated_index, last_index + 1)
                
                start_index = min_translated_index
                if start_index > 0:
                    st.info(f"🔄 继续翻译：跳过前 {start_index} 行（已翻译）")

            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # 为每种语言创建统计
            stats = {lang: {'success': 0, 'error': 0} for lang in languages}
            
            total_rows = len(df)
            
            # 生成保存文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            progress_filename = f"translation_progress_multilang_{timestamp}.xlsx"
            progress_path = os.path.join(save_directory, progress_filename)

            try:
                for index in range(start_index, total_rows):
                    row = df.iloc[index]
                    progress = (index + 1) / total_rows
                    progress_bar.progress(progress)
                    
                    # 构建状态信息
                    stats_str = " | ".join([f"{lang}: ✓{stats[lang]['success']} ✗{stats[lang]['error']}" for lang in languages])
                    status_text.text(f"📝 正在翻译第 {index + 1}/{total_rows} 行... | {stats_str}")

                    text = str(row[text_col])
                    role = row[role_col] if role_col and role_col in row else None

                    # 跳过原文为空的行
                    if pd.isna(text) == "" or str(text).strip() == "" or text == "nan":
                        print("为空")
                        continue
                    
                    # 对每种语言进行翻译（独立翻译，互不影响）
                    for lang in languages:
                        col_name = column_names[lang]
                        
                        # 如果该行该语言已经翻译过，跳过
                        existing_translation = df.at[index, col_name]
                        if not pd.isna(existing_translation) and str(existing_translation).strip() != '' and not str(existing_translation).startswith('[翻译失败'):
                            stats[lang]['success'] += 1
                            continue

                        try:
                            # 每种语言独立翻译，使用各自的上下文和术语库
                            translated_text = translator.translate_text(
                                text, lang, custom_requirements, role
                            )
                            
                            df.at[index, col_name] = translated_text
                            stats[lang]['success'] += 1
                            
                        except Exception as e:
                            error_msg = str(e)
                            st.warning(f"⚠️ [{lang}] 第 {index + 1} 行翻译失败: {error_msg}")
                            df.at[index, col_name] = f"[翻译失败: {error_msg}]"
                            stats[lang]['error'] += 1
                        
                        # 短暂延迟，避免API速率限制
                        time.sleep(0.15)

                    # 自动保存
                    if (index + 1) % auto_save_interval == 0:
                        try:
                            with pd.ExcelWriter(progress_path, engine='openpyxl') as writer:
                                df.to_excel(writer, index=False, sheet_name='翻译进度')
                            st.info(f"💾 已自动保存进度: {index + 1}/{total_rows} 行")
                        except Exception as save_error:
                            st.warning(f"⚠️ 自动保存失败: {save_error}")

                # 最终保存
                final_filename = f"translation_final_multilang_{timestamp}.xlsx"
                final_path = os.path.join(save_directory, final_filename)
                
                with pd.ExcelWriter(final_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='翻译结果')
                
                progress_bar.progress(1.0)
                
                # 显示最终统计
                st.success("✅ 多语言翻译完成！")
                
                stats_cols = st.columns(len(languages))
                for idx, lang in enumerate(languages):
                    with stats_cols[idx]:
                        st.metric(
                            f"{lang}",
                            f"✓ {stats[lang]['success']}",
                            f"✗ {stats[lang]['error']}" if stats[lang]['error'] > 0 else None,
                            delta_color="inverse"
                        )

                st.subheader("📊 翻译结果预览")
                
                # 构建显示列顺序：原文列 + 角色列 + 所有翻译结果列
                display_cols = [text_col]
                if role_col:
                    display_cols.append(role_col)
                display_cols.extend([column_names[lang] for lang in languages])
                
                st.dataframe(df[display_cols].head(20))

                # 提供下载按钮
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='翻译结果')

                st.download_button(
                    label="💾 下载多语言翻译结果",
                    data=output.getvalue(),
                    file_name=f"translated_multilang_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.success(f"📁 文件已自动保存至: {final_path}")

            except KeyboardInterrupt:
                st.warning("⚠️ 翻译被中断，正在保存当前进度...")
                try:
                    interrupt_filename = f"translation_interrupted_{timestamp}.xlsx"
                    interrupt_path = os.path.join(save_directory, interrupt_filename)
                    with pd.ExcelWriter(interrupt_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='翻译进度')
                    st.info(f"💾 进度已保存至: {interrupt_path}")
                except Exception as save_error:
                    st.error(f"❌ 保存进度失败: {save_error}")

        except Exception as e:
            st.error(f"❌ 翻译过程中出现错误: {e}")
            import traceback
            st.error(traceback.format_exc())
            
            # 尝试保存当前进度
            try:
                error_filename = f"translation_error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                error_path = os.path.join(save_directory, error_filename)
                with pd.ExcelWriter(error_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='翻译进度')
                st.info(f"💾 错误前的进度已保存至: {error_path}")
            except Exception as save_error:
                st.error(f"❌ 保存进度失败: {save_error}")
def get_api_providers():
    providers = {
        "DeepSeek": {
            "url": "https://api.deepseek.com/v1/chat/completions",
            "models": ["deepseek-chat", "deepseek-coder"]
        },
        "OpenAI": {
            "url": "https://api.openai.com/v1/chat/completions",
            "models": ["gpt-3.5-turbo", "gpt-4", "gpt-4-turbo"]
        },
        "自定义API": {
            "url": "https://tb.api.mkeai.com/v1/chat/completions",
            "models": ["custom-model"]
        }
    }
    return providers

def get_preset_options():
    presets = {
        "游戏UI简约风格": "游戏UI简约风格",
        "角色对话自然流畅": "角色对话自然流畅", 
        "专业术语统一": "专业术语统一",
        "保持原文风格": "保持原文风格",
        "本地化适配": "本地化适配",
        "保持上下文一致性": "保持上下文一致性",
        "根据角色调整语气": "根据角色调整语气"
    }
    return presets

def get_preset_languages():
    return ["英文", "日文", "韩文", "法文", "德文", "西班牙文", "自定义"]

def get_default_custom_requirements():
    return "角色对话自然流畅；专业术语统一；保持原文风格；本地化适配；保持上下文一致性；根据角色调整语气；请注意使用语体，且所有角色除了微型机和炽长，其他都为女生用语，不要用男性用语，现在角色们都十分熟悉彼此了，不需要使用太正式尊重的语体了例如日语的话不需要ですます型了。"

def parse_ai_translation_result(text):
    try:
        text = text.strip()
        lines = text.split('\n')
        translations = {}
        
        # 查找表格开始位置
        table_start = -1
        header_found = False
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line or not '|' in line:
                continue
            
            # 检查是否是表头行
            if ('原文' in line or '中文' in line) and ('翻译' in line or 'Translation' in line or '英文' in line or '日文' in line):
                header_found = True
                # 检查下一行是否是分隔线
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    if '|' in next_line and ('---' in next_line or '===' in next_line or '--' in next_line):
                        table_start = i + 2  # 数据从分隔线后开始
                    else:
                        table_start = i + 1  # 没有分隔线，数据从表头后开始
                else:
                    table_start = i + 1
                break
        
        # 如果没找到标准表头，尝试查找任何包含 | 的行作为数据起点
        if not header_found:
            for i, line in enumerate(lines):
                if '|' in line and not ('---' in line or '===' in line):
                    # 尝试解析这一行，看是否有两列数据
                    parts = [part.strip() for part in line.split('|') if part.strip()]
                    if len(parts) >= 2:
                        table_start = i
                        break
        
        if table_start == -1:
            st.warning("未找到表格结构，尝试使用备用解析方法...")
            return parse_fallback_format(text)
        
        # 解析表格数据
        success_count = 0
        for i in range(table_start, len(lines)):
            line = lines[i].strip()
            
            # 跳过空行和分隔线
            if not line or not '|' in line:
                continue
            if '---' in line or '===' in line:
                continue
                
            # 分割行，移除空白部分
            parts = [part.strip() for part in line.split('|')]
            # 移除首尾的空字符串（来自行首尾的|）
            parts = [p for p in parts if p]
            
            if len(parts) >= 2:
                original_text = parts[0]
                translation_text = parts[1]
                
                # 清理Markdown格式符号
                original_text = re.sub(r'\*\*|\*|`|#', '', original_text).strip()
                translation_text = re.sub(r'\*\*|\*|`|#', '', translation_text).strip()
                
                # 只添加非空的有效翻译
                if original_text and translation_text:
                    if original_text not in translations:
                        translations[original_text] = translation_text
                        success_count += 1
        
        if success_count > 0:
            st.success(f"✅ 成功解析 {success_count} 条翻译")
        else:
            st.warning("表格解析成功但未找到有效数据，尝试备用方法...")
            return parse_fallback_format(text)
        
        return translations
        
    except Exception as e:
        st.error(f"解析AI翻译结果时出错: {e}")
        st.warning("尝试使用备用解析方法...")
        return parse_fallback_format(text)

def parse_fallback_format(text):
    try:
        translations = {}
        lines = text.strip().split('\n')
        success_count = 0
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 跳过分隔线和表头
            if '---' in line or '===' in line:
                continue
            if '原文' in line or 'Translation' in line:
                continue
                
            # 尝试多种分隔符
            if '|' in line:
                # 处理表格格式
                parts = [p.strip() for p in line.split('|')]
                # 移除空字符串
                parts = [p for p in parts if p]
                
                if len(parts) >= 2:
                    original = parts[0]
                    translation = parts[1]
                    
                    # 清理文本
                    original = re.sub(r'\*\*|\*|`|#', '', original).strip()
                    translation = re.sub(r'\*\*|\*|`|#', '', translation).strip()
                    
                    if original and translation:
                        if original not in translations:
                            translations[original] = translation
                            success_count += 1
            elif '\t' in line:
                # 处理制表符分隔
                parts = line.split('\t')
                if len(parts) >= 2:
                    original = parts[0].strip()
                    translation = parts[1].strip()
                    
                    original = re.sub(r'\*\*|\*|`|#', '', original).strip()
                    translation = re.sub(r'\*\*|\*|`|#', '', translation).strip()
                    
                    if original and translation:
                        if original not in translations:
                            translations[original] = translation
                            success_count += 1
        
        if success_count > 0:
            st.info(f"📝 备用方法成功解析 {success_count} 条翻译")
        else:
            st.error("❌ 备用方法也未能解析出有效数据")
        
        return translations
    except Exception as e:
        st.error(f"备用解析方法也失败: {e}")
        return {}

def merge_translations_with_excel(original_df, text_col, translations, target_language):
    try:
        result_df = original_df.copy()
        result_df[f'{target_language}翻译结果'] = ''
        
        matched_count = 0
        unmatched_texts = []
        
        for index, row in result_df.iterrows():
            original_text = str(row[text_col]) if not pd.isna(row[text_col]) else ''
            if original_text and original_text in translations:
                result_df.at[index, f'{target_language}翻译结果'] = translations[original_text]
                matched_count += 1
            elif original_text:
                unmatched_texts.append(original_text)
        
        return result_df, matched_count, unmatched_texts
    except Exception as e:
        st.error(f"合并翻译结果时出错: {e}")
        return original_df, 0, []



# 页面2: 提示词生成器（原第二个程序）
def prompt_generator_page():
    st.title("📝 单语种翻译提示词生成器")
    st.markdown("### 根据待翻译文本、术语库和角色性格信息，生成针对单一目标语言的翻译提示词。")
    st.markdown("**注意：** 本页面仅用于生成提示词文本，不进行实际的API翻译调用。")
    
    if 'prompt_translator' not in st.session_state:
        st.session_state.prompt_translator = MultiAPIExcelTranslator(
            api_key="", 
            api_provider="DeepSeek", 
            api_url=get_api_providers()["DeepSeek"]["url"], 
            model="deepseek-chat"
        )
    
    translator = st.session_state.prompt_translator
    
    if 'term_base_loaded' not in st.session_state:
        st.session_state.term_base_loaded = False
    if 'role_personality_loaded' not in st.session_state:
        st.session_state.role_personality_loaded = False
    
    st.header("🎯 基本设置")
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        language_option = st.selectbox(
            "🌍 选择目标语言",
            options=get_preset_languages(),
            index=0,
            key="prompt_language_option"
        )
    
    with col2:
        if language_option == "自定义":
            custom_language = st.text_input(
                "✏️ 输入自定义语言",
                value=st.session_state.get('prompt_custom_language', ''),
                placeholder="例如：俄文、葡萄牙文、阿拉伯文等",
                key="prompt_custom_language_input"
            )
            st.session_state.prompt_custom_language = custom_language
            target_language = custom_language
        else:
            target_language = language_option
            st.session_state.prompt_custom_language = ""
    
    if target_language:
        translator.set_target_language(target_language)
        st.info(f"🎯 当前目标语言: {target_language}")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.header("📁 待翻译文本")
        
        uploaded_file = st.file_uploader(
            "📄 上传待翻译文本文件 (Excel)",
            type=['xlsx', 'xls'],
            key="prompt_file_uploader"
        )
        
        df_text = None
        text_col = None
        role_col = None
        personality_col = None
        
        if uploaded_file is not None:
            try:
                df_text = pd.read_excel(uploaded_file)
                # 清理列名中的换行符和空白字符
                df_text.columns = df_text.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.prompt_df_text = df_text
                st.success(f"✅ 成功读取文件，共 {len(df_text)} 行数据")
                
                with st.expander("📊 文件预览"):
                    st.dataframe(df_text.head(10))
                
                cols = df_text.columns.tolist()
                text_col = st.selectbox(
                    "📝 选择文本列",
                    options=cols,
                    index=0,
                    key="prompt_text_col_select"
                )
                
                role_col = st.selectbox(
                    "👥 选择说话人/角色列 (可选)",
                    options=["无"] + cols,
                    index=0,
                    key="prompt_role_col_select"
                )
                role_col = role_col if role_col != "无" else None
                
                personality_col = st.selectbox(
                    "💬 选择性格描述列 (可选)",
                    options=["无"] + cols,
                    index=0,
                    key="prompt_personality_col_select"
                )
                personality_col = personality_col if personality_col != "无" else None
                
                st.session_state.prompt_text_col = text_col
                st.session_state.prompt_role_col = role_col
                st.session_state.prompt_personality_col = personality_col
                
            except Exception as e:
                st.error(f"❌ 文件读取失败: {e}")
        else:
            if 'prompt_df_text' in st.session_state:
                df_text = st.session_state.prompt_df_text
                text_col = st.session_state.get('prompt_text_col')
                role_col = st.session_state.get('prompt_role_col')
                personality_col = st.session_state.get('prompt_personality_col')
                
                # 显示当前已加载的文件信息
                if df_text is not None:
                    st.info(f"✅ 已加载文件：{len(df_text)} 行数据")
                    if text_col:
                        st.write(f"📝 文本列: {text_col}")
                    if role_col:
                        st.write(f"👥 角色列: {role_col}")
                    if personality_col:
                        st.write(f"💬 性格列: {personality_col}")
        
        st.subheader("✂️ 分批次设置")
        batch_size = st.number_input(
            "每批次行数",
            min_value=1,
            max_value=200,
            value=50,
            step=10,
            key="prompt_batch_size"
        )
    
    with col2:
        st.header("📚 术语库和性格库")
        
        st.subheader("📚 术语库功能")
        uploaded_term_base = st.file_uploader(
            "📚 上传术语库文件 (Excel)",
            type=['xlsx', 'xls'],
            key="prompt_term_base_uploader"
        )
        
        if uploaded_term_base is not None:
            try:
                df = pd.read_excel(uploaded_term_base)
                # 清理列名中的换行符和空白字符
                df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.prompt_term_base_df = df
                st.success(f"✅ 成功读取术语库，共 {len(df)} 条记录")
                
                with st.expander("📊 术语库预览"):
                    st.dataframe(df.head(10))
                
                cols = df.columns.tolist()
                
                source_col = st.selectbox(
                    "📝 选择中文源文列",
                    options=cols,
                    index=0,
                    key="prompt_term_source_col"
                )
                
                target_col = st.selectbox(
                    "📤 选择翻译列",
                    options=cols,
                    index=min(1, len(cols)-1) if len(cols) > 1 else 0,
                    key="prompt_term_target_col"
                )
                
                if st.button("📥 加载术语库", key="prompt_load_term_base"):
                    if translator.load_term_base(df, source_col, target_col):
                        st.session_state.term_base_loaded = True
                        st.success("✅ 术语库加载成功")
                        st.rerun()
                
            except Exception as e:
                st.error(f"❌ 处理术语库文件失败: {e}")
        
        if st.session_state.get('term_base_loaded', False):
            st.info(f"✅ 术语库已加载: {len(translator.term_base_list)} 条术语")
        
        st.divider()
        
        st.subheader("👤 角色性格库功能")
        uploaded_role = st.file_uploader(
            "📋 上传角色性格库文件 (Excel)",
            type=['xlsx', 'xls'],
            key="prompt_role_personality_uploader"
        )
        
        if uploaded_role is not None:
            try:
                df = pd.read_excel(uploaded_role)
                # 清理列名中的换行符和空白字符
                df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.prompt_role_personality_df = df
                st.success(f"✅ 成功读取角色性格库，共 {len(df)} 条记录")
                
                with st.expander("📊 角色性格库预览"):
                    st.dataframe(df.head(10))
                
                cols = df.columns.tolist()
                role_col = st.selectbox(
                    "👥 选择角色名称列",
                    options=cols,
                    index=0,
                    key="prompt_role_name_col"
                )
                
                personality_col = st.selectbox(
                    "💬 选择性格描述列",
                    options=cols,
                    index=min(1, len(cols)-1) if len(cols) > 1 else 0,
                    key="prompt_personality_desc_col"
                )
                
                if st.button("📥 加载角色性格库", key="prompt_load_role_personality"):
                    if translator.load_role_personality(df, role_col, personality_col):
                        st.session_state.role_personality_loaded = True
                        st.success("✅ 角色性格库加载成功")
                        st.rerun()
                
            except Exception as e:
                st.error(f"❌ 处理角色性格库文件失败: {e}")
        
        if st.session_state.get('role_personality_loaded', False):
            st.info(f"✅ 角色性格库已加载: {len(translator.role_personality_dict)} 条角色")
    
    st.divider()
    
    st.header("🎯 翻译要求设置")
    
    st.subheader("🏷️ 预设选项")
    presets = get_preset_options()
    preset_options = st.multiselect(
        "选择预设翻译要求（可多选）",
        options=list(presets.keys()),
        default=st.session_state.get('prompt_preset_options', []),
        key="prompt_preset_multiselect"
    )
    st.session_state.prompt_preset_options = preset_options
    
    custom_requirements = st.text_area(
        "💬 自定义翻译要求",
        value=st.session_state.get('prompt_custom_requirements', get_default_custom_requirements()),
        placeholder=f"例如：游戏UI简约风格、角色对话自然流畅、专业术语统一、{target_language}本地化适配等",
        height=100,
        key="prompt_custom_requirements_text"
    )
    st.session_state.prompt_custom_requirements = custom_requirements
    
    all_requirements = []
    if preset_options:
        all_requirements.extend(preset_options)
    if custom_requirements:
        all_requirements.append(custom_requirements)
    
    final_requirements = "；".join(all_requirements) if all_requirements else ""
    
    if st.button("🚀 生成提示词", key="generate_prompt_btn", use_container_width=True):
        if df_text is None or text_col is None:
            st.error("❌ 请先上传待翻译文本文件并选择文本列。")
            return
        
        # 从session_state获取最新的列选择
        text_col = st.session_state.get('prompt_text_col')
        role_col = st.session_state.get('prompt_role_col')
        personality_col = st.session_state.get('prompt_personality_col')
        
        if not text_col:
            st.error("❌ 请选择文本列。")
            return
        
        if not target_language or target_language.strip() == "":
            st.error("❌ 请先选择或输入目标语言。")
            return
        
        term_base_loaded = st.session_state.get('term_base_loaded', False)
        role_personality_loaded = st.session_state.get('role_personality_loaded', False)
        
        if term_base_loaded:
            st.info(f"✅ 术语库已加载: {len(translator.term_base_list)} 条术语")
        else:
            st.warning("⚠️ 未加载术语库，提示词中将不包含术语匹配信息")
        
        if role_personality_loaded:
            st.info(f"✅ 角色性格库已加载: {len(translator.role_personality_dict)} 条角色")
        else:
            st.warning("⚠️ 未加载角色性格库，提示词中将不包含角色性格信息")
        
        fixed_requirements = f"""
## 翻译要求：(固定)
你是一名专业的二次元游戏本地化翻译专家，擅长将中文二次元游戏文案翻译为{target_language}。请将用户输入的中文游戏文本，以表格形式输出对应的{target_language}翻译。表格应包含两列：原文（中文）、{target_language}翻译。"""
        
        language_specific_requirements = translator.get_language_specific_requirements(target_language)
        
        other_requirements = f"""
## 其他要求：(用户输入)
{final_requirements if final_requirements else "无"}
"""
        
        important_notes = f"""
## 重要说明！：(固定)
• 请只返回翻译后的文本结果，以表格形式输出中文原文，{target_language}翻译
• 不要添加任何解释或备注
• 术语库中的特定词汇翻译需要严格采用相同的翻译
• 请根据角色性格特点调整翻译风格和语气。
• 本次翻译目标语言为：{target_language}
"""
        
        num_batches = (len(df_text) + batch_size - 1) // batch_size
        all_prompts = []
        
        for i in range(num_batches):
            start_index = i * batch_size
            end_index = min((i + 1) * batch_size, len(df_text))
            batch_df = df_text.iloc[start_index:end_index].copy()
            
            translator.reset_context()
            
            # 获取文本列数据
            text_list = batch_df[text_col].tolist()
            
            # 安全获取角色列数据
            if role_col and role_col != "无" and role_col in batch_df.columns:
                role_list = batch_df[role_col].tolist()
            else:
                role_list = [None] * len(batch_df)
            
            # 安全获取性格列数据
            if personality_col and personality_col != "无" and personality_col in batch_df.columns:
                personality_list = batch_df[personality_col].tolist()
            else:
                personality_list = [None] * len(batch_df)
            
            all_text_in_batch = " ".join([str(t) for t in text_list if not pd.isna(t)])
            
            term_base_prompt = ""
            if term_base_loaded:
                term_base_prompt = translator.build_term_base_prompt(all_text_in_batch)
            else:
                term_base_prompt = "\n\n### 术语库匹配：\n无术语库加载，跳过术语匹配。"
            
            text_and_personality_prompt = "### 待翻译文本及说话人性格格式：(用户输入)\n"
            text_and_personality_prompt += "说话人\t原文\t说话人性格\n"
            
            for j in range(len(batch_df)):
                role = role_list[j]
                text = text_list[j]
                personality = personality_list[j]
                
                if personality_col and not pd.isna(personality):
                    personality_desc = str(personality).strip()
                elif role_col and role and role_personality_loaded:
                    personality_desc = translator.find_role_personality(role)
                    personality_desc = personality_desc if personality_desc else "无"
                else:
                    personality_desc = "无"
                
                role_name = str(role).strip() if role and not pd.isna(role) else "无"
                text_content = str(text).strip() if not pd.isna(text) else ""
                
                text_and_personality_prompt += f"{role_name}\t{text_content}\t{personality_desc}\n"
            
            full_prompt = f"""
# 批次 {i+1}/{num_batches} - 目标语言: {target_language}

{fixed_requirements}

{language_specific_requirements}

{other_requirements}

{term_base_prompt}

{text_and_personality_prompt}

{important_notes}
"""
            all_prompts.append(full_prompt.strip())
        
        st.subheader("✅ 生成结果")
        
        st.session_state.all_prompts = all_prompts
        st.session_state.num_batches = num_batches
        st.session_state.current_batch_index = 0
        st.session_state.target_language = target_language
        
        st.success(f"✅ 提示词生成成功，共 {num_batches} 个批次，目标语言: {target_language}。")
    
    if st.session_state.get('all_prompts'):
        all_prompts = st.session_state.all_prompts
        num_batches = st.session_state.num_batches
        current_batch_index = st.session_state.current_batch_index
        target_language = st.session_state.get('target_language', '英文')
        
        st.subheader(f"批次 {current_batch_index + 1}/{num_batches} 提示词 - 目标语言: {target_language}")
        
        current_prompt = all_prompts[current_batch_index]
        
        st.code(current_prompt, language=None)
        st.info("👆 请使用上方代码块右下角的复制按钮进行一键复制。")
        
        col_prev, col_info, col_next = st.columns([1, 2, 1])
        
        with col_prev:
            if st.button("上一批次", disabled=(current_batch_index == 0), key="prompt_prev_batch"):
                st.session_state.current_batch_index -= 1
                st.rerun()
                
        with col_info:
            st.markdown(f"<p style='text-align: center;'>当前批次: {current_batch_index + 1} / {num_batches} | 目标语言: {target_language}</p>", unsafe_allow_html=True)
            
        with col_next:
            if st.button("下一批次", disabled=(current_batch_index == num_batches - 1), key="prompt_next_batch"):
                st.session_state.current_batch_index += 1
                st.rerun()
                
        st.markdown("---")
        
        final_output = f"# 翻译提示词 - 目标语言: {target_language}\n\n" + ("-"*80) + "\n\n".join(all_prompts)
        
        st.download_button(
            label=f"📥 下载所有提示词 (.txt)",
            data=final_output,
            file_name=f"translation_prompts_{target_language}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
            mime="text/plain",
            use_container_width=True
        )

class ExcelSearchReplace:
    def __init__(self):
        self.excel_files = []
        self.search_results = {}
        self.case_sensitive = False
        self.match_whole_word = False
        
    def find_excel_files(self, folder_path):
        """查找文件夹中的所有Excel文件"""
        self.excel_files = []
        folder_path = Path(folder_path)
        
        if not folder_path.exists():
            return False, "文件夹路径不存在"
        
        # 支持的Excel文件扩展名
        excel_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']
        
        for ext in excel_extensions:
            self.excel_files.extend(folder_path.rglob(f'*{ext}'))
        
        return True, f"找到 {len(self.excel_files)} 个Excel文件"
    
    def search_in_excel(self, search_term, case_sensitive=False, match_whole_word=False):
        """在Excel文件中搜索词语"""
        self.search_results = {}
        self.case_sensitive = case_sensitive
        self.match_whole_word = match_whole_word
        total_matches = 0
        
        for file_path in self.excel_files:
            try:
                # 读取Excel文件的所有工作表
                excel_data = pd.read_excel(file_path, sheet_name=None, dtype=str)
                file_matches = []
                
                for sheet_name, df in excel_data.items():
                    sheet_matches = self._search_in_dataframe(df, search_term, sheet_name, str(file_path))
                    file_matches.extend(sheet_matches)
                
                if file_matches:
                    self.search_results[str(file_path)] = {
                        'matches': file_matches,
                        'match_count': len(file_matches)
                    }
                    total_matches += len(file_matches)
                    
            except Exception as e:
                st.error(f"读取文件 {file_path.name} 时出错: {e}")
        
        return total_matches
    
    def _search_in_dataframe(self, df, search_term, sheet_name, file_path):
        """在DataFrame中搜索词语"""
        matches = []
        
        # 构建正则表达式模式
        if self.match_whole_word:
            pattern = r'\b' + re.escape(search_term) + r'\b'
        else:
            pattern = re.escape(search_term)
        
        flags = 0 if self.case_sensitive else re.IGNORECASE
        
        for row_idx, row in df.iterrows():
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue
                
                cell_str = str(cell_value)
                matches_found = list(re.finditer(pattern, cell_str, flags))
                
                for match in matches_found:
                    matches.append({
                        'file_path': file_path,
                        'sheet_name': sheet_name,
                        'row': row_idx + 2,  # +2 因为Excel从1开始，且有标题行
                        'column': df.columns[col_idx] if col_idx < len(df.columns) else f'Col{col_idx+1}',
                        'original_text': cell_str,
                        'matched_text': match.group(),
                        'start_pos': match.start(),
                        'end_pos': match.end()
                    })
        
        return matches
    
    def replace_in_excel(self, search_term, replace_term, backup=True):
        """替换Excel文件中的词语"""
        replaced_files = 0
        total_replacements = 0
        
        for file_path_str, file_data in self.search_results.items():
            file_path = Path(file_path_str)
            
            try:
                # 备份原文件
                if backup:
                    backup_path = file_path.parent / f"{file_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_path.suffix}"
                    shutil.copy2(file_path, backup_path)
                    st.info(f"已创建备份: {backup_path.name}")
                
                # 读取Excel文件
                excel_data = pd.read_excel(file_path, sheet_name=None, dtype=str)
                replacements_in_file = 0
                
                # 构建替换模式
                if self.match_whole_word:
                    pattern = r'\b' + re.escape(search_term) + r'\b'
                else:
                    pattern = re.escape(search_term)
                
                flags = 0 if self.case_sensitive else re.IGNORECASE
                
                # 对每个工作表进行替换
                for sheet_name, df in excel_data.items():
                    df_replaced = df.applymap(
                        lambda x: self._replace_text(x, pattern, replace_term, flags) 
                        if pd.notna(x) else x
                    )
                    excel_data[sheet_name] = df_replaced
                    
                    # 计算替换数量
                    for row_idx, row in df.iterrows():
                        for col_idx, cell_value in enumerate(row):
                            if pd.isna(cell_value):
                                continue
                            cell_str = str(cell_value)
                            replacements = len(re.findall(pattern, cell_str, flags))
                            replacements_in_file += replacements
                
                # 保存替换后的文件
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet_name, df in excel_data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                replaced_files += 1
                total_replacements += replacements_in_file
                
                st.success(f"✅ {file_path.name}: 完成 {replacements_in_file} 处替换")
                
            except Exception as e:
                st.error(f"替换文件 {file_path.name} 时出错: {e}")
        
        return replaced_files, total_replacements
    
    def _replace_text(self, text, pattern, replace_term, flags):
        """替换文本中的匹配项"""
        if pd.isna(text):
            return text
        
        text_str = str(text)
        replaced_text = re.sub(pattern, replace_term, text_str, flags=flags)
        return replaced_text

import streamlit as st
import pandas as pd
from pathlib import Path
import os
import subprocess
import platform
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
def grand_match():
    model_GRAND_match.model_grand_match.grand_match()
def excel_replace_page():
    st.set_page_config(
        page_title="Excel文件批量搜索替换工具",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    st.title("🔍 Excel文件批量搜索替换工具")
    st.markdown("### 批量搜索和替换文件夹中所有Excel文件的内容")
    
    # 初始化搜索替换工具
    if 'search_tool' not in st.session_state:
        st.session_state.search_tool = ExcelSearchReplace()
    
    # 初始化会话状态变量
    if 'folder_path' not in st.session_state:
        st.session_state.folder_path = ""
    if 'search_term' not in st.session_state:
        st.session_state.search_term = ""
    if 'replace_term' not in st.session_state:
        st.session_state.replace_term = ""
    if 'case_sensitive' not in st.session_state:
        st.session_state.case_sensitive = False
    if 'match_whole_word' not in st.session_state:
        st.session_state.match_whole_word = False
    if 'replace_confirmed' not in st.session_state:
        st.session_state.replace_confirmed = False
    if 'show_confirm_checkbox' not in st.session_state:
        st.session_state.show_confirm_checkbox = False
    if 'edited_data' not in st.session_state:
        st.session_state.edited_data = {}
    
    search_tool = st.session_state.search_tool
    
    # 侧边栏 - 文件夹选择
    st.sidebar.header("📁 文件夹设置")
    folder_path = st.sidebar.text_input(
        "请输入文件夹路径:",
        value=st.session_state.folder_path,
        placeholder="例如: C:/Users/用户名/Documents/Excel文件",
        help="请输入包含Excel文件的文件夹完整路径"
    )
    
    if folder_path and folder_path != st.session_state.folder_path:
        st.session_state.folder_path = folder_path
        success, message = search_tool.find_excel_files(folder_path)
        if success:
            st.sidebar.success(message)
        else:
            st.sidebar.error(message)
    
    # 显示找到的文件列表
    if search_tool.excel_files:
        st.sidebar.subheader("📊 找到的Excel文件")
        for i, file_path in enumerate(search_tool.excel_files[:10]):  # 只显示前10个
            st.sidebar.write(f"{i+1}. {file_path.name}")
        
        if len(search_tool.excel_files) > 10:
            st.sidebar.info(f"... 还有 {len(search_tool.excel_files) - 10} 个文件")
    
    # 主界面 - 搜索设置
    st.header("🔍 搜索设置")
    
    col1, col2 = st.columns(2)
    
    with col1:
        search_term = st.text_input(
            "搜索词语:",
            value=st.session_state.search_term,
            placeholder="请输入要搜索的词语",
            help="支持正则表达式语法"
        )
        st.session_state.search_term = search_term
    
    with col2:
        # 搜索选项
        st.subheader("⚙️ 搜索选项")
        case_sensitive = st.checkbox(
            "大小写敏感",
            value=st.session_state.case_sensitive,
            help="勾选后区分大小写"
        )
        st.session_state.case_sensitive = case_sensitive
        
        match_whole_word = st.checkbox(
            "全词匹配",
            value=st.session_state.match_whole_word,
            help="勾选后只匹配完整词语"
        )
        st.session_state.match_whole_word = match_whole_word
    
    # 搜索按钮
    if st.button("🚀 开始搜索", key="search_btn", use_container_width=True):
        if not folder_path:
            st.error("❌ 请输入文件夹路径")
            return
        
        if not search_term:
            st.error("❌ 请输入搜索词语")
            return
        
        # 执行多线程搜索
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        with st.spinner("正在搜索Excel文件..."):
            total_matches = multithreaded_search(
                search_tool,
                search_term,
                case_sensitive,
                match_whole_word,
                progress_bar,
                status_text
            )
        
        progress_bar.empty()
        status_text.empty()
        
        if total_matches > 0:
            st.success(f"✅ 搜索完成！共找到 {total_matches} 个匹配项")
        else:
            st.warning("⚠️ 未找到匹配项")
    
    # 显示搜索结果
    if search_tool.search_results:
        st.header("📊 搜索结果预览")
        
        total_files = len(search_tool.search_results)
        total_matches = sum(data['match_count'] for data in search_tool.search_results.values())
        
        st.info(f"**统计信息:** 在 {total_files} 个文件中找到 {total_matches} 个匹配项")
        
        # 文件列表
        selected_file = st.selectbox(
            "选择文件查看详情:",
            options=list(search_tool.search_results.keys()),
            format_func=lambda x: f"{Path(x).name} ({search_tool.search_results[x]['match_count']} 处)"
        )
        
        if selected_file:
            file_data = search_tool.search_results[selected_file]
            matches = file_data['matches']
            
            # 文件信息和操作按钮
            st.subheader(f"📄 文件: {Path(selected_file).name}")
            
            # 显示完整文件路径
            st.code(selected_file, language=None)
            
            # 操作按钮行
            col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 3])
            
            with col_btn1:
                if st.button("📂 打开文件夹", key=f"open_folder_{selected_file}"):
                    open_folder(selected_file)
            
            with col_btn2:
                if st.button("📊 打开Excel", key=f"open_excel_{selected_file}"):
                    open_file(selected_file)
            
            st.write(f"**匹配数量:** {len(matches)} 处")
            
            # 显示匹配详情 - 使用可编辑的表格
            display_rows = []
            row_identifiers = []  # 存储行标识符用于后续保存
            
            for i, match in enumerate(matches[:50]):  # 只显示前50个
                # 获取该行的完整数据
                row_data = get_row_data_as_list(selected_file, match['sheet_name'], match['row'])
                
                # 添加位置信息作为第一列
                row_dict = {
                    "位置": f"{match['sheet_name']} | 行{match['row']} 列{match['column']}"
                }
                
                # 添加该行的所有列数据
                if row_data:
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        row_dict[f"列{col_idx}"] = cell_value
                
                display_rows.append(row_dict)
                row_identifiers.append({
                    'sheet': match['sheet_name'],
                    'row': match['row'],
                    'original_data': row_data
                })
            
            if display_rows:
                st.markdown("### ✏️ 可编辑表格（直接修改单元格内容）")
                
                # 创建DataFrame
                df_display = pd.DataFrame(display_rows)
                
                # 使用data_editor创建可编辑表格
                edited_df = st.data_editor(
                    df_display,
                    use_container_width=True,
                    height=400,
                    key=f"editable_table_{selected_file}",
                    column_config={
                        "位置": st.column_config.TextColumn(
                            "位置",
                            disabled=True,  # 位置列不可编辑
                            width="medium"
                        )
                    }
                )
                
                # 检测是否有修改
                if not df_display.equals(edited_df):
                    st.warning("⚠️ 检测到内容已修改，请点击下方按钮保存更改")
                    
                    # 显示修改对比
                    with st.expander("📋 查看修改详情"):
                        changes_found = False
                        for idx in range(len(df_display)):
                            for col in df_display.columns:
                                if col != "位置":  # 跳过位置列
                                    old_val = df_display.iloc[idx][col]
                                    new_val = edited_df.iloc[idx][col]
                                    if old_val != new_val:
                                        changes_found = True
                                        st.markdown(f"**{df_display.iloc[idx]['位置']} - {col}:**")
                                        st.markdown(f"- 原值: `{old_val}`")
                                        st.markdown(f"- 新值: `{new_val}`")
                                        st.markdown("---")
                        
                        if not changes_found:
                            st.info("未检测到有效修改")
                    
                    # 保存修改按钮
                    col_save1, col_save2 = st.columns([1, 3])
                    
                    with col_save1:
                        if st.button("💾 保存修改到Excel", key=f"save_edits_{selected_file}", type="primary"):
                            try:
                                # 加载Excel文件
                                wb = openpyxl.load_workbook(selected_file)
                                
                                # 遍历所有修改
                                changes_count = 0
                                for idx in range(len(edited_df)):
                                    sheet_name = row_identifiers[idx]['sheet']
                                    row_num = row_identifiers[idx]['row']
                                    ws = wb[sheet_name]
                                    
                                    # 检查每一列的修改
                                    for col in edited_df.columns:
                                        if col != "位置":
                                            col_idx = int(col.replace("列", ""))
                                            new_val = edited_df.iloc[idx][col]
                                            old_val = df_display.iloc[idx][col]
                                            
                                            if new_val != old_val:
                                                # 写入新值到Excel
                                                ws.cell(row=row_num, column=col_idx, value=new_val)
                                                changes_count += 1
                                
                                # 保存文件
                                wb.save(selected_file)
                                wb.close()
                                
                                st.success(f"✅ 成功保存 {changes_count} 处修改！")
                                
                                # 提示重新搜索
                                st.info("💡 建议重新搜索以查看最新内容")
                                
                            except Exception as e:
                                st.error(f"❌ 保存失败: {str(e)}")
                    
                    with col_save2:
                        if st.button("🔄 撤销修改", key=f"reset_edits_{selected_file}"):
                            st.rerun()
            
            if len(matches) > 50:
                st.info(f"仅显示前 50 个匹配项，共有 {len(matches)} 个匹配项")
        
        # 替换功能
        st.header("🔄 批量替换功能")
        
        col1, col2 = st.columns(2)
        
        with col1:
            replace_term = st.text_input(
                "替换为:",
                value=st.session_state.replace_term,
                placeholder="请输入替换后的词语",
                help="将搜索到的词语替换为此词语"
            )
            st.session_state.replace_term = replace_term
        
        with col2:
            backup_files = st.checkbox(
                "创建备份文件",
                value=True,
                help="替换前自动创建备份文件"
            )
        
        # 替换预览
        if search_term and replace_term:
            st.subheader("🔍 替换预览")
            
            # 显示替换前后对比示例
            example_before = f"这是包含 {search_term} 的示例文本"
            example_after = example_before.replace(search_term, f"**{replace_term}**")
            
            col_before, col_arrow, col_after = st.columns([1, 0.1, 1])
            
            with col_before:
                st.text_area("替换前:", value=example_before, height=60, disabled=True)
            
            with col_arrow:
                st.markdown("<br><h2>→</h2>", unsafe_allow_html=True)
            
            with col_after:
                st.text_area("替换后:", value=example_after, height=60, disabled=True)
        
        # 选择性替换功能
        st.subheader("📋 选择要替换的项目")
        
        # 初始化选择状态
        if 'selected_replacements' not in st.session_state:
            st.session_state.selected_replacements = {}
        
        # 全选/全不选按钮
        col_select1, col_select2, col_select3 = st.columns([1, 1, 3])
        with col_select1:
            if st.button("✅ 全选", use_container_width=True):
                for file_path in search_tool.search_results.keys():
                    st.session_state.selected_replacements[file_path] = {
                        'selected': True,
                        'rows': 'all'
                    }
                st.rerun()
        
        with col_select2:
            if st.button("❌ 全不选", use_container_width=True):
                st.session_state.selected_replacements = {}
                st.rerun()
        
        # 为每个文件创建选择界面
        for file_path, file_data in search_tool.search_results.items():
            with st.expander(f"📄 {Path(file_path).name} ({file_data['match_count']} 处匹配)", expanded=False):
                # 文件级别的选择
                file_key = f"file_{file_path}"
                
                # 初始化该文件的选择状态
                if file_path not in st.session_state.selected_replacements:
                    st.session_state.selected_replacements[file_path] = {
                        'selected': False,
                        'rows': 'all',
                        'selected_rows': set()
                    }
                
                col_file1, col_file2 = st.columns([1, 3])
                
                with col_file1:
                    file_selected = st.checkbox(
                        "选择此文件",
                        value=st.session_state.selected_replacements[file_path]['selected'],
                        key=f"cb_{file_key}"
                    )
                    st.session_state.selected_replacements[file_path]['selected'] = file_selected
                
                with col_file2:
                    if file_selected:
                        replace_mode = st.radio(
                            "替换模式:",
                            options=['all', 'selected'],
                            format_func=lambda x: "替换所有匹配项" if x == 'all' else "选择特定行",
                            key=f"mode_{file_key}",
                            horizontal=True
                        )
                        st.session_state.selected_replacements[file_path]['rows'] = replace_mode
                        
                        # 如果选择了特定行模式，显示行选择界面
                        if replace_mode == 'selected':
                            st.markdown("**选择要替换的行:**")
                            
                            matches = file_data['matches']
                            # 按工作表分组
                            sheets_data = {}
                            for match in matches:
                                sheet_name = match['sheet_name']
                                if sheet_name not in sheets_data:
                                    sheets_data[sheet_name] = []
                                sheets_data[sheet_name].append(match)
                            
                            # 为每个工作表显示行选择
                            for sheet_name, sheet_matches in sheets_data.items():
                                st.markdown(f"*工作表: {sheet_name}*")
                                
                                # 获取唯一的行号
                                unique_rows = sorted(set(m['row'] for m in sheet_matches))
                                
                                cols = st.columns(5)
                                for idx, row_num in enumerate(unique_rows):
                                    with cols[idx % 5]:
                                        row_key = f"{file_path}_{sheet_name}_{row_num}"
                                        row_selected = st.checkbox(
                                            f"行 {row_num}",
                                            value=row_key in st.session_state.selected_replacements[file_path]['selected_rows'],
                                            key=f"row_{row_key}"
                                        )
                                        
                                        if row_selected:
                                            st.session_state.selected_replacements[file_path]['selected_rows'].add(row_key)
                                        elif row_key in st.session_state.selected_replacements[file_path]['selected_rows']:
                                            st.session_state.selected_replacements[file_path]['selected_rows'].remove(row_key)
        
        # 显示替换统计
        st.subheader("📊 替换统计")
        selected_files_count = sum(1 for f in st.session_state.selected_replacements.values() if f['selected'])
        total_selected_matches = 0
        
        for file_path, selection in st.session_state.selected_replacements.items():
            if selection['selected']:
                if selection['rows'] == 'all':
                    total_selected_matches += search_tool.search_results[file_path]['match_count']
                else:
                    # 计算选中的行数
                    total_selected_matches += len(selection['selected_rows'])
        
        col_stat1, col_stat2 = st.columns(2)
        with col_stat1:
            st.metric("选中的文件数", selected_files_count)
        with col_stat2:
            st.metric("预计替换项数", total_selected_matches)
        
        # 执行替换按钮
        if st.button("🔄 执行批量替换", key="replace_btn", type="primary", use_container_width=True):
            if not replace_term:
                st.error("❌ 请输入替换词语")
                return
            
            if selected_files_count == 0:
                st.error("❌ 请至少选择一个文件进行替换")
                return
            
            # 如果还没有确认,显示确认复选框
            if not st.session_state.replace_confirmed:
                st.warning(f"⚠️ 此操作将在 {selected_files_count} 个文件中执行约 {total_selected_matches} 处替换！")
                st.session_state.show_confirm_checkbox = True
        
        # 显示确认复选框
        if st.session_state.show_confirm_checkbox:
            confirm_replace = st.checkbox("我确认要执行批量替换操作")
            
            if confirm_replace:
                st.session_state.replace_confirmed = True
                st.session_state.show_confirm_checkbox = False
                st.rerun()
        
        # 如果已经确认,执行替换操作
        if st.session_state.replace_confirmed:
            # 执行选择性替换
            with st.spinner("正在执行替换操作..."):
                replaced_files, total_replacements = selective_replace(
                    search_tool,
                    search_term,
                    replace_term,
                    st.session_state.selected_replacements,
                    backup_files
                )
            
            if replaced_files > 0:
                st.success(f"✅ 替换完成！在 {replaced_files} 个文件中完成了 {total_replacements} 处替换")
                
                # 重置状态
                st.session_state.replace_confirmed = False
                st.session_state.show_confirm_checkbox = False
                st.session_state.selected_replacements = {}
                
                # 清空搜索结果,提示重新搜索
                search_tool.search_results = {}
                st.info("💡 替换完成，请重新搜索以查看更新后的内容")
                
                # 清空搜索和替换词
                st.session_state.search_term = ""
                st.session_state.replace_term = ""
            else:
                st.error("❌ 替换操作失败")
                st.session_state.replace_confirmed = False
                st.session_state.show_confirm_checkbox = False
    
    # 使用说明
    with st.expander("📖 使用说明"):
        st.markdown("""
        ## 使用说明
        
        ### 基本流程：
        1. **设置文件夹路径** - 在侧边栏输入包含Excel文件的文件夹路径
        2. **输入搜索词语** - 在主界面输入要搜索的词语
        3. **设置搜索选项** - 选择是否大小写敏感、全词匹配
        4. **开始搜索** - 点击"开始搜索"按钮
        5. **查看和编辑结果** - 浏览搜索结果，直接在表格中修改内容
        6. **保存单个文件修改** - 在可编辑表格中修改后点击"保存修改到Excel"
        7. **快速操作** - 使用"打开文件夹"或"打开Excel"按钮快速访问文件
        8. **批量替换** - 使用批量替换功能对多个文件执行统一替换
        
        ### 功能特点：
        - ✏️ **直接编辑** - 在搜索结果表格中直接修改单元格内容
        - 💾 **即时保存** - 修改后立即保存到Excel文件
        - 🔍 **多线程批量搜索** - 自动使用多线程加速搜索，充分利用CPU资源
        - 📊 **原表格展示** - 以原始表格形式显示匹配行的完整数据
        - 📂 **快速访问** - 一键打开文件所在文件夹或直接打开Excel文件
        - 🎯 **选择性替换** - 可以选择特定文件、特定行进行批量替换
        - ⚙️ **灵活选项** - 支持大小写敏感和全词匹配
        - 💾 **自动备份** - 批量替换前可自动创建备份文件
        - 📁 **多格式支持** - 支持 .xlsx, .xls, .xlsm, .xlsb 格式
        - ⚡ **实时进度** - 显示搜索和替换的实时进度
        
        ### 两种修改方式：
        1. **直接编辑（推荐用于少量精确修改）**
           - 在搜索结果表格中直接修改单元格
           - 点击"保存修改到Excel"即时保存
           - 适合修改个别单元格内容
        
        2. **批量替换（推荐用于大量统一替换）**
           - 选择要替换的文件和行
           - 执行统一的查找替换操作
           - 可创建备份文件
           - 适合大规模统一修改
        
        ### 注意事项：
        - 直接编辑会立即保存到文件，请谨慎操作
        - 批量替换操作会修改原文件，建议先创建备份
        - 建议先在小范围测试
        - 支持正则表达式语法（在搜索词语中）
        - 大型文件可能需要较长时间处理
        - 修改保存后建议重新搜索查看最新内容
        """)


def multithreaded_search(search_tool, search_term, case_sensitive, match_whole_word, progress_bar, status_text):
    """
    使用多线程搜索Excel文件
    
    Args:
        search_tool: ExcelSearchReplace工具实例
        search_term: 搜索词
        case_sensitive: 是否大小写敏感
        match_whole_word: 是否全词匹配
        progress_bar: 进度条对象
        status_text: 状态文本对象
    
    Returns:
        总匹配数
    """
    import openpyxl
    import re
    
    # 清空之前的搜索结果
    search_tool.search_results = {}
    
    if not search_tool.excel_files:
        return 0
    
    # 准备搜索模式
    if match_whole_word:
        pattern = r'\b' + re.escape(search_term) + r'\b'
    else:
        pattern = re.escape(search_term)
    
    flags = 0 if case_sensitive else re.IGNORECASE
    regex = re.compile(pattern, flags)
    
    # 线程锁，用于安全更新共享数据
    lock = threading.Lock()
    total_matches = 0
    completed_files = 0
    total_files = len(search_tool.excel_files)
    
    def search_single_file(file_path):
        """搜索单个Excel文件"""
        nonlocal total_matches, completed_files
        
        file_matches = []
        match_count = 0
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value is not None:
                            cell_text = str(cell.value)
                            
                            # 搜索匹配
                            matches = list(regex.finditer(cell_text))
                            
                            if matches:
                                for match in matches:
                                    file_matches.append({
                                        'sheet_name': sheet_name,
                                        'row': row_idx,
                                        'column': col_idx,
                                        'original_text': cell_text,
                                        'matched_text': match.group(),
                                        'start_pos': match.start(),
                                        'end_pos': match.end()
                                    })
                                    match_count += 1
            
            wb.close()
            
            # 使用锁更新共享数据
            if file_matches:
                with lock:
                    search_tool.search_results[str(file_path)] = {
                        'matches': file_matches,
                        'match_count': match_count
                    }
                    total_matches += match_count
            
        except Exception as e:
            # 忽略无法读取的文件
            pass
        
        # 更新进度
        with lock:
            completed_files += 1
            progress = completed_files / total_files
            progress_bar.progress(progress)
            status_text.text(f"正在搜索... {completed_files}/{total_files} 个文件")
        
        return match_count
    
    # 使用线程池执行搜索
    # 根据CPU核心数设置线程数，最大为16
    max_workers = min(16, (os.cpu_count() or 4) * 2)
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        futures = {executor.submit(search_single_file, file_path): file_path 
                  for file_path in search_tool.excel_files}
        
        # 等待所有任务完成
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                # 记录错误但继续处理其他文件
                pass
    
    return total_matches


def get_row_data_as_list(file_path, sheet_name, row_num):
    """
    获取指定Excel文件中某一行的完整数据（以列表形式返回）
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        row_num: 行号
    
    Returns:
        该行所有列的数据列表
    """
    try:
        # 读取Excel文件的指定工作表
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # 获取指定行的数据（注意：row_num是1-based，需要转换为0-based）
        if row_num <= len(df):
            row_data = df.iloc[row_num - 1]
            # 将该行数据转换为列表，保留NaN显示为空字符串
            return [str(val) if pd.notna(val) else "" for val in row_data]
        else:
            return ["(行号超出范围)"]
    except Exception as e:
        return [f"(读取失败: {str(e)})"]


def selective_replace(search_tool, search_term, replace_term, selected_replacements, backup_files):
    """
    执行选择性替换（多线程版本）
    
    Args:
        search_tool: ExcelSearchReplace工具实例
        search_term: 搜索词
        replace_term: 替换词
        selected_replacements: 选中的替换项字典
        backup_files: 是否备份文件
    
    Returns:
        (替换的文件数, 总替换次数)
    """
    import openpyxl
    import re
    from datetime import datetime
    import shutil
    
    # 线程锁
    lock = threading.Lock()
    replaced_files = 0
    total_replacements = 0
    
    def replace_single_file(file_path, selection):
        """替换单个文件"""
        nonlocal replaced_files, total_replacements
        
        try:
            if not selection['selected']:
                return
            
            # 创建备份
            if backup_files:
                backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                shutil.copy2(file_path, backup_path)
            
            # 加载工作簿
            wb = openpyxl.load_workbook(file_path)
            file_replaced = False
            file_replacement_count = 0
            
            # 获取该文件的匹配项
            matches = search_tool.search_results[file_path]['matches']
            
            # 如果是全部替换模式
            if selection['rows'] == 'all':
                for match in matches:
                    sheet = wb[match['sheet_name']]
                    cell = sheet.cell(row=match['row'], column=match['column'])
                    
                    if cell.value:
                        # 执行替换
                        new_value = str(cell.value).replace(search_term, replace_term)
                        cell.value = new_value
                        file_replacement_count += 1
                        file_replaced = True
            
            # 如果是选择特定行模式
            else:
                selected_rows = selection['selected_rows']
                for match in matches:
                    row_key = f"{file_path}_{match['sheet_name']}_{match['row']}"
                    if row_key in selected_rows:
                        sheet = wb[match['sheet_name']]
                        cell = sheet.cell(row=match['row'], column=match['column'])
                        
                        if cell.value:
                            # 执行替换
                            new_value = str(cell.value).replace(search_term, replace_term)
                            cell.value = new_value
                            file_replacement_count += 1
                            file_replaced = True
            
            # 保存文件
            if file_replaced:
                wb.save(file_path)
                with lock:
                    replaced_files += 1
                    total_replacements += file_replacement_count
            
            wb.close()
        
        except Exception as e:
            st.error(f"替换文件 {Path(file_path).name} 时出错: {str(e)}")
    
    # 使用线程池执行替换
    max_workers = min(8, (os.cpu_count() or 4))
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(replace_single_file, file_path, selection): file_path 
                  for file_path, selection in selected_replacements.items()}
        
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                pass
    
    return replaced_files, total_replacements


def open_folder(file_path):
    """
    打开文件所在的文件夹
    
    Args:
        file_path: 文件路径
    """
    try:
        folder_path = os.path.dirname(os.path.abspath(file_path))
        system = platform.system()
        
        if system == "Windows":
            # Windows系统
            os.startfile(folder_path)
        elif system == "Darwin":
            # macOS系统
            subprocess.run(["open", folder_path])
        else:
            # Linux系统
            subprocess.run(["xdg-open", folder_path])
        
        st.success(f"✅ 已打开文件夹: {folder_path}")
    except Exception as e:
        st.error(f"❌ 打开文件夹失败: {str(e)}")


def open_file(file_path):
    """
    打开Excel文件
    
    Args:
        file_path: 文件路径
    """
    try:
        system = platform.system()
        
        if system == "Windows":
            # Windows系统
            os.startfile(file_path)
        elif system == "Darwin":
            # macOS系统
            subprocess.run(["open", file_path])
        else:
            # Linux系统
            subprocess.run(["xdg-open", file_path])
        
        st.success(f"✅ 已打开文件: {Path(file_path).name}")
    except Exception as e:
        st.error(f"❌ 打开文件失败: {str(e)}")

# 页面3: 翻译结果处理
def translation_result_processor_page():
    st.title("📊 AI翻译结果处理工具")
    st.markdown("### 将AI返回的翻译结果与原始Excel文件进行匹配，生成包含翻译结果的Excel文件。")
    
    if 'result_translator' not in st.session_state:
        st.session_state.result_translator = MultiAPIExcelTranslator(
            api_key="", 
            api_provider="DeepSeek", 
            api_url=get_api_providers()["DeepSeek"]["url"], 
            model="deepseek-chat"
        )
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.header("📁 原始Excel文件")
        
        original_excel_file = st.file_uploader(
            "📄 上传原始Excel文件",
            type=['xlsx', 'xls'],
            key="result_original_excel"
        )
        
        original_df = None
        text_col = None
        
        if original_excel_file is not None:
            try:
                original_df = pd.read_excel(original_excel_file)
                # 清理列名中的换行符和空白字符
                original_df.columns = original_df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.result_original_df = original_df
                st.success(f"✅ 成功读取原始文件，共 {len(original_df)} 行数据")
                
                with st.expander("📊 原始文件预览"):
                    st.dataframe(original_df.head(10))
                
                cols = original_df.columns.tolist()
                text_col = st.selectbox(
                    "📝 选择文本列（与生成提示词时选择的列相同）",
                    options=cols,
                    index=0,
                    key="result_text_col_select"
                )
                
                st.session_state.result_text_col = text_col
                
            except Exception as e:
                st.error(f"❌ 原始文件读取失败: {e}")
    
    with col2:
        st.header("📝 AI翻译结果")
        
        st.subheader("📋 输入AI翻译结果")
        ai_result_text = st.text_area(
            "AI翻译结果文本",
            value=st.session_state.get('result_ai_result_text', ''),
            height=300,
            placeholder="""请粘贴AI返回的翻译结果文本，支持多种表格格式：

格式1（标准Markdown）：
| 原文（中文） | 英文翻译 |
|-------------|----------|
| 喂！你们干什么！ | Hey! What are you doing! |

格式2（简化格式）：
| 原文（中文） | 英文 Translation |
| 喂！你们干什么！ | Hey! What are you doing! |

格式3（无表头）：
| 喂！你们干什么！ | Hey! What are you doing! |
| 突突突——嗡！—— | Vroom vroom—Broom!— |""",
            key="result_ai_result_text_area"
        )
        st.session_state.result_ai_result_text = ai_result_text
        
        st.subheader("🌍 目标语言")
        target_language = st.selectbox(
            "选择目标语言",
            options=get_preset_languages(),
            index=0,
            key="result_target_language_select"
        )
        
        if target_language == "自定义":
            custom_language = st.text_input(
                "输入自定义语言",
                value=st.session_state.get('result_custom_language', ''),
                placeholder="例如：俄文、葡萄牙文、阿拉伯文等",
                key="result_custom_language_input"
            )
            target_language = custom_language if custom_language else target_language
        
        st.session_state.result_target_language = target_language
        st.info(f"🎯 目标语言: {target_language}")
    
    if st.button("🚀 处理翻译结果", key="process_results_btn", use_container_width=True):
        if original_df is None or text_col is None:
            st.error("❌ 请先上传原始Excel文件并选择文本列。")
            return
        
        if not ai_result_text:
            st.error("❌ 请输入AI翻译结果文本。")
            return
        
        if not target_language:
            st.error("❌ 请先选择目标语言。")
            return
        
        with st.spinner("正在解析AI翻译结果..."):
            translations = parse_ai_translation_result(ai_result_text)
        
        if not translations:
            st.error("❌ 无法解析AI翻译结果，请检查文本格式是否正确。")
            return
        
        st.success(f"✅ 成功解析 {len(translations)} 条翻译结果")
        
        with st.expander("📊 解析结果预览"):
            if translations:
                preview_data = []
                for i, (original, translation) in enumerate(list(translations.items())[:10]):
                    preview_data.append({
                        "序号": i + 1,
                        "原文": original,
                        "翻译结果": translation
                    })
                preview_df = pd.DataFrame(preview_data)
                st.dataframe(preview_df)
            else:
                st.warning("未解析到任何翻译结果")
        
        with st.spinner("正在合并翻译结果..."):
            result_df, matched_count, unmatched_texts = merge_translations_with_excel(
                original_df, text_col, translations, target_language
            )
        
        st.success(f"✅ 成功匹配 {matched_count}/{len(original_df)} 条记录")
        
        if unmatched_texts:
            st.warning(f"⚠️ 有 {len(unmatched_texts)} 条记录未能匹配")
        
        with st.expander("📊 合并结果预览"):
            st.dataframe(result_df.head(10))
        
        st.session_state.result_df = result_df
        st.session_state.matched_count = matched_count
        st.session_state.unmatched_count = len(unmatched_texts)
        st.session_state.original_filename = original_excel_file.name if original_excel_file else "translation_results"
        
        st.success("✅ 翻译结果处理完成！")
    
    if st.session_state.get('result_df') is not None:
        result_df = st.session_state.result_df
        matched_count = st.session_state.matched_count
        unmatched_count = st.session_state.unmatched_count
        target_language = st.session_state.get('result_target_language', '英文')
        original_filename = st.session_state.get('original_filename', 'translation_results')
        
        st.header("📥 下载结果")
        
        if original_filename and original_filename != "translation_results":
            original_name = Path(original_filename).stem
            output_filename = f"{original_name}_{target_language}ori.xlsx"
        else:
            output_filename = f"translation_results_{target_language}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='翻译结果')
        
        excel_buffer.seek(0)
        
        st.download_button(
            label=f"📥 下载翻译结果Excel文件 ({matched_count}/{len(result_df)} 条匹配)",
            data=excel_buffer,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.info(f"**统计信息：** 总记录数: {len(result_df)} | 成功匹配: {matched_count} | 未匹配: {unmatched_count} | 匹配率: {matched_count/len(result_df)*100:.1f}%")
def find_excel_files(folder_path):
    """查找文件夹中的所有Excel文件"""
    excel_files = []
    folder_path = Path(folder_path)
    
    if not folder_path.exists():
        return False, "文件夹路径不存在"
    
    # 支持的Excel文件扩展名
    excel_extensions = ['*.xlsx', '*.xls', '*.xlsm', '*.xlsb']
    
    for ext in excel_extensions:
        excel_files.extend(folder_path.rglob(ext))
    
    return True, excel_files


def excel_comparison_page():
    st.title("🔍 Excel表格对比工具")
    st.markdown("### 比较两个相似Excel表格，找出差异和改动")

    st.info("💡 此功能适用于比较两个版本相似的Excel文件，找出被修改的内容")

    # 文件上传区域
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("📄 原始表格 (版本A)")
        file_a = st.file_uploader(
            "上传原始Excel文件",
            type=['xlsx', 'xls'],
            key="comparison_file_a"
        )

        if file_a is not None:
            try:
                df_a = pd.read_excel(file_a)
                st.success(f"✅ 成功读取文件A: {len(df_a)} 行, {len(df_a.columns)} 列")

                with st.expander("📊 文件A预览"):
                    st.dataframe(df_a.head(10))

            except Exception as e:
                st.error(f"❌ 读取文件A失败: {e}")

    with col2:
        st.subheader("📄 修改后表格 (版本B)")
        file_b = st.file_uploader(
            "上传修改后的Excel文件",
            type=['xlsx', 'xls'],
            key="comparison_file_b"
        )

        if file_b is not None:
            try:
                df_b = pd.read_excel(file_b)
                st.success(f"✅ 成功读取文件B: {len(df_b)} 行, {len(df_b.columns)} 列")

                with st.expander("📊 文件B预览"):
                    st.dataframe(df_b.head(10))

            except Exception as e:
                st.error(f"❌ 读取文件B失败: {e}")

    # 比较设置
    st.markdown("---")
    st.subheader("⚙️ 比较设置")

    col1, col2, col3 = st.columns(3)

    with col1:
        # 选择关键列（用于行匹配）
        key_column = st.text_input(
            "关键列名（用于行匹配）:",
            placeholder="例如: ID、序号等",
            help="用于匹配两个表格中对应行的列名，留空则按行号匹配"
        )

    with col2:
        # 比较模式
        compare_mode = st.selectbox(
            "比较模式:",
            options=["精确匹配", "模糊匹配", "仅比较文本内容"],
            index=0,
            help="精确匹配：完全一致；模糊匹配：允许微小差异；仅比较文本：忽略格式"
        )

    with col3:
        # 敏感度设置
        sensitivity = st.slider(
            "差异敏感度:",
            min_value=1,
            max_value=10,
            value=5,
            help="数值越高，对微小差异越敏感"
        )

    # 高级选项
    with st.expander("🔧 高级选项"):
        col1, col2 = st.columns(2)

        with col1:
            ignore_case = st.checkbox("忽略大小写", value=True)
            ignore_whitespace = st.checkbox("忽略空白字符", value=True)
            show_unchanged = st.checkbox("显示未更改的行", value=False)

        with col2:
            highlight_changes = st.checkbox("高亮显示更改", value=True)
            include_additions = st.checkbox("检测新增行", value=True)
            include_deletions = st.checkbox("检测删除行", value=True)

    # 执行比较
    if st.button("🚀 开始比较", type="primary", use_container_width=True):
        if file_a is None or file_b is None:
            st.error("❌ 请先上传两个Excel文件")
            return

        try:
            # 读取数据
            df_a = pd.read_excel(file_a)
            df_b = pd.read_excel(file_b)

            # 执行比较
            with st.spinner("🔍 正在比较两个表格..."):
                comparison_results = compare_dataframes_simple(
                    df_a, df_b, key_column, compare_mode, sensitivity,
                    ignore_case, ignore_whitespace, include_additions, include_deletions
                )

            # 显示比较结果
            display_comparison_results_simple(
                comparison_results, highlight_changes, show_unchanged
            )

        except Exception as e:
            st.error(f"❌ 比较过程中出错: {e}")
            import traceback
            st.error(traceback.format_exc())


def compare_dataframes_simple(df_a, df_b, key_column=None, compare_mode="精确匹配",
                              sensitivity=5, ignore_case=True, ignore_whitespace=True,
                              include_additions=True, include_deletions=True):
    """
    简化的DataFrame比较函数
    """
    results = {
        'added_rows': [],
        'deleted_rows': [],
        'modified_rows': [],
        'modified_cells': [],
        'summary': {
            'total_rows_a': len(df_a),
            'total_rows_b': len(df_b),
            'added_count': 0,
            'deleted_count': 0,
            'modified_count': 0,
            'similarity_score': 0
        }
    }

    # 预处理数据
    df_a_clean = preprocess_dataframe_simple(df_a, ignore_case, ignore_whitespace)
    df_b_clean = preprocess_dataframe_simple(df_b, ignore_case, ignore_whitespace)

    # 如果有关键列，使用关键列进行行匹配
    if key_column and key_column in df_a.columns and key_column in df_b.columns:
        # 使用关键列匹配行
        a_keys = df_a[key_column].astype(str).tolist()
        b_keys = df_b[key_column].astype(str).tolist()

        # 找出新增和删除的行
        if include_additions:
            for i, key in enumerate(b_keys):
                if key not in a_keys:
                    results['added_rows'].append({
                        'key': key,
                        'row_index_b': i,
                        'row_data': df_b.iloc[i].to_dict()
                    })

        if include_deletions:
            for i, key in enumerate(a_keys):
                if key not in b_keys:
                    results['deleted_rows'].append({
                        'key': key,
                        'row_index_a': i,
                        'row_data': df_a.iloc[i].to_dict()
                    })

        # 比较共同的行
        common_keys = set(a_keys) & set(b_keys)
        for key in common_keys:
            idx_a = a_keys.index(key)
            idx_b = b_keys.index(key)

            row_a = df_a_clean.iloc[idx_a]
            row_b = df_b_clean.iloc[idx_b]

            # 比较行内容
            changes = compare_rows_simple(row_a, row_b, df_a.columns.tolist(),
                                          compare_mode, sensitivity)

            if changes:
                results['modified_rows'].append({
                    'key': key,
                    'row_index_a': idx_a,
                    'row_index_b': idx_b,
                    'row_data_a': df_a.iloc[idx_a].to_dict(),
                    'row_data_b': df_b.iloc[idx_b].to_dict(),
                    'changes': changes,
                    'change_count': len(changes)
                })

                # 记录修改的单元格
                for change in changes:
                    results['modified_cells'].append({
                        'key': key,
                        'row_index_a': idx_a,
                        'row_index_b': idx_b,
                        'column': change['column'],
                        'value_a': change['value_a'],
                        'value_b': change['value_b'],
                        'change_type': change['change_type']
                    })

    else:
        # 没有关键列，按行号匹配
        max_rows = min(len(df_a), len(df_b))

        for i in range(max_rows):
            row_a = df_a_clean.iloc[i]
            row_b = df_b_clean.iloc[i]

            # 比较行内容
            changes = compare_rows_simple(row_a, row_b, df_a.columns.tolist(),
                                          compare_mode, sensitivity)

            if changes:
                results['modified_rows'].append({
                    'key': f"行{i + 1}",
                    'row_index_a': i,
                    'row_index_b': i,
                    'row_data_a': df_a.iloc[i].to_dict(),
                    'row_data_b': df_b.iloc[i].to_dict(),
                    'changes': changes,
                    'change_count': len(changes)
                })

                for change in changes:
                    results['modified_cells'].append({
                        'key': f"行{i + 1}",
                        'row_index_a': i,
                        'row_index_b': i,
                        'column': change['column'],
                        'value_a': change['value_a'],
                        'value_b': change['value_b'],
                        'change_type': change['change_type']
                    })

        # 处理新增/删除的行（按行号）
        if include_additions and len(df_b) > len(df_a):
            for i in range(len(df_a), len(df_b)):
                results['added_rows'].append({
                    'key': f"新增行{i + 1}",
                    'row_index_b': i,
                    'row_data': df_b.iloc[i].to_dict()
                })

        if include_deletions and len(df_a) > len(df_b):
            for i in range(len(df_b), len(df_a)):
                results['deleted_rows'].append({
                    'key': f"删除行{i + 1}",
                    'row_index_a': i,
                    'row_data': df_a.iloc[i].to_dict()
                })

    # 计算统计信息
    results['summary']['added_count'] = len(results['added_rows'])
    results['summary']['deleted_count'] = len(results['deleted_rows'])
    results['summary']['modified_count'] = len(results['modified_rows'])

    # 计算相似度得分
    total_cells = results['summary']['total_rows_a'] * len(df_a.columns) if len(df_a.columns) > 0 else 0
    if total_cells > 0:
        changed_cells = len(results['modified_cells'])
        similarity = 1 - (changed_cells / total_cells)
        results['summary']['similarity_score'] = round(similarity * 100, 2)

    return results


def preprocess_dataframe_simple(df, ignore_case=True, ignore_whitespace=True):
    """
    简化的DataFrame预处理函数
    """
    df_clean = df.copy()

    # 处理NaN值
    df_clean = df_clean.fillna('')

    # 转换为字符串类型进行比较
    for col in df_clean.columns:
        df_clean[col] = df_clean[col].astype(str)

        if ignore_case:
            df_clean[col] = df_clean[col].str.lower()

        if ignore_whitespace:
            df_clean[col] = df_clean[col].str.strip()
            df_clean[col] = df_clean[col].str.replace(r'\s+', ' ', regex=True)

    return df_clean


def compare_rows_simple(row_a, row_b, columns, compare_mode="精确匹配", sensitivity=5):
    """
    简化的行比较函数
    """
    changes = []

    for col in columns:
        val_a = str(row_a[col]) if pd.notna(row_a[col]) else ""
        val_b = str(row_b[col]) if pd.notna(row_b[col]) else ""

        # 空值处理
        if val_a == "" and val_b == "":
            continue

        change_type = "未变化"

        if compare_mode == "精确匹配":
            if val_a != val_b:
                change_type = "修改"
        elif compare_mode == "模糊匹配":
            similarity = calculate_similarity(val_a, val_b)
            threshold = sensitivity / 10.0  # 将敏感度转换为0-1的阈值
            if similarity < threshold:
                change_type = "修改"
        elif compare_mode == "仅比较文本内容":
            # 移除数字和特殊字符，只比较文本内容
            text_a = re.sub(r'[^a-zA-Z\u4e00-\u9fa5]', '', val_a)
            text_b = re.sub(r'[^a-zA-Z\u4e00-\u9fa5]', '', val_b)
            if text_a != text_b:
                change_type = "修改"

        if change_type == "修改":
            changes.append({
                'column': col,
                'value_a': val_a,
                'value_b': val_b,
                'change_type': change_type,
                'similarity': calculate_similarity(val_a, val_b) if compare_mode == "模糊匹配" else None
            })

    return changes


def calculate_similarity(str1, str2):
    """
    计算两个字符串的相似度（0-1）
    """
    if not str1 and not str2:
        return 1.0
    if not str1 or not str2:
        return 0.0

    return difflib.SequenceMatcher(None, str1, str2).ratio()


def display_comparison_results_simple(results, highlight_changes=True, show_unchanged=False):
    """
    简化的比较结果显示函数，避免使用图表功能
    """
    st.markdown("---")
    st.header("📊 比较结果")

    # 显示统计信息
    summary = results['summary']
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("文件A行数", summary['total_rows_a'])
    with col2:
        st.metric("文件B行数", summary['total_rows_b'])
    with col3:
        st.metric("修改行数", summary['modified_count'])
    with col4:
        st.metric("相似度", f"{summary['similarity_score']}%")

    # 显示新增行
    if results['added_rows']:
        st.subheader("🆕 新增行")
        added_data = []
        for row in results['added_rows']:
            row_data = {'关键值': row['key'], '行号(B)': row['row_index_b'] + 1}
            row_data.update(row['row_data'])
            added_data.append(row_data)

        added_df = pd.DataFrame(added_data)
        st.dataframe(added_df, use_container_width=True)

    # 显示删除行
    if results['deleted_rows']:
        st.subheader("🗑️ 删除行")
        deleted_data = []
        for row in results['deleted_rows']:
            row_data = {'关键值': row['key'], '行号(A)': row['row_index_a'] + 1}
            row_data.update(row['row_data'])
            deleted_data.append(row_data)

        deleted_df = pd.DataFrame(deleted_data)
        st.dataframe(deleted_df, use_container_width=True)

    # 显示修改的行
    if results['modified_rows']:
        st.subheader("✏️ 修改的行")

        for mod_row in results['modified_rows']:
            with st.expander(f"🔍 {mod_row['key']} - {mod_row['change_count']} 处修改", expanded=True):
                # 创建对比表格
                comparison_data = []

                # 获取所有列
                all_columns = set(mod_row['row_data_a'].keys()) | set(mod_row['row_data_b'].keys())

                for col in sorted(all_columns):
                    val_a = mod_row['row_data_a'].get(col, '')
                    val_b = mod_row['row_data_b'].get(col, '')

                    # 检查此列是否有修改
                    is_changed = False
                    for change in mod_row['changes']:
                        if change['column'] == col:
                            is_changed = True
                            break

                    if is_changed or show_unchanged:
                        # 准备显示值（高亮修改）
                        if highlight_changes and is_changed:
                            display_a = f"**{val_a}**" if val_a else ""
                            display_b = f"**{val_b}**" if val_b else ""
                        else:
                            display_a = val_a
                            display_b = val_b

                        comparison_data.append({
                            '列名': col,
                            '文件A值': display_a,
                            '文件B值': display_b,
                            '状态': '✅ 未修改' if not is_changed else '❌ 已修改'
                        })

                # 显示对比表格
                comp_df = pd.DataFrame(comparison_data)
                st.dataframe(comp_df, use_container_width=True)

                # 显示详细修改
                st.write("**详细修改:**")
                for change in mod_row['changes']:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.text_area(
                            f"文件A - {change['column']}",
                            value=change['value_a'],
                            height=50,
                            key=f"a_{mod_row['key']}_{change['column']}"
                        )
                    with col2:
                        st.text_area(
                            f"文件B - {change['column']}",
                            value=change['value_b'],
                            height=50,
                            key=f"b_{mod_row['key']}_{change['column']}"
                        )

    # 显示修改统计（使用表格而不是图表）
    if results['modified_cells']:
        st.subheader("📈 修改统计")

        # 按列统计修改次数
        col_changes = {}
        for cell in results['modified_cells']:
            col = cell['column']
            if col not in col_changes:
                col_changes[col] = 0
            col_changes[col] += 1

        # 使用表格显示修改统计，避免使用图表
        if col_changes:
            stat_data = []
            for col, count in col_changes.items():
                stat_data.append({
                    '列名': col,
                    '修改次数': count
                })

            stat_df = pd.DataFrame(stat_data).sort_values('修改次数', ascending=False)
            st.dataframe(stat_df, use_container_width=True)

            # 使用进度条显示修改比例
            st.write("**修改比例:**")
            for _, row in stat_df.iterrows():
                col_name = row['列名']
                count = row['修改次数']
                max_count = max(col_changes.values())
                percentage = (count / max_count) * 100

                st.write(f"{col_name}: {count} 次修改")
                st.progress(int(percentage))

    # 提供结果下载
    st.subheader("💾 下载比较结果")

    # 准备下载数据
    download_data = []
    for cell in results['modified_cells']:
        download_data.append({
            '关键值': cell['key'],
            '行号(A)': cell.get('row_index_a', '') + 1,
            '行号(B)': cell.get('row_index_b', '') + 1,
            '列名': cell['column'],
            '文件A值': cell['value_a'],
            '文件B值': cell['value_b'],
            '修改类型': cell['change_type']
        })

    if download_data:
        download_df = pd.DataFrame(download_data)
        csv_data = download_df.to_csv(index=False).encode('utf-8-sig')

        st.download_button(
            label="📥 下载差异报告(CSV)",
            data=csv_data,
            file_name=f"excel_comparison_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True
        )
    else:
        st.info("📝 没有检测到差异，无需下载报告")
# 全局锁用于线程安全的进度更新
progress_lock = Lock()

def similar(a, b):
    """计算两个字符串的相似度"""
    return SequenceMatcher(None, str(a), str(b)).ratio()

def load_single_file(file_path):
    """加载单个文件（Excel或CSV）"""
    try:
        if file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
            # 读取Excel文件的所有sheet
            excel_file = pd.read_excel(file_path, sheet_name=None)
            results = {}
            for sheet_name, df in excel_file.items():
                if not df.empty:
                    key = f"{file_path.name} - {sheet_name}"
                    results[key] = {
                        'dataframe': df,
                        'file_path': file_path,
                        'sheet_name': sheet_name,
                        'file_type': 'excel'
                    }
            return results
        elif file_path.suffix.lower() == '.csv':
            # 尝试不同的编码格式读取CSV
            encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is not None and not df.empty:
                return {
                    file_path.name: {
                        'dataframe': df,
                        'file_path': file_path,
                        'sheet_name': 'CSV',
                        'file_type': 'csv'
                    }
                }
    except Exception as e:
        st.warning(f"无法读取文件 {file_path}: {str(e)}")
    
    return {}

def load_all_files_parallel(folder_path, max_workers=4):
    """并行加载文件夹中的所有Excel和CSV文件"""
    all_files = {}
    folder_path = Path(folder_path)
    
    # 收集所有文件路径
    file_paths = []
    for pattern in ['*.xlsx', '*.xls', '*.xlsm', '*.csv']:
        file_paths.extend(folder_path.rglob(pattern))
    
    if not file_paths:
        return all_files
    
    # 使用线程池并行加载文件
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        future_to_path = {executor.submit(load_single_file, path): path for path in file_paths}
        
        # 收集结果
        for future in concurrent.futures.as_completed(future_to_path):
            try:
                result = future.result()
                all_files.update(result)
            except Exception as e:
                path = future_to_path[future]
                st.warning(f"处理文件 {path} 时出错: {str(e)}")
    
    return all_files

def load_source_files_parallel(folder_path, max_workers=4):
    """并行加载源文件夹中的Excel和CSV文件"""
    source_files = {}
    folder_path = Path(folder_path)
    
    # 收集所有文件路径
    file_paths = []
    for pattern in ['*.xlsx', '*.xls', '*.xlsm', '*.csv']:
        file_paths.extend(folder_path.glob(pattern))
    
    if not file_paths:
        return source_files
    
    # 使用线程池并行加载文件
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        future_to_path = {executor.submit(load_single_file, path): path for path in file_paths}
        
        # 收集结果
        for future in concurrent.futures.as_completed(future_to_path):
            try:
                result = future.result()
                source_files.update(result)
            except Exception as e:
                path = future_to_path[future]
                st.error(f"处理文件 {path} 时出错: {str(e)}")
    
    return source_files

def find_matching_text(search_text, files_dict, source_col, target_col, match_strategy, similarity_threshold):
    """在文件字典中查找匹配的文本"""
    if pd.isna(search_text) or search_text == '':
        return None, None, 0
        
    search_text = str(search_text).strip()
    
    best_match = None
    best_similarity = 0
    best_source = None
    
    for file_info in files_dict.values():
        df = file_info['dataframe']
        
        # 检查所需的列是否存在
        if source_col in df.columns and target_col in df.columns:
            # 根据匹配策略进行匹配
            if match_strategy == "精确匹配":
                # 精确匹配
                matches = df[df[source_col].astype(str).str.strip() == search_text]
                if not matches.empty:
                    return matches[target_col].iloc[0], matches[source_col].iloc[0], 1.0
            else:
                # 模糊匹配
                for idx, row in df.iterrows():
                    source_text = str(row[source_col])
                    if pd.isna(source_text) or source_text == '':
                        continue
                    
                    if match_strategy == "包含匹配":
                        # 包含匹配
                        if search_text in source_text or source_text in search_text:
                            similarity = similar(search_text, source_text)
                            if similarity > best_similarity:
                                best_similarity = similarity
                                best_match = row[target_col]
                                best_source = source_text
                    else:  # 相似度匹配
                        # 计算相似度
                        similarity = similar(search_text, source_text)
                        if similarity > best_similarity and similarity >= similarity_threshold:
                            best_similarity = similarity
                            best_match = row[target_col]
                            best_source = source_text
    
    # 对于模糊匹配，返回最佳匹配（如果找到）
    if match_strategy != "精确匹配" and best_match is not None:
        return best_match, best_source, best_similarity
    
    return None, None, 0

def process_single_row(args):
    """处理单行数据的匹配"""
    index, row, folder1_match_col, folder1_fill_col, folder2_files, folder2_source_col, folder2_target_col, match_strategy, similarity_threshold, skip_filled = args
    
    # 检查是否需要跳过已填充的行
    if skip_filled and not pd.isna(row.get(folder1_fill_col, None)) and str(row[folder1_fill_col]).strip() != '':
        return index, None, None, None, 0, "跳过已填充"
    
    search_text = row[folder1_match_col]
    
    # 跳过空值
    if pd.isna(search_text) or search_text == '':
        return index, None, None, None, 0, "空值"
    
    # 查找匹配
    matched_text, matched_source, similarity = find_matching_text(
        search_text, folder2_files, folder2_source_col, folder2_target_col, match_strategy, similarity_threshold
    )
    
    match_status = "匹配成功" if matched_text is not None else "未匹配"
    
    return index, matched_text, matched_source, search_text, similarity, match_status

def process_single_file(args):
    """处理单个文件的匹配"""
    filename, file_info, folder1_match_col, folder1_fill_col, folder2_files, folder2_source_col, folder2_target_col, match_strategy, similarity_threshold, skip_filled, thread_id = args
    
    df = file_info['dataframe'].copy()
    file_matches = 0
    file_total = 0
    file_skipped = 0
    
    # 检查必要的列是否存在
    if folder1_match_col not in df.columns:
        return filename, None, {"error": f"文件 {filename} 中找不到列 '{folder1_match_col}'"}
        
    if folder1_fill_col not in df.columns:
        return filename, None, {"error": f"文件 {filename} 中找不到列 '{folder1_fill_col}'"}
    
    # 准备处理数据
    rows_to_process = []
    for index, row in df.iterrows():
        file_total += 1
        rows_to_process.append((index, row, folder1_match_col, folder1_fill_col, folder2_files, folder2_source_col, folder2_target_col, match_strategy, similarity_threshold, skip_filled))
    
    # 并行处理行数据
    matched_results = {}
    match_details = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        future_to_index = {executor.submit(process_single_row, args): args[0] for args in rows_to_process}
        
        for future in concurrent.futures.as_completed(future_to_index):
            try:
                index, matched_text, matched_source, search_text, similarity, match_status = future.result()
                
                # 记录匹配详情
                match_details.append({
                    'index': index,
                    'search_text': search_text,
                    'matched_text': matched_text,
                    'matched_source': matched_source,
                    'similarity': similarity,
                    'status': match_status
                })
                
                if match_status == "跳过已填充":
                    file_skipped += 1
                elif matched_text is not None:
                    matched_results[index] = matched_text
                    file_matches += 1
            except Exception as e:
                index = future_to_index[future]
                st.warning(f"处理文件 {filename} 的第 {index} 行时出错: {str(e)}")
    
    # 应用匹配结果
    for index, matched_text in matched_results.items():
        df.at[index, folder1_fill_col] = matched_text
    
    # 生成报告
    report = {
        'total_rows': file_total,
        'matched_rows': file_matches,
        'unmatched_rows': file_total - file_matches - file_skipped,
        'skipped_rows': file_skipped,
        'match_details': match_details
    }
    
    return filename, df, report

def process_file_matching_parallel(folder1_path, folder2_path, folder1_match_col, folder1_fill_col, 
                                  folder2_source_col, folder2_target_col, match_strategy, 
                                  similarity_threshold, skip_filled, max_workers=4):
    """并行处理文件匹配"""
    
    # 加载文件夹1中的文件
    st.info("正在加载第一个文件夹中的文件...")
    folder1_files = load_source_files_parallel(folder1_path, max_workers)
    
    if not folder1_files:
        st.error("在第一个文件夹中未找到Excel或CSV文件")
        return None, None
    
    # 加载文件夹2中的文件
    st.info("正在加载第二个文件夹中的文件...")
    folder2_files = load_all_files_parallel(folder2_path, max_workers)
    
    if not folder2_files:
        st.error("在第二个文件夹中未找到Excel或CSV文件")
        return None, None
    
    # 显示找到的文件信息
    st.success(f"在第一个文件夹中找到 {len(folder1_files)} 个文件")
    st.success(f"在第二个文件夹中找到 {len(folder2_files)} 个文件")
    
    # 处理匹配
    st.info("开始匹配处理...")
    results = {}
    match_report = {
        'total_files': len(folder1_files),
        'total_rows': 0,
        'matched_rows': 0,
        'unmatched_rows': 0,
        'skipped_rows': 0,
        'file_details': {}
    }
    
    # 准备处理数据
    files_to_process = []
    for i, (filename, file_info) in enumerate(folder1_files.items()):
        files_to_process.append((
            filename, file_info, folder1_match_col, folder1_fill_col,
            folder2_files, folder2_source_col, folder2_target_col, 
            match_strategy, similarity_threshold, skip_filled, i % max_workers
        ))
    
    # 并行处理文件
    progress_bar = st.progress(0)
    processed_count = 0
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        future_to_filename = {executor.submit(process_single_file, args): args[0] for args in files_to_process}
        
        # 收集结果
        for future in concurrent.futures.as_completed(future_to_filename):
            try:
                filename, processed_df, report = future.result()
                if processed_df is not None:
                    results[filename] = processed_df
                    match_report['file_details'][filename] = report
                    match_report['total_rows'] += report['total_rows']
                    match_report['matched_rows'] += report['matched_rows']
                    match_report['unmatched_rows'] += report['unmatched_rows']
                    match_report['skipped_rows'] += report['skipped_rows']
                elif "error" in report:
                    st.error(report["error"])
            except Exception as e:
                filename = future_to_filename[future]
                st.error(f"处理文件 {filename} 时出错: {str(e)}")
            
            # 更新进度
            with progress_lock:
                processed_count += 1
                progress_bar.progress(processed_count / len(files_to_process))
    
    progress_bar.empty()
    return results, match_report

def save_processed_files(processed_files):
    """保存处理后的文件到ZIP包"""
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for filename, df in processed_files.items():
            # 根据原文件类型保存
            if filename.lower().endswith('.csv'):
                # 保存为CSV
                csv_buffer = BytesIO()
                df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
                csv_buffer.seek(0)
                zip_file.writestr(filename, csv_buffer.getvalue())
            else:
                # 保存为Excel
                excel_buffer = BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                excel_buffer.seek(0)
                zip_file.writestr(filename, excel_buffer.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer

def excel_matchpro_page():
    st.set_page_config(
        page_title="Excel/CSV文件匹配工具(增强版)",
        page_icon="⚡",
        layout="wide"
    )
    
    st.title("⚡ Excel/CSV文件匹配工具(增强版)")
    st.markdown("""
    这个工具使用多线程技术加速处理，可以快速匹配两个文件夹中的Excel和CSV文件内容。  
    支持精确匹配、包含匹配和相似度匹配，可跳过已有翻译文本的行。
    """)
    
    # 侧边栏配置
    st.sidebar.header("配置参数")
    
    with st.sidebar.expander("文件夹设置", expanded=True):
        folder1_path = st.text_input("第一个文件夹路径（固定格式文件）", value="./folder1")
        folder2_path = st.text_input("第二个文件夹路径（翻译文件）", value="./folder2")
        max_workers = st.slider("线程数", min_value=1, max_value=16, value=4, step=1)
    
    with st.sidebar.expander("匹配策略设置", expanded=True):
        match_strategy = st.selectbox(
            "匹配策略",
            ["精确匹配", "包含匹配", "相似度匹配"],
            help="精确匹配: 完全相同的文本; 包含匹配: 文本互相包含; 相似度匹配: 基于文本相似度"
        )
        
        similarity_threshold = st.slider(
            "相似度阈值(仅对相似度匹配有效)",
            min_value=0.1,
            max_value=1.0,
            value=0.8,
            step=0.05,
            help="相似度高于此阈值的文本将被视为匹配"
        )
        
        skip_filled = st.checkbox(
            "跳过已有翻译文本的行",
            value=True,
            help="如果目标列已有内容，则跳过该行不进行匹配"
        )
    
    with st.sidebar.expander("列映射设置", expanded=True):
        st.markdown("**第一个文件夹列设置**")
        folder1_match_col = st.text_input("匹配列名（用于查找的列）", value="中文文本")
        folder1_fill_col = st.text_input("填充列名（要填入翻译的列）", value="英文文本")
        
        st.markdown("**第二个文件夹列设置**")
        folder2_source_col = st.text_input("原文列名", value="原文")
        folder2_target_col = st.text_input("翻译列名", value="翻译结果")
    
    # 主界面
    col1, col2 = st.columns([2, 1])
    
    with col1:
        if st.button("开始处理", type="primary", use_container_width=True):
            if not folder1_path or not folder2_path:
                st.error("请填写两个文件夹的路径")
                return
                
            if not os.path.exists(folder1_path):
                st.error(f"第一个文件夹路径不存在: {folder1_path}")
                return
                
            if not os.path.exists(folder2_path):
                st.error(f"第二个文件夹路径不存在: {folder2_path}")
                return
            
            # 处理匹配
            start_time = time.time()
            with st.spinner("正在处理文件匹配..."):
                processed_files, match_report = process_file_matching_parallel(
                    folder1_path, folder2_path, 
                    folder1_match_col, folder1_fill_col,
                    folder2_source_col, folder2_target_col,
                    match_strategy, similarity_threshold, skip_filled, max_workers
                )
            
            end_time = time.time()
            
            if processed_files is not None:
                # 显示报告
                st.success(f"处理完成！耗时: {end_time - start_time:.2f} 秒")
                
                # 汇总报告
                st.subheader("匹配报告")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("总文件数", match_report['total_files'])
                    st.metric("总行数", match_report['total_rows'])
                
                with col2:
                    st.metric("匹配行数", match_report['matched_rows'])
                    match_rate = (match_report['matched_rows'] / (match_report['total_rows'] - match_report['skipped_rows']) * 100) if (match_report['total_rows'] - match_report['skipped_rows']) > 0 else 0
                    st.metric("匹配率", f"{match_rate:.1f}%")
                
                with col3:
                    st.metric("未匹配行数", match_report['unmatched_rows'])
                    st.metric("跳过的行数", match_report['skipped_rows'])
                
                with col4:
                    st.metric("线程数", max_workers)
                    st.metric("处理时间", f"{end_time - start_time:.2f}s")
                
                # 详细报告
                with st.expander("详细文件报告"):
                    for filename, details in match_report['file_details'].items():
                        if 'error' not in details:
                            file_match_rate = (details['matched_rows'] / (details['total_rows'] - details['skipped_rows']) * 100) if (details['total_rows'] - details['skipped_rows']) > 0 else 0
                            st.write(f"**{filename}**: {details['matched_rows']}/{details['total_rows']} 匹配 ({file_match_rate:.1f}%)，跳过 {details['skipped_rows']} 行")
                
                # 匹配详情表格
                with st.expander("匹配详情"):
                    all_match_details = []
                    for filename, details in match_report['file_details'].items():
                        if 'error' not in details and 'match_details' in details:
                            for match_detail in details['match_details']:
                                match_detail['filename'] = filename
                                all_match_details.append(match_detail)
                    
                    if all_match_details:
                        details_df = pd.DataFrame(all_match_details)
                        # 只显示部分列以保持简洁
                        display_df = details_df[['filename', 'search_text', 'matched_source', 'similarity', 'status']].copy()
                        display_df = display_df.rename(columns={
                            'filename': '文件名',
                            'search_text': '搜索文本',
                            'matched_source': '匹配到的原文',
                            'similarity': '相似度',
                            'status': '状态'
                        })
                        st.dataframe(display_df)
                    else:
                        st.info("无匹配详情数据")
                
                # 预览处理后的数据
                with st.expander("预览处理后的数据"):
                    selected_file = st.selectbox("选择要预览的文件", list(processed_files.keys()))
                    if selected_file:
                        st.dataframe(processed_files[selected_file].head(10))
                
                # 下载处理后的文件
                st.subheader("下载处理结果")
                zip_buffer = save_processed_files(processed_files)
                
                st.download_button(
                    label="📥 下载所有处理后的文件 (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name="processed_files.zip",
                    mime="application/zip",
                    use_container_width=True
                )
    
    with col2:
        st.markdown("### 使用说明")
        st.markdown("""
        1. **设置文件夹路径**: 输入两个文件夹的完整路径
        2. **配置匹配策略**: 选择适合的匹配方式和参数
        3. **配置列映射**: 设置源文件和目标文件的列名对应关系
        4. **调整线程数**: 根据CPU核心数调整线程数以获得最佳性能
        5. **开始处理**: 点击按钮开始匹配过程
        6. **下载结果**: 处理完成后下载ZIP包
        """)
        
        st.markdown("### 匹配策略说明")
        st.markdown("""
        - **精确匹配**: 完全相同的文本才会匹配
        - **包含匹配**: 文本互相包含即可匹配(如"关卡1"和"关卡")
        - **相似度匹配**: 基于文本相似度算法，可处理错别字
        """)
        
        st.markdown("### 支持的文件格式")
        st.markdown("""
        - Excel: .xlsx, .xls, .xlsm
        - CSV: .csv
        """)
        
        st.markdown("### 性能优化")
        st.markdown("""
        - 使用多线程并行处理文件
        - 跳过已有翻译文本的行
        - 支持模糊匹配和相似度计算
        - 实时进度显示和详细报告
        """)
def search_in_dataframe(df, col_name, target_values, keyword, case_sensitive=False, match_whole_word=False):
    """在DataFrame中搜索满足条件的行"""
    matches = []
    
    # 检查列是否存在
    if col_name not in df.columns:
        return matches, f"列 '{col_name}' 不存在"
    
    # 第一步：过滤出指定列包含目标值的行
    # 确保比较时使用字符串类型
    filtered_df = df[df[col_name].astype(str).isin([str(v) for v in target_values])]
    
    if len(filtered_df) == 0:
        return matches, "未找到包含指定目标值的行"
    
    # 如果没有提供关键词，返回所有匹配行
    if not keyword or not keyword.strip():
        for idx, row in filtered_df.iterrows():
            matches.append({
                'row_index': idx,
                'row_data': row.to_dict(),
                'matched_column': col_name,
                'matched_value': row[col_name],
                'keyword_found': False,
                'keyword_matches': [],
                'match_count': 0
            })
        return matches, f"找到 {len(matches)} 行包含目标值，但未搜索关键词"
    
    # 第二步：在过滤后的行中搜索关键词
    keyword = keyword.strip()
    flags = 0 if case_sensitive else re.IGNORECASE
    
    # 构建搜索模式
    if match_whole_word:
        pattern = r'\b' + re.escape(keyword) + r'\b'
    else:
        pattern = re.escape(keyword)
    
    keyword_matches = 0
    
    for idx, row in filtered_df.iterrows():
        row_matches = []
        
        # 搜索行的每一列
        for col_idx, (col, cell_value) in enumerate(row.items()):
            if pd.isna(cell_value):
                continue
                
            cell_str = str(cell_value)
            cell_matches = list(re.finditer(pattern, cell_str, flags))
            
            for match in cell_matches:
                row_matches.append({
                    'column': col,
                    'original_value': cell_str,
                    'match_text': match.group(),
                    'start_pos': match.start(),
                    'end_pos': match.end(),
                    'replaced_value': cell_str[:match.start()] + f"**[{match.group()}]**" + cell_str[match.end():]
                })
        
        if row_matches:
            keyword_matches += 1
            matches.append({
                'row_index': idx,
                'row_data': row.to_dict(),
                'matched_column': col_name,
                'matched_value': row[col_name],
                'keyword_found': True,
                'keyword_matches': row_matches,
                'match_count': len(row_matches)
            })
    
    return matches, f"在 {len(filtered_df)} 行目标行中找到 {keyword_matches} 行包含关键词"
def highlight_keyword(text, keyword, case_sensitive=False):
    """高亮显示关键词"""
    if not text or not keyword:
        return text
    
    flags = 0 if case_sensitive else re.IGNORECASE
    pattern = re.escape(keyword)
    
    if case_sensitive:
        highlighted = re.sub(f'({pattern})', r'**\1**', text)
    else:
        highlighted = re.sub(f'({pattern})', r'**\1**', text, flags=re.IGNORECASE)
    
    return highlighted

def replace_in_excel(file_path, replacements, backup=True):
    """在Excel文件中执行替换操作"""
    try:
        file_path = Path(file_path)
        
        # 备份原文件
        if backup:
            backup_path = file_path.parent / f"{file_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_path.suffix}"
            shutil.copy2(file_path, backup_path)
            st.info(f"📁 已创建备份文件: {backup_path.name}")
        
        # 读取Excel文件
        df = pd.read_excel(file_path, engine='openpyxl')
        
        # 执行替换
        replaced_count = 0
        for replacement in replacements:
            row_idx = replacement['row_index']
            col_name = replacement['column']
            old_text = replacement['original_value']
            new_text = replacement['new_value']
            
            if row_idx < len(df) and col_name in df.columns:
                # 获取当前单元格值
                current_value = df.at[row_idx, col_name]
                if pd.isna(current_value):
                    current_value = ""
                
                current_str = str(current_value)
                
                # 执行替换
                if replacement.get('replace_all', False):
                    # 替换所有出现的关键词
                    flags = 0 if replacement.get('case_sensitive', False) else re.IGNORECASE
                    pattern = re.escape(replacement['search_keyword'])
                    if replacement.get('match_whole_word', False):
                        pattern = r'\b' + pattern + r'\b'
                    
                    new_value = re.sub(pattern, replacement['replace_keyword'], current_str, flags=flags)
                else:
                    # 替换特定位置的匹配
                    start_pos = replacement['start_pos']
                    end_pos = replacement['end_pos']
                    new_value = current_str[:start_pos] + replacement['replace_keyword'] + current_str[end_pos:]
                
                df.at[row_idx, col_name] = new_value
                replaced_count += 1
        
        # 保存文件
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        return True, f"✅ 成功替换 {replaced_count} 处内容"
        
    except Exception as e:
        return False, f"❌ 替换失败: {str(e)}"
import tempfile


# ============================================================================
# 工具函数 - Niconico
# ============================================================================

def find_yt_dlp():
    """查找yt-dlp可执行文件"""
    if sys.platform.startswith('win'):
        candidates = ["yt-dlp.exe", "yt-dlp"]
    else:
        candidates = ["./yt-dlp", "yt-dlp"]
    
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return "yt-dlp"  # 依赖于PATH

def normalize_url(url):
    """将非标准的Niconico URL格式转换为标准格式"""
    return url.replace("www.video.nicovideo.jp", "www.nicovideo.jp")

def extract_watch_id(url):
    """从Niconico视频URL中提取watch ID (sm/nm号)"""
    match = re.search(r'(sm|nm)\d+', url)
    if match:
        return match.group(0)
    return "unknown_id"

def extract_bilibili_id(url):
    """从Bilibili视频URL中提取video ID (BV号或av号)"""
    # 尝试提取BV号
    bv_match = re.search(r'BV[a-zA-Z0-9]+', url)
    if bv_match:
        return bv_match.group(0)
    
    # 尝试提取av号
    av_match = re.search(r'av(\d+)', url)
    if av_match:
        return f"av{av_match.group(1)}"
    
    return "unknown_id"

def run_yt_dlp_to_get_json(url, output_filename_base="danmaku"):
    """运行yt-dlp命令来抓取弹幕数据并保存为JSON文件"""
    yt_dlp_path = find_yt_dlp()
    
    command = [
        yt_dlp_path,
        "--skip-download",
        "--write-sub",
        "--all-subs",
        "--sub-format", "json",
        "--output", f"{output_filename_base}.%(ext)s",
        url
    ]
    
    try:
        result = subprocess.run(command, capture_output=True, text=True, check=True)
        
        json_filename = f"{output_filename_base}.comments.json"
        if os.path.exists(json_filename):
            return json_filename
        else:
            return None
            
    except subprocess.CalledProcessError as e:
        st.error(f"yt-dlp执行失败: {e.stderr}")
        return None
    except FileNotFoundError:
        st.error(f"找不到yt-dlp可执行文件。请确保yt-dlp已安装或在PATH中。")
        return None

def process_niconico_json_to_dataframe(json_path):
    """读取yt-dlp生成的JSON文件，处理Niconico弹幕数据"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        st.error(f"JSON文件处理失败: {e}")
        return None

    danmaku_list = []
    for comment in data:
        vpos_ms = comment.get("vposMs", 0)
        time_sec = vpos_ms / 1000
        video_time = time.strftime('%H:%M:%S', time.gmtime(time_sec))
        
        posted_at_str = comment.get("postedAt")
        try:
            posted_at = datetime.fromisoformat(posted_at_str)
            send_time = posted_at.strftime('%Y-%m-%d %H:%M:%S')
        except:
            send_time = posted_at_str
            
        commands = " ".join(comment.get("commands", []))
        
        danmaku_info = {
            "弹幕内容": comment.get("body"),
            "视频时间": video_time,
            "时间(秒)": time_sec,
            "格式/颜色": commands,
            "用户ID": comment.get("userId"),
            "发送时间": send_time,
            "编号": comment.get("no"),
        }
        danmaku_list.append(danmaku_info)
        
    if not danmaku_list:
        return None
        
    df = pd.DataFrame(danmaku_list)
    df = df[['编号', '视频时间', '时间(秒)', '弹幕内容', '格式/颜色', '用户ID', '发送时间']]
    return df

def scrape_niconico_danmaku(url):
    """抓取Niconico弹幕"""
    normalized_url = normalize_url(url)
    watch_id = extract_watch_id(normalized_url)
    
    with st.spinner(f"正在抓取Niconico视频 {watch_id} 的弹幕..."):
        json_path = run_yt_dlp_to_get_json(normalized_url, output_filename_base=watch_id)
        
        if json_path:
            df = process_niconico_json_to_dataframe(json_path)
            
            # 清理临时文件
            try:
                os.remove(json_path)
            except OSError:
                pass
            
            return df, watch_id
        else:
            return None, watch_id

# ============================================================================
# 工具函数 - Bilibili
# ============================================================================

def scrape_bilibili_danmaku(url, cookies_file=None):
    """抓取Bilibili弹幕"""
    video_id = extract_bilibili_id(url)
    
    with st.spinner(f"正在抓取Bilibili视频 {video_id} 的弹幕..."):
        # 使用yt-dlp抓取Bilibili弹幕
        yt_dlp_path = find_yt_dlp()
        
        command = [
            yt_dlp_path,
            "--skip-download",
            "--write-sub",
            "--all-subs",
            "--sub-format", "json",
            "--output", f"{video_id}.%(ext)s",
        ]
        
        # 如果提供了Cookie文件，添加到命令中
        if cookies_file and os.path.exists(cookies_file):
            command.extend(["--cookies", cookies_file])
        
        command.append(url)
        
        try:
            result = subprocess.run(command, capture_output=True, text=True, check=True, timeout=60)
            
            json_filename = f"{video_id}.comments.json"
            if os.path.exists(json_filename):
                try:
                    with open(json_filename, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                except (FileNotFoundError, json.JSONDecodeError) as e:
                    st.error(f"JSON文件处理失败: {e}")
                    return None, video_id
                
                danmaku_list = []
                for comment in data:
                    # Bilibili的弹幕格式与Niconico略有不同
                    danmaku_info = {
                        "弹幕内容": comment.get("body", comment.get("text", "")),
                        "发送时间": comment.get("postedAt", comment.get("timestamp", "")),
                        "用户ID": comment.get("userId", comment.get("author", "")),
                    }
                    danmaku_list.append(danmaku_info)
                
                if danmaku_list:
                    df = pd.DataFrame(danmaku_list)
                    
                    # 清理临时文件
                    try:
                        os.remove(json_filename)
                    except OSError:
                        pass
                    
                    return df, video_id
                else:
                    st.warning("未找到弹幕数据")
                    return None, video_id
            else:
                st.error(f"yt-dlp未生成预期的文件")
                return None, video_id
                
        except subprocess.CalledProcessError as e:
            error_msg = e.stderr if e.stderr else str(e)
            if "412" in error_msg or "Precondition Failed" in error_msg:
                st.error(
                    "❌ Bilibili API速率限制（HTTP 412错误）\n\n"
                    "这通常是因为：\n"
                    "1. 没有提供有效的Cookie认证\n"
                    "2. 该IP地址的请求已达到限制\n\n"
                    "**解决方案：**\n"
                    "请在左侧栏上传您的Bilibili Cookie文件，然后重试。"
                )
            else:
                st.error(f"yt-dlp执行失败: {error_msg}")
            return None, video_id
        except subprocess.TimeoutExpired:
            st.error("❌ 请求超时，请检查网络连接或稍后重试")
            return None, video_id
        except FileNotFoundError:
            st.error(f"找不到yt-dlp可执行文件")
            return None, video_id

# ============================================================================
# UI布局
# ============================================================================

def danmu_page():
    # 标题和描述
    st.markdown("""
    # 🎬 弹幕抓取工具
    
    支持从 **Niconico** 和 **Bilibili** 抓取视频弹幕，并导出为 Excel 文件。
    """)
    # 侧边栏配置
    with st.sidebar:
        st.header("⚙️ 配置1")
        platform = st.radio(
            "选择视频平台",
            options=["Niconico", "Bilibili"],
            help="选择您要抓取弹幕的视频平台",
            key="video_pla_selector"
        )
        
        st.divider()
        
        # Bilibili Cookie配置
        bilibili_cookies_file = None
        if platform == "Bilibili":
            st.subheader("🔐 Bilibili Cookie配置")
            st.markdown(
                """Bilibili需要Cookie认证以避免速率限制。\n\n
                **获取Cookie的方法：**
                1. 打开浏览器访问 https://www.bilibili.com
                2. 登录您的账号
                3. 按F12打开开发者工具 → Application → Cookies
                4. 复制所有Cookie内容到文本文件
                5. 上传该文件
                """
            )
            
            uploaded_file = st.file_uploader(
                "上传Cookie文件",
                type=["txt"],
                help="上传从浏览器导出的Cookie文件"
            )
            
            if uploaded_file is not None:
                # 保存上传的Cookie文件
                cookies_content = uploaded_file.read().decode('utf-8')
                bilibili_cookies_file = "temp_cookies.txt"
                with open(bilibili_cookies_file, 'w', encoding='utf-8') as f:
                    f.write(cookies_content)
                st.success("✅ Cookie文件已加载")
            else:
                st.warning("⚠️ 未上传Cookie文件，可能导致速率限制错误")
        
        st.divider()
        
        st.markdown("""
        ### 📌 使用说明
        
        **Niconico:**
        - 输入格式: `https://www.nicovideo.jp/watch/sm500873`
        - 或: `https://www.video.nicovideo.jp/watch/sm500873` (会自动转换)
        
        **Bilibili:**
        - 输入格式: `https://www.bilibili.com/video/BV1xx411c7mD`
        - 或: `https://www.bilibili.com/video/av123456789`
        - 建议上传Cookie文件以避免速率限制
        """)
    
    # 主内容区域
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader(f"输入{platform}视频链接")
        video_url = st.text_input(
            "视频链接",
            placeholder=f"请输入{platform}视频链接...",
            label_visibility="collapsed"
        )
    
    with col2:
        st.subheader("操作")
        scrape_button = st.button(
            "🔍 开始抓取",
            use_container_width=True,
            type="primary"
        )
    
    st.divider()
    
    # 处理抓取请求
    if scrape_button:
        if not video_url.strip():
            st.error("❌ 请输入视频链接")
        else:
            if platform == "Niconico":
                df, video_id = scrape_niconico_danmaku(video_url)
            else:  # Bilibili
                df, video_id = scrape_bilibili_danmaku(video_url, cookies_file=bilibili_cookies_file)
            
            if df is not None and len(df) > 0:
                st.success(f"✅ 成功抓取 {len(df)} 条弹幕！")
                
                # 显示数据预览
                st.subheader("📊 弹幕数据预览")
                st.dataframe(df, use_container_width=True, height=400)
                
                # 导出选项
                st.subheader("💾 导出选项")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    # Excel导出
                    excel_buffer = pd.ExcelWriter(
                        f"danmaku_{video_id}.xlsx",
                        engine='openpyxl'
                    )
                    df.to_excel(excel_buffer, index=False, sheet_name='弹幕数据')
                    excel_buffer.close()
                    
                    with open(f"danmaku_{video_id}.xlsx", 'rb') as f:
                        st.download_button(
                            label="📥 下载 Excel",
                            data=f.read(),
                            file_name=f"danmaku_{video_id}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    # 清理临时文件
                    try:
                        os.remove(f"danmaku_{video_id}.xlsx")
                    except OSError:
                        pass
                    
                    # 清理临时Cookie文件
                    if bilibili_cookies_file and os.path.exists(bilibili_cookies_file):
                        try:
                            os.remove(bilibili_cookies_file)
                        except OSError:
                            pass
                
                with col2:
                    # CSV导出
                    csv_buffer = df.to_csv(index=False)
                    st.download_button(
                        label="📥 下载 CSV",
                        data=csv_buffer,
                        file_name=f"danmaku_{video_id}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                
                with col3:
                    # JSON导出
                    json_buffer = df.to_json(orient='records', force_ascii=False)
                    st.download_button(
                        label="📥 下载 JSON",
                        data=json_buffer,
                        file_name=f"danmaku_{video_id}.json",
                        mime="application/json",
                        use_container_width=True
                    )
                
                # 统计信息
                st.subheader("📈 统计信息")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("总弹幕数", len(df))
                
                with col2:
                    if '用户ID' in df.columns:
                        unique_users = df['用户ID'].nunique()
                        st.metric("独立用户数", unique_users)
                
                with col3:
                    if '弹幕内容' in df.columns:
                        avg_length = df['弹幕内容'].str.len().mean()
                        st.metric("平均弹幕长度", f"{avg_length:.1f} 字符")
            
            elif df is not None and len(df) == 0:
                st.warning("⚠️ 未找到弹幕数据，请检查视频链接是否正确")
            else:
                st.error("❌ 抓取失败，请检查视频链接或网络连接")
    
    # 页脚
    st.divider()
    st.markdown("""
    ---
    **弹幕抓取工具** | 基于 Streamlit 和 yt-dlp
    
    💡 **提示:**
    - 某些视频的弹幕可能需要登录才能访问
    - 如果遇到问题，请检查您的网络连接
    - 支持的视频平台：Niconico、Bilibili
    """)

def excel_sreplace_page():
    # 创建侧边栏用于输入参数
    st.sidebar.header("🔧 搜索参数设置")
    
    # 获取文件夹路径
    folder_path = st.sidebar.text_input(
        "📁 请输入文件夹路径:",
        placeholder="例如: C:/Users/用户名/Documents/Excel文件",
        help="请输入包含Excel文件的文件夹完整路径"
    )
    
    # 搜索参数设置
    col_name = st.sidebar.text_input(
        "📊 要搜索的列名:",
        value="角色名",
        placeholder="例如: 角色名",
        help="请输入要搜索的Excel列名称"
    )
    
    # 目标值输入（支持多个值）
    target_values_input = st.sidebar.text_input(
        "🎯 列目标值（用逗号分隔）:",
        value="班长,班长大人",
        placeholder="例如: 班长,班长大人",
        help="请输入要在指定列中查找的值，多个值用逗号分隔"
    )
    
    # 要搜索的关键词
    search_keyword = st.sidebar.text_input(
        "🔤 要查找的关键词YYY:",
        value="私",
        placeholder="请输入要在行中查找的关键词",
        help="在满足条件的行中查找此关键词"
    )
    
    # 高级选项
    st.sidebar.markdown("---")
    st.sidebar.subheader("⚙️ 高级选项")
    
    case_sensitive = st.sidebar.checkbox(
        "区分大小写",
        value=False,
        help="勾选后搜索时区分英文大小写"
    )
    
    match_whole_word = st.sidebar.checkbox(
        "全词匹配",
        value=False,
        help="勾选后只匹配完整的词语"
    )
    
    # 处理目标值（分割逗号分隔的值）
    target_values = [v.strip() for v in target_values_input.split(',') if v.strip()]
    
    # 检查输入参数
    if not folder_path:
        st.warning("⚠️ 请输入文件夹路径")
        return
    
    if not col_name:
        st.warning("⚠️ 请输入要搜索的列名")
        return
    
    if not target_values:
        st.warning("⚠️ 请输入至少一个目标值")
        return
    
    # 查找Excel文件
    success, result = find_excel_files(folder_path)
    
    if not success:
        st.error(f"❌ {result}")
        return
    
    excel_files = result
    
    if not excel_files:
        st.warning("⚠️ 在指定文件夹中未找到Excel文件")
        return
    
    st.success(f"✅ 找到 {len(excel_files)} 个Excel文件")
    
    # 显示找到的文件列表
    with st.expander("📁 找到的Excel文件"):
        for i, file_path in enumerate(excel_files[:10]):  # 只显示前10个
            st.write(f"{i+1}. {file_path.name}")
        
        if len(excel_files) > 10:
            st.info(f"... 还有 {len(excel_files) - 10} 个文件")
    
    # 执行搜索
    if st.button("🚀 开始搜索", type="primary", use_container_width=True):
        if not search_keyword.strip():
            st.warning("⚠️ 关键词YYY为空，将只显示包含目标值的行")
        
        all_matches = []
        files_with_matches = 0
        total_matches = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, file_path in enumerate(excel_files):
            progress = (i + 1) / len(excel_files)
            progress_bar.progress(progress)
            status_text.text(f"🔍 正在处理文件 {i+1}/{len(excel_files)}: {file_path.name}")
            
            try:
                # 读取Excel文件
                df = pd.read_excel(file_path, engine='openpyxl')
                
                # 搜索数据
                matches, message = search_in_dataframe(
                    df, col_name, target_values, search_keyword, 
                    case_sensitive, match_whole_word
                )
                
                if matches:
                    files_with_matches += 1
                    total_matches += len(matches)
                    
                    for match in matches:
                        match['file_path'] = str(file_path)
                        match['file_name'] = file_path.name
                        all_matches.append(match)
                
                # 短暂延迟以便显示进度
                import time
                time.sleep(0.1)
                
            except Exception as e:
                st.error(f"处理文件 {file_path.name} 时出错: {e}")
        
        progress_bar.progress(1.0)
        status_text.text(f"✅ 搜索完成！")
        
        # 保存搜索结果到session state
        st.session_state.search_results = all_matches
        st.session_state.search_keyword = search_keyword
        st.session_state.case_sensitive = case_sensitive
        st.session_state.match_whole_word = match_whole_word
        
        # 显示搜索结果
        st.header("📊 搜索结果")
        
        if total_matches == 0:
            st.warning("⚠️ 未找到满足条件的行")
            
            # 显示搜索统计
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("扫描文件数", len(excel_files))
            with col2:
                st.metric("包含匹配的文件", files_with_matches)
            with col3:
                st.metric("总匹配行数", total_matches)
        else:
            st.success(f"✅ 在 {files_with_matches} 个文件中找到 {total_matches} 行匹配结果")
            
            # 显示搜索统计
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("扫描文件数", len(excel_files))
            with col2:
                st.metric("包含匹配的文件", files_with_matches)
            with col3:
                st.metric("总匹配行数", total_matches)
            
            # 按文件分组显示结果
            files_group = {}
            for match in all_matches:
                file_name = match['file_name']
                if file_name not in files_group:
                    files_group[file_name] = []
                files_group[file_name].append(match)
            
            # 显示每个文件的结果
            for file_name, file_matches in files_group.items():
                with st.expander(f"📄 {file_name} ({len(file_matches)} 行匹配)", expanded=True):
                    st.write(f"**文件路径:** {file_matches[0]['file_path']}")
                    
                    # 创建结果显示表格
                    display_data = []
                    for i, match in enumerate(file_matches):
                        row_data = match['row_data']
                        
                        # 准备行显示数据
                        row_display = {}
                        for col_name_display, cell_value in row_data.items():
                            if pd.isna(cell_value):
                                display_value = ""
                            else:
                                cell_str = str(cell_value)
                                
                                # 高亮关键词
                                if search_keyword and search_keyword.strip():
                                    highlighted = highlight_keyword(
                                        cell_str, search_keyword, case_sensitive
                                    )
                                    row_display[col_name_display] = highlighted
                                else:
                                    row_display[col_name_display] = cell_str
                        
                        display_data.append({
                            '序号': i + 1,
                            '匹配列值': match['matched_value'],
                            '关键词匹配数': match.get('match_count', 0) if match.get('keyword_found', False) else '无',
                            **row_display
                        })
                    
                    # 转换为DataFrame显示
                    if display_data:
                        # 获取所有列名
                        all_columns = set()
                        for item in display_data:
                            all_columns.update(item.keys())
                        
                        # 创建规范的显示顺序
                        base_columns = ['序号', '匹配列值', '关键词匹配数']
                        other_columns = [col for col in all_columns if col not in base_columns]
                        display_columns = base_columns + sorted(other_columns)
                        
                        # 创建显示DataFrame
                        display_df = pd.DataFrame(display_data)
                        
                        # 确保所有列都存在
                        for col in display_columns:
                            if col not in display_df.columns:
                                display_df[col] = ""
                        
                        # 重新排列列顺序
                        display_df = display_df[display_columns]
                        
                        # 显示表格
                        st.dataframe(display_df, use_container_width=True)
                    
                    # 显示匹配详情
                    if search_keyword and search_keyword.strip():
                        st.subheader("🔍 匹配详情")
                        for i, match in enumerate(file_matches):
                            if match.get('keyword_found', False) and match.get('keyword_matches'):
                                with st.expander(f"匹配详情 - 行 {i+1}"):
                                    st.write(f"**文件:** {file_name}")
                                    st.write(f"**行索引:** {match['row_index'] + 2}")  # +2 因为Excel从1开始且有标题
                                    st.write(f"**在列 '{col_name}' 中找到值:** {match['matched_value']}")
                                    st.write(f"**关键词匹配位置:**")
                                    
                                    for kw_match in match['keyword_matches']:
                                        col_name_kw = kw_match['column']
                                        cell_value = kw_match['original_value']
                                        match_text = kw_match['match_text']
                                        start_pos = kw_match['start_pos']
                                        end_pos = kw_match['end_pos']
                                        
                                        # 显示上下文
                                        context_start = max(0, start_pos - 20)
                                        context_end = min(len(cell_value), end_pos + 20)
                                        context = cell_value[context_start:context_end]
                                        
                                        # 高亮显示
                                        if context_start > 0:
                                            context = "..." + context
                                        if context_end < len(cell_value):
                                            context = context + "..."
                                        
                                        highlighted_context = highlight_keyword(
                                            context, search_keyword, case_sensitive
                                        )
                                        
                                        st.write(f"**列 '{col_name_kw}':** {highlighted_context}")
            
            # 提供结果下载
            if all_matches:
                st.header("💾 下载搜索结果")
                
                # 准备下载数据
                download_data = []
                for match in all_matches:
                    row_data = {
                        '文件路径': match['file_path'],
                        '文件名称': match['file_name'],
                        '行索引': match['row_index'] + 2,  # 转换为Excel行号
                        '匹配列': match['matched_column'],
                        '匹配列值': match['matched_value'],
                        '是否找到关键词': '是' if match.get('keyword_found', False) else '否',
                        '关键词匹配数': match.get('match_count', 0)
                    }
                    
                    # 添加所有列数据
                    for col_name_dl, cell_value in match['row_data'].items():
                        row_data[col_name_dl] = cell_value if not pd.isna(cell_value) else ""
                    
                    download_data.append(row_data)
                
                download_df = pd.DataFrame(download_data)
                
                # 提供下载
                csv_data = download_df.to_csv(index=False).encode('utf-8-sig')
                
                st.download_button(
                    label="📥 下载搜索结果(CSV)",
                    data=csv_data,
                    file_name=f"excel_search_results.csv",
                    mime="text/csv",
                    use_container_width=True
                )

    # 替换功能界面
    if st.session_state.get('search_results'):
        st.markdown("---")
        st.header("🔄 替换功能")
        
        # 获取搜索结果
        search_results = st.session_state.search_results
        search_keyword = st.session_state.search_keyword
        case_sensitive = st.session_state.case_sensitive
        match_whole_word = st.session_state.match_whole_word
        
        # 替换设置
        col1, col2 = st.columns(2)
        
        with col1:
            replace_keyword = st.text_input(
                "🔄 替换为:",
                value="僕",
                placeholder="请输入替换后的词语",
                help="将查找到的关键词替换为此词语"
            )
        
        with col2:
            replace_mode = st.radio(
                "替换模式:",
                ["替换所有匹配", "仅替换选中项"],
                help="选择替换全部匹配项还是仅替换选中的匹配项",
                key = "tihuanA"
            )
            
            create_backup = st.checkbox(
                "创建备份文件",
                value=True,
                help="替换前自动创建备份文件"
            )
        
        # 显示替换预览
        st.subheader("👁️ 替换预览")
        
        # 收集所有匹配项用于替换
        all_replacements = []
        files_to_replace = set()
        
        for match in search_results:
            if match.get('keyword_found', False) and match.get('keyword_matches'):
                file_path = match['file_path']
                files_to_replace.add(file_path)
                
                for kw_match in match['keyword_matches']:
                    replacement_info = {
                        'file_path': file_path,
                        'file_name': match['file_name'],
                        'row_index': match['row_index'],
                        'column': kw_match['column'],
                        'original_value': kw_match['original_value'],
                        'search_keyword': search_keyword,
                        'replace_keyword': replace_keyword,
                        'start_pos': kw_match['start_pos'],
                        'end_pos': kw_match['end_pos'],
                        'case_sensitive': case_sensitive,
                        'match_whole_word': match_whole_word,
                        'replace_all': (replace_mode == "替换所有匹配")
                    }
                    
                    # 计算替换后的值
                    if replacement_info['replace_all']:
                        flags = 0 if case_sensitive else re.IGNORECASE
                        pattern = re.escape(search_keyword)
                        if match_whole_word:
                            pattern = r'\b' + pattern + r'\b'
                        
                        new_value = re.sub(pattern, replace_keyword, kw_match['original_value'], flags=flags)
                    else:
                        new_value = kw_match['original_value'][:kw_match['start_pos']] + replace_keyword + kw_match['original_value'][kw_match['end_pos']:]
                    
                    replacement_info['new_value'] = new_value
                    all_replacements.append(replacement_info)
        
        # 显示替换预览
        if all_replacements:
            st.info(f"📊 共找到 {len(all_replacements)} 处可替换内容，涉及 {len(files_to_replace)} 个文件")
            
            # 按文件分组显示预览
            for file_path in files_to_replace:
                file_replacements = [r for r in all_replacements if r['file_path'] == file_path]
                
                with st.expander(f"📄 {Path(file_path).name} - {len(file_replacements)} 处替换"):
                    preview_data = []
                    
                    for i, replacement in enumerate(file_replacements[:10]):  # 只显示前10个
                        original_text = replacement['original_value']
                        new_text = replacement['new_value']
                        
                        # 高亮显示变化
                        highlighted_original = highlight_keyword(original_text, search_keyword, case_sensitive)
                        highlighted_new = highlight_keyword(new_text, replace_keyword, case_sensitive)
                        
                        preview_data.append({
                            '序号': i + 1,
                            '行号': replacement['row_index'] + 2,
                            '列名': replacement['column'],
                            '原内容': highlighted_original,
                            '新内容': highlighted_new,
                            '变化': "✅ 有变化" if original_text != new_text else "⚠️ 无变化"
                        })
                    
                    if preview_data:
                        preview_df = pd.DataFrame(preview_data)
                        st.dataframe(preview_df, use_container_width=True)
                    
                    if len(file_replacements) > 10:
                        st.info(f"... 还有 {len(file_replacements) - 10} 处替换未显示")
            
            # 替换确认
            st.subheader("✅ 替换确认")
            
            st.warning("⚠️ 此操作将修改原始Excel文件！请确认以下信息：")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("总替换数", len(all_replacements))
            with col2:
                st.metric("涉及文件数", len(files_to_replace))
            with col3:
                st.metric("备份文件", "会创建" if create_backup else "不创建")
            
            # 二次确认
            confirm_replace = st.checkbox("我确认要执行替换操作，理解此操作会修改原始文件")
            
            if confirm_replace:
                if st.button("🔄 执行替换", type="primary", use_container_width=True):
                    # 按文件分组执行替换
                    total_replaced = 0
                    success_files = 0
                    
                    for file_path in files_to_replace:
                        file_replacements = [r for r in all_replacements if r['file_path'] == file_path]
                        
                        with st.spinner(f"正在替换文件 {Path(file_path).name}..."):
                            success, message = replace_in_excel(file_path, file_replacements, create_backup)
                            
                            if success:
                                success_files += 1
                                total_replaced += len(file_replacements)
                                st.success(f"✅ {Path(file_path).name}: {message}")
                            else:
                                st.error(f"❌ {Path(file_path).name}: {message}")
                    
                    st.success(f"🎉 替换完成！成功处理 {success_files}/{len(files_to_replace)} 个文件，共替换 {total_replaced} 处内容")
                    
                    # 清空搜索结果，提示重新搜索
                    st.session_state.search_results = None
                    st.info("💡 替换完成，请重新搜索以查看更新后的内容")
        else:
            st.warning("未找到可替换的内容")
def term_lookup_page():
    st.title("🔍 术语查询")
    st.markdown("### 智能搜索术语库与角色数据库 | by Jacky_9S")
    
    # 初始化会话状态
    if 'lookup_translator' not in st.session_state:
        st.session_state.lookup_translator = MultiAPIExcelTranslator(
            api_key="", 
            api_provider="DeepSeek", 
            api_url=get_api_providers()["DeepSeek"]["url"], 
            model="deepseek-chat"
        )
    
    if 'lookup_term_base_loaded' not in st.session_state:
        st.session_state.lookup_term_base_loaded = False
    
    if 'lookup_role_base_loaded' not in st.session_state:
        st.session_state.lookup_role_base_loaded = False
    
    if 'lookup_term_target_cols' not in st.session_state:
        st.session_state.lookup_term_target_cols = []
    
    if 'lookup_term_df' not in st.session_state:
        st.session_state.lookup_term_df = None
    
    if 'lookup_role_df' not in st.session_state:
        st.session_state.lookup_role_df = None
    
    if 'lookup_term_source_col' not in st.session_state:
        st.session_state.lookup_term_source_col = None
    
    if 'lookup_role_name_col' not in st.session_state:
        st.session_state.lookup_role_name_col = None
    
    if 'lookup_role_personality_col' not in st.session_state:
        st.session_state.lookup_role_personality_col = None
    
    translator = st.session_state.lookup_translator
    
    # 快速导入区域
    st.header("🚀 快速导入")
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🗑️ 清空所有数据", use_container_width=True):
            # 清空所有数据
            translator.term_base_list = []
            translator.role_personality_dict = {}
            st.session_state.lookup_term_base_loaded = False
            st.session_state.lookup_role_base_loaded = False
            st.session_state.lookup_term_target_cols = []
            st.session_state.lookup_term_df = None
            st.session_state.lookup_role_df = None
            st.session_state.lookup_term_source_col = None
            st.session_state.lookup_role_name_col = None
            st.session_state.lookup_role_personality_col = None
            st.success("✅ 所有数据已清空!")
            st.rerun()
    
    st.markdown("---")
    
    # 批量上传区域
    st.header("📤 批量上传")
    batch_files = st.file_uploader(
        "📁 同时选择术语库和角色档案文件（可多选）",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        key="lookup_batch_files"
    )
    
    if batch_files:
        st.info(f"已选择 {len(batch_files)} 个文件")
        for i, file in enumerate(batch_files):
            st.write(f"{i+1}. {file.name}")
        
        if st.button("🔄 处理批量上传的文件", key="process_batch_files"):
            term_file = None
            role_file = None
            
            # 识别文件类型
            for file in batch_files:
                file_name = file.name.lower()
                if '术语库' in file_name or 'term' in file_name or '术语' in file_name:
                    term_file = file
                elif '角色档案' in file_name or 'role' in file_name or '角色' in file_name:
                    role_file = file
            
            # 处理识别到的文件
            if term_file:
                try:
                    df = pd.read_excel(term_file)
                    df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                    st.session_state.lookup_term_df = df
                    st.success(f"✅ 已识别并加载术语库文件: {term_file.name}")
                except Exception as e:
                    st.error(f"❌ 术语库文件处理失败: {e}")
            
            if role_file:
                try:
                    df = pd.read_excel(role_file)
                    df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                    st.session_state.lookup_role_df = df
                    st.success(f"✅ 已识别并加载角色档案文件: {role_file.name}")
                except Exception as e:
                    st.error(f"❌ 角色档案文件处理失败: {e}")
            
            if not term_file and not role_file:
                st.warning("⚠️ 未能识别术语库或角色档案文件")
                st.info("💡 提示：文件名应包含'术语库'或'角色档案'等关键词")
    
    st.markdown("---")
    
    # 术语库和角色库上传区域
    col1, col2 = st.columns(2)
    
    with col1:
        st.header("📚 术语库上传")
        
        uploaded_term_base = st.file_uploader(
            "📁 选择术语库文件（Excel）",
            type=['xlsx', 'xls'],
            key="lookup_term_base_uploader"
        )
        
        if uploaded_term_base is not None:
            try:
                df = pd.read_excel(uploaded_term_base)
                df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.lookup_term_df = df
                st.success(f"✅ 成功读取术语库，共 {len(df)} 条记录")
                
                with st.expander("📊 术语库预览"):
                    st.dataframe(df.head(10))
            except Exception as e:
                st.error(f"❌ 处理术语库文件失败: {e}")
        
        # 如果术语库已加载，显示列选择
        if st.session_state.lookup_term_df is not None:
            df = st.session_state.lookup_term_df
            cols = df.columns.tolist()
            
            st.subheader("📝 选择列")
            source_col_name = st.selectbox(
                "选择中文列",
                options=cols,
                index=0,
                key="lookup_term_source_col_select"
            )
            
            target_cols_names = st.multiselect(
                "选择翻译列（可多选）",
                options=cols,
                default=[cols[1]] if len(cols) > 1 else [],
                key="lookup_term_target_cols_select"
            )
            
            if st.button("📥 加载术语库", key="lookup_load_term_base"):
                if not target_cols_names:
                    st.error("❌ 请至少选择一个翻译列")
                else:
                    # 存储列名和索引
                    st.session_state.lookup_term_source_col = source_col_name
                    st.session_state.lookup_term_target_cols = target_cols_names
                    
                    # 构建术语列表（用于兼容原有的查找逻辑）
                    term_list = []
                    for _, row in df.iterrows():
                        source = row[source_col_name]
                        if pd.isna(source):
                            continue
                        source = str(source).strip()
                        if not source:
                            continue
                        
                        for target_col in target_cols_names:
                            target = row[target_col]
                            if pd.isna(target) or str(target).strip() in ['', '-']:
                                continue
                            target = str(target).strip()
                            
                            term_list.append({
                                'source': source,
                                'target': target,
                                'target_col': target_col
                            })
                    
                    translator.term_base_list = term_list
                    st.session_state.lookup_term_base_loaded = True
                    st.success(f"✅ 术语库加载成功：共 {len(df)} 条记录，{len(term_list)} 个术语-翻译对")
                    st.rerun()
        
        if st.session_state.lookup_term_base_loaded:
            st.metric("📊 已加载术语数", len(translator.term_base_list))
    
    with col2:
        st.header("👤 角色性格库上传")
        
        uploaded_role_base = st.file_uploader(
            "📁 选择角色性格库文件（Excel）",
            type=['xlsx', 'xls'],
            key="lookup_role_base_uploader"
        )
        
        if uploaded_role_base is not None:
            try:
                df = pd.read_excel(uploaded_role_base)
                df.columns = df.columns.str.strip().str.replace('\n', '').str.replace('\r', '')
                st.session_state.lookup_role_df = df
                st.success(f"✅ 成功读取角色性格库，共 {len(df)} 条记录")
                
                with st.expander("📊 角色性格库预览"):
                    st.dataframe(df.head(10))
            except Exception as e:
                st.error(f"❌ 处理角色性格库文件失败: {e}")
        
        # 如果角色库已加载，显示列选择
        if st.session_state.lookup_role_df is not None:
            df = st.session_state.lookup_role_df
            cols = df.columns.tolist()
            
            st.subheader("📝 选择列")
            role_name_col = st.selectbox(
                "选择角色名称列",
                options=cols,
                index=0,
                key="lookup_role_name_col_select"
            )
            
            role_personality_col = st.selectbox(
                "选择性格描述列",
                options=cols,
                index=min(1, len(cols)-1) if len(cols) > 1 else 0,
                key="lookup_role_personality_col_select"
            )
            
            if st.button("📥 加载角色性格库", key="lookup_load_role_base"):
                st.session_state.lookup_role_name_col = role_name_col
                st.session_state.lookup_role_personality_col = role_personality_col
                
                # 加载到translator
                if translator.load_role_personality(df, role_name_col, role_personality_col):
                    st.session_state.lookup_role_base_loaded = True
                    st.success(f"✅ 角色性格库加载成功：共 {len(translator.role_personality_dict)} 个角色")
                    st.rerun()
        
        if st.session_state.lookup_role_base_loaded:
            st.metric("📊 已加载角色数", len(translator.role_personality_dict))
    
    st.markdown("---")
    
    # 智能查询区域
    st.header("🔍 智能查询")
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        query_input = st.text_input(
            "输入查询词语",
            placeholder="例如：吃、杰克、不要、美味等",
            key="lookup_query_input"
        )
    
    with col2:
        st.write("")
        st.write("")
        search_button = st.button("🚀 开始查询", use_container_width=True, key="lookup_search_button")
    
    match_mode = st.radio(
        "匹配模式",
        ["模糊匹配", "精确匹配"],
        horizontal=True,
        key="lookup_match_mode"
    )
    
    # 执行查询
    if search_button or (query_input and len(query_input) > 0):
        if not query_input:
            st.warning("⚠️ 请输入查询词语")
        elif not st.session_state.lookup_term_base_loaded and not st.session_state.lookup_role_base_loaded:
            st.warning("⚠️ 请先加载术语库或角色性格库")
        else:
            with st.spinner("🔍 正在查询..."):
                time.sleep(0.3)  # 模拟加载
                
                st.subheader("📊 查询结果")
                
                total_matches = 0
                
                # 查询术语库
                if st.session_state.lookup_term_base_loaded and st.session_state.lookup_term_df is not None:
                    st.markdown("### 📚 术语库匹配结果")
                    
                    df = st.session_state.lookup_term_df
                    source_col = st.session_state.lookup_term_source_col
                    target_cols = st.session_state.lookup_term_target_cols
                    
                    # 收集匹配的行
                    matched_rows = []
                    for idx, row in df.iterrows():
                        source_text = row[source_col]
                        if pd.isna(source_text):
                            continue
                        
                        source_text = str(source_text)
                        
                        # 检查源列是否匹配
                        if match_mode == "模糊匹配":
                            match = query_input in source_text
                        else:
                            match = query_input == source_text
                        
                        # 如果源列不匹配，检查翻译列
                        if not match:
                            for target_col in target_cols:
                                target_text = row[target_col]
                                if pd.isna(target_text) or str(target_text).strip() in ['', '-']:
                                    continue
                                target_text = str(target_text)
                                
                                if match_mode == "模糊匹配":
                                    if query_input in target_text:
                                        match = True
                                        break
                                else:
                                    if query_input == target_text:
                                        match = True
                                        break
                        
                        if match:
                            matched_rows.append(row)
                    
                    # 合并相同词条的翻译
                    merged_terms = {}
                    for row in matched_rows:
                        source_text = str(row[source_col])
                        if source_text not in merged_terms:
                            merged_terms[source_text] = {col: [] for col in target_cols}
                        
                        for target_col in target_cols:
                            value = row[target_col]
                            if pd.isna(value) or str(value).strip() in ['', '-']:
                                continue
                            value = str(value)
                            if value not in merged_terms[source_text][target_col]:
                                merged_terms[source_text][target_col].append(value)
                    
                    if merged_terms:
                        st.success(f"✅ 找到 {len(merged_terms)} 个不同词条，共 {len(matched_rows)} 条记录")
                        total_matches += len(merged_terms)
                        
                        for source_text, translations in merged_terms.items():
                            with st.container():
                                st.markdown(f"#### 🔸 {source_text}")
                                
                                cols_display = st.columns(len(target_cols))
                                for i, target_col in enumerate(target_cols):
                                    with cols_display[i]:
                                        values = translations[target_col]
                                        display_value = ', '.join(values) if values else '-'
                                        st.info(f"**{target_col}:** {display_value}")
                                
                                st.markdown("---")
                    else:
                        st.warning(f"⚠️ 术语库中未找到包含「{query_input}」的词条")
                
                # 查询角色性格库
                if st.session_state.lookup_role_base_loaded and st.session_state.lookup_role_df is not None:
                    st.markdown("### 👤 角色性格库匹配结果")
                    
                    df = st.session_state.lookup_role_df
                    role_name_col = st.session_state.lookup_role_name_col
                    role_personality_col = st.session_state.lookup_role_personality_col
                    
                    # 收集匹配的角色
                    matched_roles = []
                    for idx, row in df.iterrows():
                        role_name = row[role_name_col]
                        if pd.isna(role_name):
                            continue
                        
                        role_name = str(role_name)
                        
                        if match_mode == "模糊匹配":
                            match = query_input in role_name
                        else:
                            match = query_input == role_name
                        
                        if match:
                            matched_roles.append(row)
                    
                    if matched_roles:
                        st.success(f"✅ 找到 {len(matched_roles)} 个匹配角色")
                        total_matches += len(matched_roles)
                        
                        for row in matched_roles:
                            role_name = str(row[role_name_col])
                            personality = str(row[role_personality_col]) if not pd.isna(row[role_personality_col]) else '无性格描述'
                            
                            with st.container():
                                st.markdown(f"#### 👤 {role_name}")
                                
                                # 在术语库中查找该角色名的翻译
                                if st.session_state.lookup_term_base_loaded and st.session_state.lookup_term_df is not None:
                                    term_df = st.session_state.lookup_term_df
                                    source_col = st.session_state.lookup_term_source_col
                                    target_cols = st.session_state.lookup_term_target_cols
                                    
                                    role_term_matches = term_df[term_df[source_col] == role_name]
                                    
                                    if not role_term_matches.empty:
                                        # 合并翻译
                                        merged_translations = {col: [] for col in target_cols}
                                        
                                        for _, term_row in role_term_matches.iterrows():
                                            for target_col in target_cols:
                                                value = term_row[target_col]
                                                if pd.isna(value) or str(value).strip() in ['', '-']:
                                                    continue
                                                value = str(value)
                                                if value not in merged_translations[target_col]:
                                                    merged_translations[target_col].append(value)
                                        
                                        # 显示翻译
                                        cols_display = st.columns(len(target_cols))
                                        for i, target_col in enumerate(target_cols):
                                            with cols_display[i]:
                                                values = merged_translations[target_col]
                                                display_value = ', '.join(values) if values else '-'
                                                st.info(f"**{target_col}:** {display_value}")
                                
                                # 显示性格描述
                                st.success(f"**💬 性格描述:** {personality}")
                                st.markdown("---")
                    else:
                        st.warning(f"⚠️ 角色性格库中未找到包含「{query_input}」的角色")
                
                # 显示统计信息
                st.info(f"🔍 查询词：**{query_input}** | 匹配模式：**{match_mode}** | 共找到：**{total_matches}** 条结果")
def excel_ABC_page():
    """Excel批量处理工具主函数"""
    
    st.title("📊 Excel批量处理工具")
    
    # 初始化session state
    if 'excel_files' not in st.session_state:
        st.session_state.excel_files = []
    if 'dataframes' not in st.session_state:
        st.session_state.dataframes = {}
    
    # 辅助函数
    def load_excel_file(file_path):
        """加载Excel文件"""
        try:
            df = pd.read_excel(file_path)
            return df
        except Exception as e:
            st.error(f"读取文件失败 {file_path}: {str(e)}")
            return None
    
    def check_condition(value, keywords, match_mode):
        """检查值是否满足关键词条件"""
        if not keywords:
            return True
        
        value_str = str(value).lower()
        keywords_list = [kw.strip().lower() for kw in keywords if kw.strip()]
        
        if not keywords_list:
            return True
        
        if match_mode == "同时包含所有关键词":
            return all(kw in value_str for kw in keywords_list)
        else:  # 包含任意一个关键词
            return any(kw in value_str for kw in keywords_list)
    
    def process_dataframe(df, col1, col2, keywords, match_mode, operation, params):
        """处理数据框"""
        df_copy = df.copy()
        modified_count = 0
        
        for idx, row in df_copy.iterrows():
            if check_condition(row[col1], keywords, match_mode):
                if operation == "删除值":
                    target_value = params.get('target_value', '')
                    if target_value:
                        cell_value = str(row[col2])
                        if target_value in cell_value:
                            df_copy.at[idx, col2] = cell_value.replace(target_value, '')
                            modified_count += 1
                        
                elif operation == "替换值":
                    old_value = params.get('old_value', '')
                    new_value = params.get('new_value', '')
                    if old_value:
                        cell_value = str(row[col2])
                        if old_value in cell_value:
                            df_copy.at[idx, col2] = cell_value.replace(old_value, new_value)
                            modified_count += 1
                        
                elif operation == "修改中间值":
                    value_a = params.get('value_a', '')
                    value_c = params.get('value_c', '')
                    new_value = params.get('new_value', '')
                    
                    cell_value = str(row[col2])
                    if value_a and value_c and value_a in cell_value and value_c in cell_value:
                        pos_a = cell_value.find(value_a)
                        pos_c = cell_value.find(value_c, pos_a + len(value_a))
                        
                        if pos_c > pos_a:
                            before = cell_value[:pos_a + len(value_a)]
                            after = cell_value[pos_c:]
                            df_copy.at[idx, col2] = before + new_value + after
                            modified_count += 1
        
        return df_copy, modified_count
    
    # 主界面
    # 1. 文件上传
    st.header("1️⃣ 上传文件夹")
    uploaded_files = st.file_uploader(
        "选择Excel文件（可多选）", 
        type=['xlsx', 'xls'], 
        accept_multiple_files=True,
        key="excel_uploader"
    )
    
    if uploaded_files:
        st.session_state.excel_files = []
        st.session_state.dataframes = {}
        
        for uploaded_file in uploaded_files:
            temp_path = f"temp_{uploaded_file.name}"
            with open(temp_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            df = load_excel_file(temp_path)
            if df is not None:
                st.session_state.excel_files.append(uploaded_file.name)
                st.session_state.dataframes[uploaded_file.name] = df
            
            os.remove(temp_path)
        
        st.success(f"已加载 {len(st.session_state.excel_files)} 个Excel文件")
        
        with st.expander("查看已加载的文件"):
            for file_name in st.session_state.excel_files:
                st.write(f"- {file_name}")
    
    # 2. 处理设置
    if st.session_state.excel_files:
        st.header("2️⃣ 配置处理规则")
        
        # 选择一个文件来预览列名
        sample_file = st.session_state.excel_files[0]
        sample_df = st.session_state.dataframes[sample_file]
        columns = list(sample_df.columns)
        
        col_left, col_right = st.columns(2)
        
        with col_left:
            st.subheader("选择列")
            col1 = st.selectbox("第一列（条件列）", columns, key="col1")
            col2 = st.selectbox("第二列（操作列）", columns, key="col2")
        
        with col_right:
            st.subheader("条件设置")
            keywords_input = st.text_area(
                "关键词（每行一个，留空则处理所有行）",
                height=100,
                placeholder="输入关键词\n可输入多个\n每行一个",
                key="keywords_input"
            )
            keywords = [kw.strip() for kw in keywords_input.split('\n') if kw.strip()]
            
            match_mode = st.radio(
                "匹配模式",
                ["同时包含所有关键词", "包含任意一个关键词"],
                disabled=len(keywords) == 0,
                key="match_mode"
            )
        
        st.divider()
        
        # 3. 操作选择
        st.subheader("选择操作")
        operation = st.radio(
            "操作类型",
            ["删除值", "替换值", "修改中间值"],
            key="operation"
        )
        
        params = {}
        
        if operation == "删除值":
            st.info("💡 删除第二列文本中包含的指定内容")
            params['target_value'] = st.text_input(
                "要删除的内容", 
                key="delete_value", 
                placeholder="例如：删除'帅'，则'我好帅'变为'我好'"
            )
            
        elif operation == "替换值":
            st.info("💡 将第二列文本中的某个内容替换为新内容")
            col_a, col_b = st.columns(2)
            with col_a:
                params['old_value'] = st.text_input(
                    "要替换的内容", 
                    key="old_value",
                    placeholder="例如：帅"
                )
            with col_b:
                params['new_value'] = st.text_input(
                    "替换为", 
                    key="new_value",
                    placeholder="例如：丑"
                )
                
        elif operation == "修改中间值":
            st.info("💡 修改夹在两个值之间的内容")
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                params['value_a'] = st.text_input("起始值 A", key="value_a")
            with col_b:
                params['value_c'] = st.text_input("结束值 C", key="value_c")
            with col_c:
                params['new_value'] = st.text_input("新的中间值", key="middle_new_value")
        
        st.divider()
        
        # 4. 预览和执行
        st.header("3️⃣ 预览和执行")
        
        col_preview, col_execute = st.columns(2)
        
        with col_preview:
            if st.button("🔍 预览效果（使用第一个文件）", type="secondary", use_container_width=True):
                preview_df = st.session_state.dataframes[sample_file].copy()
                processed_df, count = process_dataframe(
                    preview_df, col1, col2, keywords, match_mode, operation, params
                )
                
                st.success(f"预览完成！共修改 {count} 行数据")
                
                col_before, col_after = st.columns(2)
                with col_before:
                    st.write("**处理前**")
                    st.dataframe(preview_df[[col1, col2]].head(20), use_container_width=True)
                with col_after:
                    st.write("**处理后**")
                    st.dataframe(processed_df[[col1, col2]].head(20), use_container_width=True)
        
        with col_execute:
            if st.button("✅ 批量处理所有文件", type="primary", use_container_width=True):
                with st.spinner("正在处理..."):
                    results = {}
                    total_modified = 0
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    for idx, file_name in enumerate(st.session_state.excel_files):
                        status_text.text(f"正在处理: {file_name}")
                        df = st.session_state.dataframes[file_name]
                        processed_df, count = process_dataframe(
                            df, col1, col2, keywords, match_mode, operation, params
                        )
                        results[file_name] = processed_df
                        total_modified += count
                        progress_bar.progress((idx + 1) / len(st.session_state.excel_files))
                    
                    status_text.empty()
                    progress_bar.empty()
                    
                    st.success(f"✅ 处理完成！共修改 {total_modified} 行数据")
                    
                    # 创建下载包
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for file_name, df in results.items():
                            excel_buffer = BytesIO()
                            df.to_excel(excel_buffer, index=False)
                            zip_file.writestr(f"processed_{file_name}", excel_buffer.getvalue())
                    
                    st.download_button(
                        label="📥 下载处理后的文件（ZIP）",
                        data=zip_buffer.getvalue(),
                        file_name="processed_excel_files.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
    
    else:
        st.info("👆 请先上传Excel文件")
    
    # 使用说明
    with st.expander("📖 使用说明"):
        st.markdown("""
        ### 功能说明
        
        1. **上传文件**：选择多个Excel文件（支持.xlsx和.xls格式）
        
        2. **选择列**：
           - 第一列：条件列，用于判断是否符合处理条件
           - 第二列：操作列，对符合条件的行进行操作
        
        3. **设置条件**：
           - 输入关键词（每行一个）
           - 留空表示处理所有行
           - 选择匹配模式：同时包含所有关键词 或 包含任意一个关键词
        
        4. **选择操作**：
           - **删除值**：删除第二列文本中包含的指定内容（例如：删除"帅"，"我好帅"变为"我好"）
           - **替换值**：将第二列文本中的某个内容替换为新内容（例如："帅"替换为"丑"，"我好帅"变为"我好丑"）
           - **修改中间值**：修改夹在A和C之间的B值（例如：A="<"，C=">"，将"<旧值>"改为"<新值>"）
        
        5. **预览和执行**：
           - 先预览第一个文件的处理效果
           - 确认无误后批量处理所有文件
           - 下载处理后的文件压缩包
        
        ### 使用示例
        
        **示例1：删除指定内容**
        - 条件：第一列包含"产品"
        - 操作：删除第二列中的"旧版"
        - 结果："旧版产品说明" → "产品说明"
        
        **示例2：替换内容**
        - 条件：第一列包含"评价"
        - 操作：将"帅"替换为"丑"
        - 结果："这个人好帅" → "这个人好丑"
        
        **示例3：修改中间值**
        - 条件：第一列包含"标签"
        - 操作：A="【"，C="】"，新值="已处理"
        - 结果："【待处理】任务" → "【已处理】任务"
        """)
    
# 主程序
def jacky_page():
    st.header("作者主页")
    col1,col2,col3 = st.columns(3)
    
    with col1:
        if st.button("📖 打开作者主页", use_container_width=True):
            st.markdown("[作者主页](https://jackyjay.cn)")
        if st.button("🔍 打开百度", use_container_width=True):
            st.markdown("[百度](https://www.baidu.com)")
    
    with col2:
        if st.button("📚 打开GitHub", use_container_width=True):
            st.markdown("[GitHub](https://github.com)")
        if st.button("💬 打开Stack Overflow", use_container_width=True):
            st.markdown("[Stack Overflow](https://stackoverflow.com)")
    
    with col3:
        if st.button("📊 打开Streamlit文档", use_container_width=True):
            st.markdown("[Streamlit文档](https://docs.streamlit.io)")
        if st.button("🐼 打开Pandas文档", use_container_width=True):
            st.markdown("[Pandas文档](https://pandas.pydata.org/docs)")
def main():
    st.set_page_config(
        page_title="API_AI_Excel翻译分析工具_Jacky",
        page_icon="🎮",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # 侧边栏页面选择
    st.sidebar.title("🎮 多API Excel智能翻译工具")
    st.sidebar.markdown("---")
    
    page = st.sidebar.radio(
        "选择功能页面",
        [
            "📝 提示词生成器",
            "📊 翻译结果处理",
            "🔄 批量翻译工具",
            "术语查找",
            "excel查找替换",
            "excel高级替换",
            "Jacky的主页",
            "🔍 Excel表格对比",
            "🔍 ExcelABC操作",
            "🔍 抓弹幕（只支持nikoniko)",
            "blbl视频弹幕评论下载",
            "文件夹单向匹配程序",
            "模板一键匹配"
        ],
        index=0
    )
    
    st.sidebar.markdown("---")
    st.sidebar.markdown("""
    ### 📖 使用说明
    
    **提示词生成器：**
    - 上传待翻译文本
    - 加载术语库和性格库
    - 生成翻译提示词
    - 复制给AI进行翻译
    
    **翻译结果处理：**
    - 上传原始Excel文件
    - 粘贴AI翻译结果
    - 自动匹配合并
    - 下载完整结果
    
    **批量翻译工具：**
    - 配置API密钥
    - 上传文件和术语库
    - 自动批量翻译
    - 支持重试机制
    
    ### ⚙️ 版本信息
    版本: v2.0 合并版
    作者: Jacky_9S
    """)
    
    # 根据选择显示不同页面
    if page == "📝 提示词生成器":
        prompt_generator_page()
    elif page == "📊 翻译结果处理":
        translation_result_processor_page()
    elif page == "🔄 批量翻译工具":
        batch_translation_page()
    elif page == '术语查找':
        term_lookup_page()
    elif page == "excel查找替换":
        excel_replace_page()
    elif page == "excel高级替换":
        excel_sreplace_page()
    elif page == "Jacky的主页":
        jacky_page()
    elif page == "🔍 Excel表格对比":  # 新增的页面
        excel_comparison_page()
    elif page == "🔍 ExcelABC操作":  # 新增的页面
        excel_ABC_page()
    elif page == "🔍 抓弹幕（只支持nikoniko)":  # 新增的页面
        danmu_page()
    elif page == "blbl视频弹幕评论下载":  # 新增的页面
        ytdlp_downloader_app()
    elif page == "文件夹单向匹配程序":  # 新增的页面
        excel_matchpro_page()
    elif page == "模板一键匹配":  # 新增的页面
        grand_match()
if __name__ == "__main__":
    # 确保 jieba 库已安装
    try:
        import jieba
    except ImportError:
        print("jieba 库未安装，正在尝试安装...")
        os.system(f"{sys.executable} -m pip install jieba")
        import jieba
    main()
