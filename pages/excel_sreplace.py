# pages/excel_sreplace.py - Excel 高级替换页面

import os
import re
import shutil
import platform
import subprocess
from pathlib import Path
from datetime import datetime

import pandas as pd
import streamlit as st
import openpyxl


def excel_sreplace_page():
    """Excel高级替换页面 - 支持选择性替换和行选择"""
    st.title("🔍 Excel高级替换工具")
    st.markdown("### 高级搜索替换功能，支持选择特定行进行替换")

    # 初始化会话状态
    if 'sreplace_folder_path' not in st.session_state:
        st.session_state.sreplace_folder_path = ""
    if 'sreplace_search_term' not in st.session_state:
        st.session_state.sreplace_search_term = ""
    if 'sreplace_replace_term' not in st.session_state:
        st.session_state.sreplace_replace_term = ""
    if 'sreplace_results' not in st.session_state:
        st.session_state.sreplace_results = {}
    if 'sreplace_selected_rows' not in st.session_state:
        st.session_state.sreplace_selected_rows = set()

    # 侧边栏设置
    st.sidebar.header("📁 文件夹设置")
    folder_path = st.sidebar.text_input(
        "文件夹路径:",
        value=st.session_state.sreplace_folder_path,
        placeholder="请输入包含Excel文件的文件夹路径"
    )
    st.session_state.sreplace_folder_path = folder_path

    # 搜索设置
    st.header("🔍 搜索设置")

    col1, col2 = st.columns(2)

    with col1:
        search_term = st.text_input(
            "搜索词语:",
            value=st.session_state.sreplace_search_term,
            placeholder="请输入要搜索的词语"
        )
        st.session_state.sreplace_search_term = search_term

    with col2:
        replace_term = st.text_input(
            "替换为:",
            value=st.session_state.sreplace_replace_term,
            placeholder="请输入替换后的词语"
        )
        st.session_state.sreplace_replace_term = replace_term

    col1, col2 = st.columns(2)
    with col1:
        case_sensitive = st.checkbox("大小写敏感", value=False)
    with col2:
        create_backup = st.checkbox("创建备份", value=True)

    # 搜索按钮
    if st.button("🔍 搜索", use_container_width=True):
        if not folder_path:
            st.error("请输入文件夹路径")
            return

        if not search_term:
            st.error("请输入搜索词语")
            return

        folder = Path(folder_path)
        if not folder.exists():
            st.error("文件夹不存在")
            return

        # 搜索Excel文件
        results = {}
        excel_files = list(folder.rglob("*.xlsx")) + list(folder.rglob("*.xls"))

        if not excel_files:
            st.warning("未找到Excel文件")
            return

        progress_bar = st.progress(0)

        for i, file_path in enumerate(excel_files):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                file_matches = []

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        for col_idx, cell_value in enumerate(row, start=1):
                            if cell_value is not None:
                                cell_str = str(cell_value)
                                if case_sensitive:
                                    if search_term in cell_str:
                                        file_matches.append({
                                            'sheet': sheet_name,
                                            'row': row_idx,
                                            'col': col_idx,
                                            'value': cell_str
                                        })
                                else:
                                    if search_term.lower() in cell_str.lower():
                                        file_matches.append({
                                            'sheet': sheet_name,
                                            'row': row_idx,
                                            'col': col_idx,
                                            'value': cell_str
                                        })

                if file_matches:
                    results[str(file_path)] = file_matches

                wb.close()
            except Exception as e:
                st.warning(f"读取文件 {file_path.name} 失败: {e}")

            progress_bar.progress((i + 1) / len(excel_files))

        progress_bar.empty()
        st.session_state.sreplace_results = results

        total_matches = sum(len(m) for m in results.values())
        if total_matches > 0:
            st.success(f"✅ 在 {len(results)} 个文件中找到 {total_matches} 个匹配项")
        else:
            st.warning("未找到匹配项")

    # 显示搜索结果
    if st.session_state.sreplace_results:
        st.header("📊 搜索结果")

        results = st.session_state.sreplace_results

        for file_path, matches in results.items():
            with st.expander(f"📄 {Path(file_path).name} ({len(matches)} 处)", expanded=False):
                # 创建数据表格
                data = []
                for i, match in enumerate(matches):
                    row_key = f"{file_path}_{match['sheet']}_{match['row']}_{match['col']}"
                    data.append({
                        "选择": row_key in st.session_state.sreplace_selected_rows,
                        "工作表": match['sheet'],
                        "行": match['row'],
                        "列": match['col'],
                        "内容": match['value'][:100] + "..." if len(match['value']) > 100 else match['value']
                    })

                df = pd.DataFrame(data)

                # 显示表格并处理选择
                for idx, row in df.iterrows():
                    row_key = f"{file_path}_{matches[idx]['sheet']}_{matches[idx]['row']}_{matches[idx]['col']}"
                    col1, col2 = st.columns([1, 10])
                    with col1:
                        selected = st.checkbox("", value=row_key in st.session_state.sreplace_selected_rows,
                                               key=f"cb_{row_key}")
                        if selected:
                            st.session_state.sreplace_selected_rows.add(row_key)
                        elif row_key in st.session_state.sreplace_selected_rows:
                            st.session_state.sreplace_selected_rows.discard(row_key)
                    with col2:
                        st.write(f"**{row['工作表']}** | 行{row['行']} 列{row['列']}: {row['内容']}")

        # 替换按钮
        st.subheader("🔄 执行替换")
        selected_count = len(st.session_state.sreplace_selected_rows)
        st.write(f"已选择 {selected_count} 项进行替换")

        if st.button("🔄 执行选择性替换", type="primary", use_container_width=True):
            if not replace_term:
                st.error("请输入替换词语")
                return

            if selected_count == 0:
                st.error("请至少选择一项进行替换")
                return

            # 按文件分组选中的项
            file_replacements = {}
            for row_key in st.session_state.sreplace_selected_rows:
                parts = row_key.rsplit('_', 3)
                if len(parts) >= 4:
                    file_path = parts[0]
                    sheet = parts[1]
                    row = int(parts[2])
                    col = int(parts[3])

                    if file_path not in file_replacements:
                        file_replacements[file_path] = []
                    file_replacements[file_path].append({
                        'sheet': sheet,
                        'row': row,
                        'col': col
                    })

            # 执行替换
            replaced_count = 0
            for file_path, replacements in file_replacements.items():
                try:
                    # 备份
                    if create_backup:
                        backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                        shutil.copy2(file_path, backup_path)

                    # 加载并替换
                    wb = openpyxl.load_workbook(file_path)

                    for rep in replacements:
                        sheet = wb[rep['sheet']]
                        cell = sheet.cell(row=rep['row'], column=rep['col'])
                        if cell.value:
                            if case_sensitive:
                                cell.value = str(cell.value).replace(search_term, replace_term)
                            else:
                                cell.value = re.sub(re.escape(search_term), replace_term,
                                                    str(cell.value), flags=re.IGNORECASE)
                            replaced_count += 1

                    wb.save(file_path)
                    wb.close()

                except Exception as e:
                    st.error(f"替换文件 {Path(file_path).name} 失败: {e}")

            st.success(f"✅ 完成 {replaced_count} 处替换")
            st.session_state.sreplace_results = {}
            st.session_state.sreplace_selected_rows = set()
            st.rerun()
