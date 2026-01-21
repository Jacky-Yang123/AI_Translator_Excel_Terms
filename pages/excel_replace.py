# pages/excel_replace.py - Excel 查找替换页面

import os
import re
import shutil
import platform
import subprocess
import threading
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import streamlit as st
import openpyxl

from utils import open_folder, open_file


class ExcelSearchReplace:
    """Excel搜索替换工具类"""

    def __init__(self):
        self.excel_files = []
        self.search_results = {}
        self.case_sensitive = False
        self.match_whole_word = False

    def find_excel_files(self, folder_path):
        """查找文件夹中的所有Excel文件"""
        self.excel_files = []
        folder_path = Path(folder_path)

        if not folder_path.exists():
            return False, "文件夹路径不存在"

        excel_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']

        for ext in excel_extensions:
            self.excel_files.extend(folder_path.rglob(f'*{ext}'))

        return True, f"找到 {len(self.excel_files)} 个Excel文件"

    def search_in_excel(self, search_term, case_sensitive=False, match_whole_word=False):
        """在Excel文件中搜索词语"""
        self.search_results = {}
        self.case_sensitive = case_sensitive
        self.match_whole_word = match_whole_word
        total_matches = 0

        for file_path in self.excel_files:
            try:
                # 读取Excel文件的所有工作表
                excel_data = pd.read_excel(file_path, sheet_name=None, dtype=str)
                file_matches = []

                for sheet_name, df in excel_data.items():
                    sheet_matches = self._search_in_dataframe(df, search_term, sheet_name, str(file_path))
                    file_matches.extend(sheet_matches)

                if file_matches:
                    self.search_results[str(file_path)] = {
                        'matches': file_matches,
                        'match_count': len(file_matches)
                    }
                    total_matches += len(file_matches)

            except Exception as e:
                st.error(f"读取文件 {file_path.name} 时出错: {e}")

        return total_matches

    def _search_in_dataframe(self, df, search_term, sheet_name, file_path):
        """在DataFrame中搜索词语"""
        matches = []

        # 构建正则表达式模式
        if self.match_whole_word:
            pattern = r'\b' + re.escape(search_term) + r'\b'
        else:
            pattern = re.escape(search_term)

        flags = 0 if self.case_sensitive else re.IGNORECASE

        for row_idx, row in df.iterrows():
            for col_idx, cell_value in enumerate(row):
                if pd.isna(cell_value):
                    continue

                cell_str = str(cell_value)
                matches_found = list(re.finditer(pattern, cell_str, flags))

                for match in matches_found:
                    matches.append({
                        'file_path': file_path,
                        'sheet_name': sheet_name,
                        'row': row_idx + 2,  # +2 因为Excel从1开始，且有标题行
                        'column': df.columns[col_idx] if col_idx < len(df.columns) else f'Col{col_idx+1}',
                        'original_text': cell_str,
                        'matched_text': match.group(),
                        'start_pos': match.start(),
                        'end_pos': match.end()
                    })

        return matches

    def replace_in_excel(self, search_term, replace_term, backup=True):
        """替换Excel文件中的词语"""
        replaced_files = 0
        total_replacements = 0

        for file_path_str, file_data in self.search_results.items():
            file_path = Path(file_path_str)

            try:
                # 备份原文件
                if backup:
                    backup_path = file_path.parent / f"{file_path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_path.suffix}"
                    shutil.copy2(file_path, backup_path)
                    st.info(f"已创建备份: {backup_path.name}")

                # 读取Excel文件
                excel_data = pd.read_excel(file_path, sheet_name=None, dtype=str)
                replacements_in_file = 0

                # 构建替换模式
                if self.match_whole_word:
                    pattern = r'\b' + re.escape(search_term) + r'\b'
                else:
                    pattern = re.escape(search_term)

                flags = 0 if self.case_sensitive else re.IGNORECASE

                # 对每个工作表进行替换
                for sheet_name, df in excel_data.items():
                    df_replaced = df.map(
                        lambda x: self._replace_text(x, pattern, replace_term, flags)
                        if pd.notna(x) else x
                    )
                    excel_data[sheet_name] = df_replaced

                    # 计算替换数量
                    for row_idx, row in df.iterrows():
                        for col_idx, cell_value in enumerate(row):
                            if pd.isna(cell_value):
                                continue
                            cell_str = str(cell_value)
                            replacements = len(re.findall(pattern, cell_str, flags))
                            replacements_in_file += replacements

                # 保存替换后的文件
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet_name, df in excel_data.items():
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                replaced_files += 1
                total_replacements += replacements_in_file

                st.success(f"✅ {file_path.name}: 完成 {replacements_in_file} 处替换")

            except Exception as e:
                st.error(f"替换文件 {file_path.name} 时出错: {e}")

        return replaced_files, total_replacements

    def _replace_text(self, text, pattern, replace_term, flags):
        """替换文本中的匹配项"""
        if pd.isna(text):
            return text

        text_str = str(text)
        replaced_text = re.sub(pattern, replace_term, text_str, flags=flags)
        return replaced_text


def multithreaded_search(search_tool, search_term, case_sensitive, match_whole_word, progress_bar, status_text):
    """使用多线程搜索Excel文件"""
    # 清空之前的搜索结果
    search_tool.search_results = {}

    if not search_tool.excel_files:
        return 0

    # 准备搜索模式
    if match_whole_word:
        pattern = r'\b' + re.escape(search_term) + r'\b'
    else:
        pattern = re.escape(search_term)

    flags = 0 if case_sensitive else re.IGNORECASE
    regex = re.compile(pattern, flags)

    # 线程锁，用于安全更新共享数据
    lock = threading.Lock()
    total_matches = 0
    completed_files = 0
    total_files = len(search_tool.excel_files)

    def search_single_file(file_path):
        """搜索单个Excel文件"""
        nonlocal total_matches, completed_files

        file_matches = []
        match_count = 0

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value is not None:
                            cell_text = str(cell.value)

                            # 搜索匹配
                            matches = list(regex.finditer(cell_text))

                            if matches:
                                for match in matches:
                                    file_matches.append({
                                        'sheet_name': sheet_name,
                                        'row': row_idx,
                                        'column': col_idx,
                                        'original_text': cell_text,
                                        'matched_text': match.group(),
                                        'start_pos': match.start(),
                                        'end_pos': match.end()
                                    })
                                    match_count += 1

            wb.close()

            # 使用锁更新共享数据
            if file_matches:
                with lock:
                    search_tool.search_results[str(file_path)] = {
                        'matches': file_matches,
                        'match_count': match_count
                    }
                    total_matches += match_count

        except Exception as e:
            # 忽略无法读取的文件
            pass

        # 更新进度
        with lock:
            completed_files += 1
            progress = completed_files / total_files
            progress_bar.progress(progress)
            status_text.text(f"正在搜索... {completed_files}/{total_files} 个文件")

        return match_count

    # 使用线程池执行搜索
    # 根据CPU核心数设置线程数，最大为16
    max_workers = min(16, (os.cpu_count() or 4) * 2)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务
        futures = {executor.submit(search_single_file, file_path): file_path
                   for file_path in search_tool.excel_files}

        # 等待所有任务完成
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                # 记录错误但继续处理其他文件
                pass

    return total_matches


def get_row_data_as_list(file_path, sheet_name, row_num):
    """获取指定Excel文件中某一行的完整数据（以列表形式返回）"""
    try:
        # 读取Excel文件的指定工作表
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # 获取指定行的数据（注意：row_num是1-based，需要转换为0-based）
        if row_num <= len(df):
            row_data = df.iloc[row_num - 1]
            # 将该行数据转换为列表，保留NaN显示为空字符串
            return [str(val) if pd.notna(val) else "" for val in row_data]
        else:
            return ["(行号超出范围)"]
    except Exception as e:
        return [f"(读取失败: {str(e)})"]


def selective_replace(search_tool, search_term, replace_term, selected_replacements, backup_files, case_sensitive=False):
    """执行选择性替换（多线程版本）
    
    Args:
        case_sensitive: 是否大小写敏感，默认False（不区分大小写）
    """
    # 线程锁
    lock = threading.Lock()
    replaced_files = 0
    total_replacements = 0

    def replace_single_file(file_path, selection):
        """替换单个文件"""
        nonlocal replaced_files, total_replacements

        try:
            if not selection['selected']:
                return

            # 创建备份
            if backup_files:
                backup_path = f"{file_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                shutil.copy2(file_path, backup_path)

            # 加载工作簿
            wb = openpyxl.load_workbook(file_path)
            file_replaced = False
            file_replacement_count = 0

            # 获取该文件的匹配项
            matches = search_tool.search_results[file_path]['matches']

            # 构建正则表达式模式
            flags = 0 if case_sensitive else re.IGNORECASE
            pattern = re.escape(search_term)
            
            # 如果是全部替换模式
            if selection['rows'] == 'all':
                for match in matches:
                    sheet = wb[match['sheet_name']]
                    cell = sheet.cell(row=match['row'], column=match['column'])

                    if cell.value:
                        # 使用正则表达式执行替换（支持大小写不敏感）
                        new_value = re.sub(pattern, replace_term, str(cell.value), flags=flags)
                        if new_value != str(cell.value):
                            cell.value = new_value
                            file_replacement_count += 1
                            file_replaced = True

            # 如果是选择特定行模式
            else:
                selected_rows = selection['selected_rows']
                for match in matches:
                    row_key = f"{file_path}_{match['sheet_name']}_{match['row']}"
                    if row_key in selected_rows:
                        sheet = wb[match['sheet_name']]
                        cell = sheet.cell(row=match['row'], column=match['column'])

                        if cell.value:
                            # 使用正则表达式执行替换（支持大小写不敏感）
                            new_value = re.sub(pattern, replace_term, str(cell.value), flags=flags)
                            if new_value != str(cell.value):
                                cell.value = new_value
                                file_replacement_count += 1
                                file_replaced = True

            # 保存文件
            if file_replaced:
                wb.save(file_path)
                with lock:
                    replaced_files += 1
                    total_replacements += file_replacement_count

            wb.close()

        except Exception as e:
            st.error(f"替换文件 {Path(file_path).name} 时出错: {str(e)}")

    # 使用线程池执行替换
    max_workers = min(8, (os.cpu_count() or 4))

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(replace_single_file, file_path, selection): file_path
                   for file_path, selection in selected_replacements.items()}

        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                pass

    return replaced_files, total_replacements


def excel_replace_page():
    st.title("🔍 Excel文件批量搜索替换工具")
    st.markdown("### 批量搜索和替换文件夹中所有Excel文件的内容")

    # 初始化搜索替换工具
    if 'search_tool' not in st.session_state:
        st.session_state.search_tool = ExcelSearchReplace()

    # 初始化会话状态变量
    if 'folder_path' not in st.session_state:
        st.session_state.folder_path = ""
    if 'search_term' not in st.session_state:
        st.session_state.search_term = ""
    if 'replace_term' not in st.session_state:
        st.session_state.replace_term = ""
    if 'case_sensitive' not in st.session_state:
        st.session_state.case_sensitive = False
    if 'match_whole_word' not in st.session_state:
        st.session_state.match_whole_word = False
    if 'replace_confirmed' not in st.session_state:
        st.session_state.replace_confirmed = False
    if 'show_confirm_checkbox' not in st.session_state:
        st.session_state.show_confirm_checkbox = False
    if 'edited_data' not in st.session_state:
        st.session_state.edited_data = {}

    search_tool = st.session_state.search_tool

    # 侧边栏 - 文件夹选择
    st.sidebar.header("📁 文件夹设置")
    folder_path = st.sidebar.text_input(
        "请输入文件夹路径:",
        value=st.session_state.folder_path,
        placeholder="例如: C:/Users/用户名/Documents/Excel文件",
        help="请输入包含Excel文件的文件夹完整路径"
    )

    if folder_path and folder_path != st.session_state.folder_path:
        st.session_state.folder_path = folder_path
        success, message = search_tool.find_excel_files(folder_path)
        if success:
            st.sidebar.success(message)
        else:
            st.sidebar.error(message)

    # 显示找到的文件列表
    if search_tool.excel_files:
        st.sidebar.subheader("📊 找到的Excel文件")
        for i, file_path in enumerate(search_tool.excel_files[:10]):  # 只显示前10个
            st.sidebar.write(f"{i+1}. {file_path.name}")

        if len(search_tool.excel_files) > 10:
            st.sidebar.info(f"... 还有 {len(search_tool.excel_files) - 10} 个文件")

    # 主界面 - 搜索设置
    st.header("🔍 搜索设置")

    col1, col2 = st.columns(2)

    with col1:
        search_term = st.text_input(
            "搜索词语:",
            value=st.session_state.search_term,
            placeholder="请输入要搜索的词语",
            help="支持正则表达式语法"
        )
        st.session_state.search_term = search_term

    with col2:
        # 搜索选项
        st.subheader("⚙️ 搜索选项")
        case_sensitive = st.checkbox(
            "大小写敏感",
            value=st.session_state.case_sensitive,
            help="勾选后区分大小写"
        )
        st.session_state.case_sensitive = case_sensitive

        match_whole_word = st.checkbox(
            "全词匹配",
            value=st.session_state.match_whole_word,
            help="勾选后只匹配完整词语"
        )
        st.session_state.match_whole_word = match_whole_word

    # 搜索按钮
    if st.button("🚀 开始搜索", key="search_btn", use_container_width=True):
        if not folder_path:
            st.error("❌ 请输入文件夹路径")
            return

        if not search_term:
            st.error("❌ 请输入搜索词语")
            return

        # 执行多线程搜索
        progress_bar = st.progress(0)
        status_text = st.empty()

        with st.spinner("正在搜索Excel文件..."):
            total_matches = multithreaded_search(
                search_tool,
                search_term,
                case_sensitive,
                match_whole_word,
                progress_bar,
                status_text
            )

        progress_bar.empty()
        status_text.empty()

        if total_matches > 0:
            st.success(f"✅ 搜索完成！共找到 {total_matches} 个匹配项")
        else:
            st.warning("⚠️ 未找到匹配项")

    # 显示搜索结果
    if search_tool.search_results:
        st.header("📊 搜索结果预览")

        total_files = len(search_tool.search_results)
        total_matches = sum(data['match_count'] for data in search_tool.search_results.values())

        st.info(f"**统计信息:** 在 {total_files} 个文件中找到 {total_matches} 个匹配项")

        # 文件列表
        selected_file = st.selectbox(
            "选择文件查看详情:",
            options=list(search_tool.search_results.keys()),
            format_func=lambda x: f"{Path(x).name} ({search_tool.search_results[x]['match_count']} 处)"
        )

        if selected_file:
            file_data = search_tool.search_results[selected_file]
            matches = file_data['matches']

            # 文件信息和操作按钮
            st.subheader(f"📄 文件: {Path(selected_file).name}")

            # 显示完整文件路径
            st.code(selected_file, language=None)

            # 操作按钮行
            col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 3])

            with col_btn1:
                if st.button("📂 打开文件夹", key=f"open_folder_{selected_file}"):
                    open_folder(selected_file)

            with col_btn2:
                if st.button("📊 打开Excel", key=f"open_excel_{selected_file}"):
                    open_file(selected_file)

            st.write(f"**匹配数量:** {len(matches)} 处")

            # 显示匹配详情 - 使用可编辑的表格
            display_rows = []
            row_identifiers = []  # 存储行标识符用于后续保存

            for i, match in enumerate(matches[:50]):  # 只显示前50个
                # 获取该行的完整数据
                row_data = get_row_data_as_list(selected_file, match['sheet_name'], match['row'])

                # 添加位置信息作为第一列
                row_dict = {
                    "位置": f"{match['sheet_name']} | 行{match['row']} 列{match['column']}"
                }

                # 添加该行的所有列数据
                if row_data:
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        row_dict[f"列{col_idx}"] = cell_value

                display_rows.append(row_dict)
                row_identifiers.append({
                    'sheet': match['sheet_name'],
                    'row': match['row'],
                    'original_data': row_data
                })

            if display_rows:
                st.markdown("### ✏️ 可编辑表格（直接修改单元格内容）")

                # 创建DataFrame
                df_display = pd.DataFrame(display_rows)

                # 使用data_editor创建可编辑表格
                edited_df = st.data_editor(
                    df_display,
                    use_container_width=True,
                    height=400,
                    key=f"editable_table_{selected_file}",
                    column_config={
                        "位置": st.column_config.TextColumn(
                            "位置",
                            disabled=True,  # 位置列不可编辑
                            width="medium"
                        )
                    }
                )

                # 检测是否有修改
                if not df_display.equals(edited_df):
                    st.warning("⚠️ 检测到内容已修改，请点击下方按钮保存更改")

                    # 显示修改对比
                    with st.expander("📋 查看修改详情"):
                        changes_found = False
                        for idx in range(len(df_display)):
                            for col in df_display.columns:
                                if col != "位置":  # 跳过位置列
                                    old_val = df_display.iloc[idx][col]
                                    new_val = edited_df.iloc[idx][col]
                                    if old_val != new_val:
                                        changes_found = True
                                        st.markdown(f"**{df_display.iloc[idx]['位置']} - {col}:**")
                                        st.markdown(f"- 原值: `{old_val}`")
                                        st.markdown(f"- 新值: `{new_val}`")
                                        st.markdown("---")

                        if not changes_found:
                            st.info("未检测到有效修改")

                    # 保存修改按钮
                    col_save1, col_save2 = st.columns([1, 3])

                    with col_save1:
                        if st.button("💾 保存修改到Excel", key=f"save_edits_{selected_file}", type="primary"):
                            try:
                                # 加载Excel文件
                                wb = openpyxl.load_workbook(selected_file)

                                # 遍历所有修改
                                changes_count = 0
                                for idx in range(len(edited_df)):
                                    sheet_name = row_identifiers[idx]['sheet']
                                    row_num = row_identifiers[idx]['row']
                                    ws = wb[sheet_name]

                                    # 检查每一列的修改
                                    for col in edited_df.columns:
                                        if col != "位置":
                                            col_idx = int(col.replace("列", ""))
                                            new_val = edited_df.iloc[idx][col]
                                            old_val = df_display.iloc[idx][col]

                                            if new_val != old_val:
                                                # 写入新值到Excel
                                                ws.cell(row=row_num, column=col_idx, value=new_val)
                                                changes_count += 1

                                # 保存文件
                                wb.save(selected_file)
                                wb.close()

                                st.success(f"✅ 成功保存 {changes_count} 处修改！")

                                # 提示重新搜索
                                st.info("💡 建议重新搜索以查看最新内容")

                            except Exception as e:
                                st.error(f"❌ 保存失败: {str(e)}")

                    with col_save2:
                        if st.button("🔄 撤销修改", key=f"reset_edits_{selected_file}"):
                            st.rerun()

            if len(matches) > 50:
                st.info(f"仅显示前 50 个匹配项，共有 {len(matches)} 个匹配项")

        # 替换功能
        st.header("🔄 批量替换功能")

        col1, col2 = st.columns(2)

        with col1:
            replace_term = st.text_input(
                "替换为:",
                value=st.session_state.replace_term,
                placeholder="请输入替换后的词语",
                help="将搜索到的词语替换为此词语"
            )
            st.session_state.replace_term = replace_term

        with col2:
            backup_files = st.checkbox(
                "创建备份文件",
                value=True,
                help="替换前自动创建备份文件"
            )

        # 替换预览
        if search_term and replace_term:
            st.subheader("🔍 替换预览")

            # 显示替换前后对比示例
            example_before = f"这是包含 {search_term} 的示例文本"
            example_after = example_before.replace(search_term, f"**{replace_term}**")

            col_before, col_arrow, col_after = st.columns([1, 0.1, 1])

            with col_before:
                st.text_area("替换前:", value=example_before, height=60, disabled=True)

            with col_arrow:
                st.markdown("<br><h2>→</h2>", unsafe_allow_html=True)

            with col_after:
                st.text_area("替换后:", value=example_after, height=60, disabled=True)

        # 选择性替换功能
        st.subheader("📋 选择要替换的项目")

        # 初始化选择状态
        if 'selected_replacements' not in st.session_state:
            st.session_state.selected_replacements = {}

        # 全选/全不选按钮
        col_select1, col_select2, col_select3 = st.columns([1, 1, 3])
        with col_select1:
            if st.button("✅ 全选", use_container_width=True):
                for file_path in search_tool.search_results.keys():
                    st.session_state.selected_replacements[file_path] = {
                        'selected': True,
                        'rows': 'all'
                    }
                st.rerun()

        with col_select2:
            if st.button("❌ 全不选", use_container_width=True):
                st.session_state.selected_replacements = {}
                st.rerun()

        # 为每个文件创建选择界面
        for file_path, file_data in search_tool.search_results.items():
            with st.expander(f"📄 {Path(file_path).name} ({file_data['match_count']} 处匹配)", expanded=False):
                # 文件级别的选择
                file_key = f"file_{file_path}"

                # 初始化该文件的选择状态
                if file_path not in st.session_state.selected_replacements:
                    st.session_state.selected_replacements[file_path] = {
                        'selected': False,
                        'rows': 'all',
                        'selected_rows': set()
                    }

                col_file1, col_file2 = st.columns([1, 3])

                with col_file1:
                    file_selected = st.checkbox(
                        "选择此文件",
                        value=st.session_state.selected_replacements[file_path]['selected'],
                        key=f"cb_{file_key}"
                    )
                    st.session_state.selected_replacements[file_path]['selected'] = file_selected

                with col_file2:
                    if file_selected:
                        replace_mode = st.radio(
                            "替换模式:",
                            options=['all', 'selected'],
                            format_func=lambda x: "替换所有匹配项" if x == 'all' else "选择特定行",
                            key=f"mode_{file_key}",
                            horizontal=True
                        )
                        st.session_state.selected_replacements[file_path]['rows'] = replace_mode

                        # 如果选择了特定行模式，显示行选择界面
                        if replace_mode == 'selected':
                            st.markdown("**选择要替换的行:**")

                            matches = file_data['matches']
                            # 按工作表分组
                            sheets_data = {}
                            for match in matches:
                                sheet_name = match['sheet_name']
                                if sheet_name not in sheets_data:
                                    sheets_data[sheet_name] = []
                                sheets_data[sheet_name].append(match)

                            # 为每个工作表显示行选择
                            for sheet_name, sheet_matches in sheets_data.items():
                                st.markdown(f"*工作表: {sheet_name}*")

                                # 获取唯一的行号
                                unique_rows = sorted(set(m['row'] for m in sheet_matches))

                                cols = st.columns(5)
                                for idx, row_num in enumerate(unique_rows):
                                    with cols[idx % 5]:
                                        row_key = f"{file_path}_{sheet_name}_{row_num}"
                                        row_selected = st.checkbox(
                                            f"行 {row_num}",
                                            value=row_key in st.session_state.selected_replacements[file_path]['selected_rows'],
                                            key=f"row_{row_key}"
                                        )

                                        if row_selected:
                                            st.session_state.selected_replacements[file_path]['selected_rows'].add(row_key)
                                        elif row_key in st.session_state.selected_replacements[file_path]['selected_rows']:
                                            st.session_state.selected_replacements[file_path]['selected_rows'].remove(row_key)

        # 显示替换统计
        st.subheader("📊 替换统计")
        selected_files_count = sum(1 for f in st.session_state.selected_replacements.values() if f['selected'])
        total_selected_matches = 0

        for file_path, selection in st.session_state.selected_replacements.items():
            if selection['selected']:
                if selection['rows'] == 'all':
                    total_selected_matches += search_tool.search_results[file_path]['match_count']
                else:
                    # 计算选中的行数
                    total_selected_matches += len(selection['selected_rows'])

        col_stat1, col_stat2 = st.columns(2)
        with col_stat1:
            st.metric("选中的文件数", selected_files_count)
        with col_stat2:
            st.metric("预计替换项数", total_selected_matches)

        # 执行替换按钮
        if st.button("🔄 执行批量替换", key="replace_btn", type="primary", use_container_width=True):
            if not replace_term:
                st.error("❌ 请输入替换词语")
                return

            if selected_files_count == 0:
                st.error("❌ 请至少选择一个文件进行替换")
                return

            # 如果还没有确认,显示确认复选框
            if not st.session_state.replace_confirmed:
                st.warning(f"⚠️ 此操作将在 {selected_files_count} 个文件中执行约 {total_selected_matches} 处替换！")
                st.session_state.show_confirm_checkbox = True

        # 显示确认复选框
        if st.session_state.show_confirm_checkbox:
            confirm_replace = st.checkbox("我确认要执行批量替换操作")

            if confirm_replace:
                st.session_state.replace_confirmed = True
                st.session_state.show_confirm_checkbox = False
                st.rerun()

        # 如果已经确认,执行替换操作
        if st.session_state.replace_confirmed:
            # 执行选择性替换
            with st.spinner("正在执行替换操作..."):
                replaced_files, total_replacements = selective_replace(
                    search_tool,
                    search_term,
                    replace_term,
                    st.session_state.selected_replacements,
                    backup_files,
                    case_sensitive
                )

            if replaced_files > 0:
                st.success(f"✅ 替换完成！在 {replaced_files} 个文件中完成了 {total_replacements} 处替换")

                # 重置状态
                st.session_state.replace_confirmed = False
                st.session_state.show_confirm_checkbox = False
                st.session_state.selected_replacements = {}

                # 清空搜索结果,提示重新搜索
                search_tool.search_results = {}
                st.info("💡 替换完成，请重新搜索以查看更新后的内容")

                # 清空搜索和替换词
                st.session_state.search_term = ""
                st.session_state.replace_term = ""
            else:
                st.error("❌ 替换操作失败")
                st.session_state.replace_confirmed = False
                st.session_state.show_confirm_checkbox = False

    # 使用说明
    with st.expander("📖 使用说明"):
        st.markdown("""
        ## 使用说明

        ### 基本流程：
        1. **设置文件夹路径** - 在侧边栏输入包含Excel文件的文件夹路径
        2. **输入搜索词语** - 在主界面输入要搜索的词语
        3. **设置搜索选项** - 选择是否大小写敏感、全词匹配
        4. **开始搜索** - 点击"开始搜索"按钮
        5. **查看和编辑结果** - 浏览搜索结果，直接在表格中修改内容
        6. **保存单个文件修改** - 在可编辑表格中修改后点击"保存修改到Excel"
        7. **快速操作** - 使用"打开文件夹"或"打开Excel"按钮快速访问文件
        8. **批量替换** - 使用批量替换功能对多个文件执行统一替换

        ### 功能特点：
        - ✏️ **直接编辑** - 在搜索结果表格中直接修改单元格内容
        - 💾 **即时保存** - 修改后立即保存到Excel文件
        - 🔍 **多线程批量搜索** - 自动使用多线程加速搜索，充分利用CPU资源
        - 📊 **原表格展示** - 以原始表格形式显示匹配行的完整数据
        - 📂 **快速访问** - 一键打开文件所在文件夹或直接打开Excel文件
        - 🎯 **选择性替换** - 可以选择特定文件、特定行进行批量替换
        - ⚙️ **灵活选项** - 支持大小写敏感和全词匹配
        - 💾 **自动备份** - 批量替换前可自动创建备份文件
        - 📁 **多格式支持** - 支持 .xlsx, .xls, .xlsm, .xlsb 格式
        - ⚡ **实时进度** - 显示搜索和替换的实时进度

        ### 两种修改方式：
        1. **直接编辑（推荐用于少量精确修改）**
           - 在搜索结果表格中直接修改单元格
           - 点击"保存修改到Excel"即时保存
           - 适合修改个别单元格内容

        2. **批量替换（推荐用于大量统一替换）**
           - 选择要替换的文件和行
           - 执行统一的查找替换操作
           - 可创建备份文件
           - 适合大规模统一修改

        ### 注意事项：
        - 直接编辑会立即保存到文件，请谨慎操作
        - 批量替换操作会修改原文件，建议先创建备份
        - 建议先在小范围测试
        - 支持正则表达式语法（在搜索词语中）
        - 大型文件可能需要较长时间处理
        - 修改保存后建议重新搜索查看最新内容
        """)
